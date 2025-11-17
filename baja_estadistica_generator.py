import mysql.connector
import psycopg2
import psycopg2.extras
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import re
import random
import requests
import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup
import warnings
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def extraer_criterio_individual(criteria_elem):
    """Extrae información de un elemento AwardingCriteria individual - versión mejorada"""
    criterio_info = {}

    # Buscar elementos específicos dentro del criterio
    for child in criteria_elem:
        tag_name = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        tag_lower = tag_name.lower()

        # Descripción del criterio - buscar múltiples variaciones
        if (any(term in tag_lower for term in ['description', 'name', 'titulo', 'desc']) and child.text):
            criterio_info['descripcion'] = child.text.strip()

        # Peso numérico - buscar múltiples variaciones
        elif (any(term in tag_lower for term in ['weight', 'peso', 'points', 'puntos', 'percentage']) and child.text):
            peso_num = child.text.strip()
            criterio_info['peso'] = f"{peso_num}%" if not peso_num.endswith('%') else peso_num

        # Subtipo de criterio
        elif 'subtype' in tag_lower or 'tipo' in tag_lower:
            subtipo = child.get('name') or child.text
            if subtipo:
                if 'descripcion' in criterio_info:
                    criterio_info['descripcion'] += f" ({subtipo.strip()})"
                else:
                    criterio_info['descripcion'] = subtipo.strip()

    # Si no encontramos descripción específica, usar el texto directo del elemento
    if not criterio_info.get('descripcion') and criteria_elem.text:
        text = criteria_elem.text.strip()
        if len(text) > 5 and len(text) < 200:
            criterio_info['descripcion'] = text

    # Buscar información en atributos del elemento principal
    for attr_name, attr_value in criteria_elem.attrib.items():
        attr_lower = attr_name.lower()
        if 'name' in attr_lower or 'description' in attr_lower:
            if not criterio_info.get('descripcion'):
                criterio_info['descripcion'] = attr_value
        elif 'weight' in attr_lower or 'peso' in attr_lower:
            if not criterio_info.get('peso'):
                criterio_info['peso'] = f"{attr_value}%"

    return criterio_info if criterio_info else None

class BajaEstadisticaGenerator:
    def __init__(self):
        self.connection = None

        # Plantillas de saludo variadas
        self.saludos = [
            "Buenos días,",
            "Estimados señores,",
            "Buenas tardes,",
            "Estimado equipo,",
            "Muy buenos días,"
        ]

        # Plantillas de despedida variadas
        self.despedidas = [
            "Un cordial saludo",
            "Saludos cordiales",
            "Atentamente",
            "Un saludo",
            "Cordialmente"
        ]

        # Frases de introducción variadas
        self.introducciones = [
            "En la selección de expedientes, priorizaremos el criterio de precio, otorgándole {puntos_precio} puntos, con un margen de {puntos_tecnico} puntos para evaluaciones técnicas.",
            "Para la evaluación de propuestas, se asignará {puntos_precio} puntos al aspecto económico y {puntos_tecnico} puntos a la valoración técnica.",
            "En el proceso de selección, el criterio económico tendrá un peso de {puntos_precio} puntos, reservando {puntos_tecnico} puntos para aspectos técnicos.",
            "La puntuación se distribuirá otorgando {puntos_precio} puntos al precio y {puntos_tecnico} puntos a criterios técnicos."
        ]

    def connect_to_database(self):
        """Conectar a la base de datos PostgreSQL oclemconcursos"""
        try:
            self.connection = psycopg2.connect(
                host=st.secrets["postgres"]["host"],
                database=st.secrets["postgres"]["database"],
                user=st.secrets["postgres"]["user"],
                password=st.secrets["postgres"]["password"],
                port=st.secrets["postgres"]["port"]
            )
            return True
        except Exception as e:
            st.error(f"Error conectando a la base de datos: {e}")
            return False

    def extract_json_data(self, json_data, numero_lote=None):
        """Extraer datos de un JSON de licitación

        Args:
            json_data: Datos JSON (string o dict)
            numero_lote: Número de lote específico a analizar (opcional)
        """
        try:
            # Si es un string, parsearlo como JSON
            if isinstance(json_data, str):
                data = json.loads(json_data)
            elif isinstance(json_data, dict):
                data = json_data
            else:
                st.error("Formato JSON no válido")
                return None

            # Si se especifica un lote, filtrar el JSON para trabajar solo con ese lote
            if numero_lote:
                lote_encontrado = False
                # Buscar arrays de lotes en el JSON
                lotes_keys = ['lotes', 'lots', 'lote', 'lot', 'items', 'partidas']
                for key in lotes_keys:
                    lotes_array = self._find_json_value(data, key)
                    if lotes_array and isinstance(lotes_array, list):
                        # Buscar el lote específico
                        for lote in lotes_array:
                            if isinstance(lote, dict):
                                # Buscar el ID o número del lote
                                lote_id = lote.get('id', lote.get('numero', lote.get('lote', lote.get('lot', ''))))
                                if str(lote_id) == numero_lote or str(lote_id).endswith(numero_lote):
                                    data = lote  # Usar solo este lote
                                    lote_encontrado = True
                                    st.info(f"✅ Filtrando análisis para el Lote {numero_lote}")
                                    break
                        if lote_encontrado:
                            break

                if not lote_encontrado:
                    st.warning(f"⚠️ No se encontró el lote {numero_lote}. Analizando todo el contrato.")

            # Estructura de datos similar a extract_xml_data
            datos = {
                'titulo': '',
                'organismo': '',
                'presupuesto': 0,
                'cpv': '',
                'tipo_procedimiento': '',
                'criterios_adjudicacion': [],
                'descripcion': '',
                'ubicacion': '',
                'debug_info': {'source': 'JSON', 'keys': list(data.keys()) if isinstance(data, dict) else []}
            }

            # Buscar campos comunes en JSON
            # Título - buscar variaciones comunes
            titulo_keys = ['titulo', 'title', 'name', 'objeto', 'description', 'asunto', 'denominacion']
            for key in titulo_keys:
                if self._find_json_value(data, key):
                    datos['titulo'] = str(self._find_json_value(data, key)).strip()
                    break

            # Organismo
            organismo_keys = ['organismo', 'entidad', 'organo', 'buyer', 'contracting_authority', 'contratante', 'administracion']
            for key in organismo_keys:
                if self._find_json_value(data, key):
                    datos['organismo'] = str(self._find_json_value(data, key)).strip()
                    break

            # Presupuesto - buscar variaciones
            presupuesto_keys = ['presupuesto', 'precio', 'valor', 'importe', 'amount', 'budget', 'value', 'estimatedValue']
            for key in presupuesto_keys:
                value = self._find_json_value(data, key)
                if value:
                    try:
                        # Limpiar y convertir a número
                        if isinstance(value, (int, float)):
                            datos['presupuesto'] = float(value)
                            break
                        elif isinstance(value, str):
                            # Extraer número del string
                            clean_value = re.sub(r'[^\d.,]', '', value.replace(',', '.'))
                            if clean_value:
                                datos['presupuesto'] = float(clean_value)
                                break
                    except:
                        continue

            # CPV
            cpv_keys = ['cpv', 'codigo', 'classification', 'classificationCode', 'cpv_code']
            for key in cpv_keys:
                value = self._find_json_value(data, key)
                if value:
                    # Si es una lista, tomar todos los códigos
                    if isinstance(value, list):
                        cpv_codes = [str(v) for v in value if str(v).isdigit() and len(str(v)) >= 8]
                        if cpv_codes:
                            datos['cpv'] = ', '.join(cpv_codes)
                            break
                    else:
                        # Si es string o número, verificar si es un CPV válido
                        cpv_str = str(value).strip()
                        if cpv_str.isdigit() and len(cpv_str) >= 8:
                            datos['cpv'] = cpv_str
                            break

            # Ubicación
            ubicacion_keys = ['ubicacion', 'lugar', 'provincia', 'location', 'place', 'region', 'city', 'address']
            for key in ubicacion_keys:
                if self._find_json_value(data, key):
                    datos['ubicacion'] = str(self._find_json_value(data, key)).strip()
                    break

            # Tipo de procedimiento
            tipo_keys = ['tipo', 'procedimiento', 'procedure', 'type', 'procurementMethod']
            for key in tipo_keys:
                if self._find_json_value(data, key):
                    datos['tipo_procedimiento'] = str(self._find_json_value(data, key)).strip()
                    break

            # Criterios de adjudicación - buscar en diferentes estructuras
            criterios = []
            criterios_keys = ['criterios', 'criteria', 'awardingCriteria', 'evaluationCriteria']

            for key in criterios_keys:
                criterios_data = self._find_json_value(data, key)
                if criterios_data:
                    if isinstance(criterios_data, list):
                        for criterio in criterios_data:
                            if isinstance(criterio, dict):
                                criterio_info = {}
                                # Buscar descripción
                                desc_keys = ['descripcion', 'description', 'name', 'titulo']
                                for desc_key in desc_keys:
                                    if desc_key in criterio and criterio[desc_key]:
                                        criterio_info['descripcion'] = str(criterio[desc_key]).strip()
                                        break

                                # Buscar peso
                                peso_keys = ['peso', 'weight', 'puntos', 'points', 'percentage']
                                for peso_key in peso_keys:
                                    if peso_key in criterio and criterio[peso_key]:
                                        peso_val = criterio[peso_key]
                                        if isinstance(peso_val, (int, float)):
                                            criterio_info['peso'] = f"{peso_val}%"
                                        else:
                                            criterio_info['peso'] = str(peso_val)
                                        break

                                if criterio_info:
                                    criterios.append(criterio_info)
                            elif isinstance(criterio, str):
                                criterios.append({'descripcion': criterio})
                    elif isinstance(criterios_data, dict):
                        # Si es un diccionario, tratar como un solo criterio
                        criterio_info = {}
                        if 'descripcion' in criterios_data or 'description' in criterios_data:
                            desc = criterios_data.get('descripcion') or criterios_data.get('description')
                            if desc:
                                criterio_info['descripcion'] = str(desc).strip()

                        if 'peso' in criterios_data or 'weight' in criterios_data:
                            peso = criterios_data.get('peso') or criterios_data.get('weight')
                            if peso:
                                criterio_info['peso'] = f"{peso}%" if isinstance(peso, (int, float)) else str(peso)

                        if criterio_info:
                            criterios.append(criterio_info)
                    break

            # Si no encontramos criterios estructurados, buscar criterios por defecto
            if not criterios:
                st.info("🔍 No se encontraron criterios específicos en JSON. Generando criterios por defecto...")
                criterios = [
                    {'descripcion': 'Oferta económica', 'peso': '80 puntos'},
                    {'descripcion': 'Criterios técnicos', 'peso': '20 puntos'}
                ]

            datos['criterios_adjudicacion'] = criterios

            # Descripción
            descripcion_keys = ['descripcion', 'description', 'details', 'summary']
            for key in descripcion_keys:
                value = self._find_json_value(data, key)
                if value and len(str(value).strip()) > 20:
                    datos['descripcion'] = str(value).strip()
                    break

            return datos

        except Exception as e:
            st.error(f"Error procesando JSON: {e}")
            return None

    def _find_json_value(self, data, key_to_find):
        """Buscar una clave en un JSON de manera recursiva (case-insensitive)"""
        if isinstance(data, dict):
            # Buscar directamente (case-insensitive)
            for key, value in data.items():
                if key.lower() == key_to_find.lower() and value:
                    return value

            # Buscar recursivamente en valores anidados
            for key, value in data.items():
                if isinstance(value, (dict, list)):
                    result = self._find_json_value(value, key_to_find)
                    if result:
                        return result
        elif isinstance(data, list):
            for item in data:
                result = self._find_json_value(item, key_to_find)
                if result:
                    return result

        return None

    def extract_xml_data(self, xml_url, numero_lote=None):
        """Extraer datos de un XML de contratación del estado

        Args:
            xml_url: URL del XML
            numero_lote: Número de lote específico a analizar (opcional)
        """
        try:
            # Descargar el XML
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(xml_url, headers=headers, timeout=30)
            response.raise_for_status()

            # Parsear el XML
            root = ET.fromstring(response.content)

            # Si se especifica un lote, filtrar el XML para trabajar solo con ese lote
            if numero_lote:
                lote_encontrado = False
                # Buscar elementos de lote en el XML
                for elem in root.iter():
                    if 'lot' in elem.tag.lower() or 'lote' in elem.tag.lower():
                        # Buscar el ID del lote
                        lote_id = None
                        for child in elem.iter():
                            if 'id' in child.tag.lower() and child.text:
                                lote_id = child.text.strip()
                                break

                        # Si el ID coincide con el número de lote buscado
                        if lote_id and (lote_id == numero_lote or lote_id == f"Lote {numero_lote}" or lote_id.endswith(numero_lote)):
                            root = elem  # Usar este elemento como raíz
                            lote_encontrado = True
                            st.info(f"✅ Filtrando análisis para el Lote {numero_lote}")
                            break

                if not lote_encontrado:
                    st.warning(f"⚠️ No se encontró el lote {numero_lote}. Analizando todo el contrato.")

            root_original = root  # Guardar referencia al root filtrado

            # Información de debug (opcional para ver la estructura)
            debug_info = {
                'total_elements': len(list(root.iter())),
                'root_tag': root.tag,
                'namespaces': set(),
                'unique_tags': set()
            }

            for elem in root.iter():
                debug_info['unique_tags'].add(elem.tag.split('}')[-1])  # Sin namespace
                if '}' in elem.tag:
                    debug_info['namespaces'].add(elem.tag.split('}')[0] + '}')

            # Extraer datos básicos
            datos = {
                'titulo': '',
                'organismo': '',
                'presupuesto': 0,
                'cpv': '',
                'tipo_procedimiento': '',
                'criterios_adjudicacion': [],
                'descripcion': '',
                'ubicacion': '',
                'debug_info': debug_info
            }

            # Buscar elementos específicos de XMLs de contratación del estado
            # Título - buscar específicamente en ProcurementProject/Name
            # Primero buscar la estructura específica de contratación del estado
            for elem in root.iter():
                if 'procurementproject' in elem.tag.lower():
                    for child in elem:
                        if 'name' in child.tag.lower() and child.text:
                            datos['titulo'] = child.text.strip()
                            break
                    if datos['titulo']:
                        break

            # Si no se encuentra, buscar de forma general
            if not datos['titulo']:
                for titulo in root.iter():
                    if any(term in titulo.tag.lower() for term in ['name', 'titulo', 'title', 'objeto', 'description']):
                        if titulo.text and len(titulo.text.strip()) > 10:
                            datos['titulo'] = titulo.text.strip()
                            break

            # Organismo - buscar específicamente en PartyName/Name
            # Primero buscar la estructura específica
            for elem in root.iter():
                if 'partyname' in elem.tag.lower():
                    for child in elem:
                        if 'name' in child.tag.lower() and child.text:
                            datos['organismo'] = child.text.strip()
                            break
                    if datos['organismo']:
                        break

            # Si no se encuentra, buscar de forma general
            if not datos['organismo']:
                for org in root.iter():
                    if any(term in org.tag.lower() for term in ['contractingparty', 'organismo', 'organo', 'entity', 'contracting', 'buyername']):
                        if org.text and len(org.text.strip()) > 3:
                            datos['organismo'] = org.text.strip()
                            break

            # Presupuesto - buscar elementos específicos de contratación del estado
            presupuesto_encontrado = False

            # Buscar específicamente TaxExclusiveAmount con currencyID="EUR"
            for elem in root.iter():
                if ('taxexclusiveamount' in elem.tag.lower() or elem.tag.endswith('TaxExclusiveAmount')):
                    # Verificar si tiene el atributo currencyID (opcional pero común)
                    currency_id = elem.get('currencyID', '')
                    if elem.text:
                        try:
                            # Limpiar y extraer número - puede ser 125000 directo
                            valor_texto = elem.text.strip()
                            # Eliminar caracteres no numéricos excepto puntos y comas
                            valor_limpio = valor_texto.replace(',', '').replace(' ', '')

                            # Convertir a float
                            valor_numerico = float(valor_limpio)
                            datos['presupuesto'] = valor_numerico
                            presupuesto_encontrado = True

                            # Debug info
                            datos['presupuesto_debug'] = {
                                'tag': elem.tag,
                                'valor_original': valor_texto,
                                'currency': currency_id,
                                'valor_procesado': valor_numerico
                            }
                            break
                        except Exception as e:
                            # Debug de errores
                            datos['presupuesto_error'] = f"Error procesando {elem.text}: {str(e)}"
                            continue

            # Si no encuentra TaxExclusiveAmount, buscar EstimatedOverallContractAmount
            if not presupuesto_encontrado:
                for elem in root.iter():
                    if 'estimatedoverallcontractamount' in elem.tag.lower() or elem.tag.endswith('EstimatedOverallContractAmount'):
                        if elem.text:
                            try:
                                valor_limpio = elem.text.strip().replace(',', '').replace(' ', '')
                                datos['presupuesto'] = float(valor_limpio)
                                presupuesto_encontrado = True
                                break
                            except:
                                continue

            # Búsqueda general si no se encuentra en elementos específicos
            if not presupuesto_encontrado:
                for precio in root.iter():
                    if any(term in precio.tag.lower() for term in ['importe', 'valor', 'value', 'amount', 'precio']):
                        if precio.text:
                            try:
                                texto_limpio = precio.text.strip().replace(',', '').replace(' ', '')
                                numeros = re.findall(r'[\d]+\.?\d*', texto_limpio)
                                if numeros:
                                    datos['presupuesto'] = float(numeros[0])
                                    break
                            except:
                                continue

            # CPV - códigos de clasificación (puede haber varios)
            cpv_codes = []

            # Buscar específicamente en RequiredCommodityClassification/ItemClassificationCode
            for elem in root.iter():
                if 'requiredcommodityclassification' in elem.tag.lower():
                    for child in elem:
                        if 'itemclassificationcode' in child.tag.lower() and child.text:
                            cpv_text = child.text.strip()
                            if len(cpv_text) >= 8 and cpv_text.isdigit():
                                cpv_codes.append(cpv_text)
                        # También buscar atributos con name para descripción
                        if 'name' in child.attrib and child.text:
                            cpv_text = child.text.strip()
                            if len(cpv_text) >= 8 and cpv_text.isdigit():
                                cpv_codes.append(cpv_text)

            # Si no se encuentra en la estructura específica, buscar de forma general
            if not cpv_codes:
                for cpv in root.iter():
                    if any(term in cpv.tag.lower() for term in ['cpv', 'classificationcode', 'codigo']):
                        if cpv.text and len(cpv.text.strip()) >= 8:
                            cpv_nums = re.findall(r'\d+', cpv.text.strip())
                            if cpv_nums and len(cpv_nums[0]) >= 8:
                                cpv_codes.append(cpv_nums[0])
                    # También buscar en atributos
                    for attr_name, attr_value in cpv.attrib.items():
                        if 'code' in attr_name.lower() and len(attr_value) >= 8:
                            cpv_nums = re.findall(r'\d+', attr_value)
                            if cpv_nums and len(cpv_nums[0]) >= 8:
                                cpv_codes.append(cpv_nums[0])

            # Guardar todos los CPVs encontrados, separados por comas
            if cpv_codes:
                datos['cpv'] = ', '.join(list(set(cpv_codes)))  # Eliminar duplicados

            # Ubicación - buscar específicamente en RealizedLocation/CountrySubentity
            for elem in root.iter():
                if 'realizedlocation' in elem.tag.lower():
                    for child in elem:
                        if 'countrysubentity' in child.tag.lower() and child.text:
                            datos['ubicacion'] = child.text.strip()
                            break
                    if datos['ubicacion']:
                        break

            # Si no se encuentra, buscar de forma general
            if not datos['ubicacion']:
                for ubicacion in root.iter():
                    if any(term in ubicacion.tag.lower() for term in ['location', 'place', 'lugar', 'provincia', 'city', 'countrysubentity']):
                        if ubicacion.text and len(ubicacion.text.strip()) > 2:
                            datos['ubicacion'] = ubicacion.text.strip()
                            break

            # Procedimiento
            for proc in root.iter():
                if any(term in proc.tag.lower() for term in ['procedimiento', 'procedure', 'tipo']):
                    if proc.text:
                        datos['tipo_procedimiento'] = proc.text.strip()
                        break

            # Criterios de adjudicación - búsqueda mejorada y más robusta
            criterios = []
            criterios_encontrados = False

            with st.expander("🔍 Ver debug completo de búsqueda de criterios", expanded=False):
                st.write("🔍 **Debug - Buscando criterios de adjudicación...**")

                # DEBUG COMPLETO: Mostrar TODA la estructura del XML
                st.write("🔍 **DEBUG COMPLETO: Analizando estructura del XML**")

                # Mostrar todos los tags únicos en el XML
                all_tags = set()
                all_text_elements = []

                for elem in root.iter():
                    tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                    all_tags.add(tag_name)

                    # Recopilar elementos con texto relevante
                    if elem.text and len(elem.text.strip()) > 10:
                        text = elem.text.strip()
                        if any(keyword in text.lower() for keyword in [
                            'criterio', 'criteria', 'económico', 'técnico', 'calidad', 'precio',
                            'economic', 'technical', 'quality', 'price', 'evaluation', 'peso',
                            'weight', 'puntos', 'points', '%'
                        ]):
                            all_text_elements.append({
                                'tag': tag_name,
                                'full_tag': elem.tag,
                                'text': text[:100] + ('...' if len(text) > 100 else ''),
                                'text_length': len(text)
                            })

                st.write(f"📋 **Tags únicos encontrados en el XML ({len(all_tags)}):**")
                sorted_tags = sorted(all_tags)
                # Mostrar tags en columnas para mejor visualización
                col1, col2, col3 = st.columns(3)
                for i, tag in enumerate(sorted_tags):
                    if i % 3 == 0:
                        col1.write(f"• {tag}")
                    elif i % 3 == 1:
                        col2.write(f"• {tag}")
                    else:
                        col3.write(f"• {tag}")

                st.write(f"📝 **Elementos con texto potencialmente relevante ({len(all_text_elements)}):**")
                for elem_info in all_text_elements[:10]:  # Mostrar solo los primeros 10
                    st.write(f"• **{elem_info['tag']}**: {elem_info['text']}")

                if len(all_text_elements) > 10:
                    st.write(f"... y {len(all_text_elements) - 10} elementos más")

                # Buscar específicamente palabras clave relacionadas con criterios
                criterios_keywords = [
                    'award', 'criteria', 'criterion', 'evaluation', 'scoring',
                    'criterio', 'evaluacion', 'puntuacion', 'adjudicacion'
                ]

                tags_with_keywords = []
                for tag in sorted_tags:
                    if any(keyword in tag.lower() for keyword in criterios_keywords):
                        tags_with_keywords.append(tag)

                if tags_with_keywords:
                    st.write(f"🎯 **Tags que contienen palabras clave de criterios:**")
                    for tag in tags_with_keywords:
                        st.write(f"• {tag}")
                else:
                    st.warning("⚠️ No se encontraron tags con palabras clave obvias de criterios")

                # Estrategia 1: Búsqueda flexible de criterios de adjudicación
                st.write("🎯 **Estrategia 1: Búsqueda flexible de criterios**")

                # Términos que pueden indicar criterios de adjudicación
                criterios_tags = [
                    'AwardingCriteria', 'AwardingTerms', 'EvaluationCriteria',
                    'Criterion', 'Criteria', 'CriterioAdjudicacion',
                    'CriteriosEvaluacion', 'Subcriteria'
                ]

                # Buscar elementos con nombres relacionados con criterios
                for elem in root.iter():
                    tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag

                    # Verificar si el tag contiene alguno de los términos de criterios
                    if any(term in tag_name for term in criterios_tags):
                        st.write(f"📍 Encontrado elemento relevante: {elem.tag}")

                        # Si es un contenedor de criterios (Terms), buscar sus hijos
                        if 'Terms' in tag_name or 'Container' in tag_name:
                            for criteria_elem in elem:
                                if any(term in criteria_elem.tag for term in criterios_tags):
                                    criterio_info = extraer_criterio_individual(criteria_elem)
                                    if criterio_info:
                                        criterios.append(criterio_info)
                                        criterios_encontrados = True
                                        st.write(f"  ✅ Criterio extraído desde contenedor: {criterio_info}")

                        # Si es directamente un criterio
                        else:
                            criterio_info = extraer_criterio_individual(elem)
                            if criterio_info:
                                criterios.append(criterio_info)
                                criterios_encontrados = True
                                st.write(f"  ✅ Criterio extraído directo: {criterio_info}")

                # Estrategia 1.5: Buscar cualquier elemento que contenga información de criterios
                if not criterios_encontrados:
                    st.write("🔍 **Estrategia 1.5: Búsqueda por contenido de criterios**")
                    for elem in root.iter():
                        # Buscar elementos que contengan palabras clave en su texto
                        if elem.text and len(elem.text.strip()) > 10:
                            text_lower = elem.text.lower()
                            if any(keyword in text_lower for keyword in [
                                'criterio', 'criteria', 'evaluación', 'evaluation',
                                'puntuación', 'scoring', 'peso', 'weight', 'puntos', 'points'
                            ]):
                                # Verificar que parece realmente un criterio
                                if any(indicator in text_lower for indicator in [
                                    'económico', 'técnico', 'calidad', 'precio', 'plazo',
                                    'economic', 'technical', 'quality', 'price', 'deadline',
                                    '%', 'puntos', 'points'
                                ]):
                                    criterio_info = {'descripcion': elem.text.strip()}
                                    criterios.append(criterio_info)
                                    criterios_encontrados = True
                                    st.write(f"  📝 Criterio encontrado por contenido: {elem.text.strip()[:60]}...")

                # Estrategia FINAL: Extraer CUALQUIER texto que parezca criterio (muy agresiva)
                if not criterios_encontrados:
                    st.write("🚨 **Estrategia FINAL: Extracción agresiva de cualquier criterio potencial**")

                    # Buscar en TODOS los elementos de texto
                    for elem in root.iter():
                        if elem.text and len(elem.text.strip()) > 15:
                            text = elem.text.strip()
                            text_lower = text.lower()

                            # Criterios muy flexibles para detectar cualquier cosa que pueda ser un criterio
                            if any(indicator in text_lower for indicator in [
                                'económico', 'técnico', 'calidad', 'precio', 'plazo', 'coste',
                                'economic', 'technical', 'quality', 'price', 'cost', 'delivery',
                                'experiencia', 'experience', 'capacidad', 'capacity',
                                'oferta', 'offer', 'propuesta', 'proposal',
                                'valoración', 'evaluation', 'puntuación', 'scoring',
                                '%', 'punto', 'point', 'peso', 'weight'
                            ]):
                                # Verificar que no sea demasiado largo (probablemente no es un criterio)
                                if len(text) < 500:
                                    criterio_info = {
                                        'descripcion': text,
                                        'fuente': 'extraccion_agresiva',
                                        'tag': elem.tag
                                    }
                                    criterios.append(criterio_info)
                                    criterios_encontrados = True
                                    st.write(f"  🎯 Criterio potencial extraído: {text[:80]}...")

                    # Si aún no encuentra nada, usar criterios por defecto
                    if not criterios_encontrados:
                        st.warning("⚠️ No se pudieron extraer criterios específicos del XML")
                        st.info("📋 Usando criterios estándar de contratación pública:")
                        criterios = [
                            {'descripcion': 'Criterio económico - Oferta económica', 'peso': '60%'},
                            {'descripcion': 'Criterio técnico - Aspectos técnicos', 'peso': '40%'}
                        ]
                        criterios_encontrados = True

                # Estrategia 2: Búsqueda más amplia por texto que contenga palabras clave
                if not criterios_encontrados:
                    st.write("🔍 **Búsqueda amplia de criterios...**")

                    terminos_busqueda = [
                        'criterio', 'criteria', 'award', 'evaluation', 'subcriteria',
                        'weighting', 'peso', 'valoracion', 'puntuacion', 'scoring',
                        'economic', 'technical', 'price', 'quality', 'técnico', 'económico'
                    ]

                    for elem in root.iter():
                        # Buscar por nombre de tag
                        if any(term in elem.tag.lower() for term in terminos_busqueda):
                            if elem.text and len(elem.text.strip()) > 5:
                                criterio_desc = elem.text.strip()
                                criterios.append({'descripcion': criterio_desc})
                                criterios_encontrados = True
                                st.write(f"  📝 Criterio encontrado por tag: {criterio_desc[:50]}...")

                        # Buscar por contenido de texto
                        elif elem.text and any(term in elem.text.lower() for term in ['criterio', 'criteria', 'puntos', 'points', '%']):
                            texto = elem.text.strip()
                            if len(texto) > 10 and len(texto) < 200:  # Filtrar textos muy cortos o muy largos
                                criterios.append({'descripcion': texto})
                                criterios_encontrados = True
                                st.write(f"  📝 Criterio encontrado por texto: {texto[:50]}...")

                # Estrategia 3: Búsqueda por atributos que contengan información de criterios
                if not criterios_encontrados:
                    st.write("🔍 **Búsqueda por atributos...**")

                    for elem in root.iter():
                        for attr_name, attr_value in elem.attrib.items():
                            if (any(term in attr_name.lower() for term in ['name', 'description', 'title']) and
                                any(term in str(attr_value).lower() for term in ['criterio', 'criteria', 'economic', 'technical', 'price'])):
                                criterios.append({'descripcion': str(attr_value)})
                                criterios_encontrados = True
                                st.write(f"  📝 Criterio encontrado por atributo: {attr_value[:50]}...")

                # Si aún no encontramos criterios, crear criterios por defecto basados en patrones comunes
                if not criterios_encontrados:
                    st.write("⚠️ **No se encontraron criterios específicos. Generando criterios por defecto...**")
                    criterios = [
                        {'descripcion': 'Oferta económica', 'peso': '80 puntos'},
                        {'descripcion': 'Criterios técnicos', 'peso': '20 puntos'}
                    ]
                    criterios_encontrados = True

                # Si aún no encontramos nada, hacer una búsqueda muy amplia
                if not criterios:
                    for elem in root.iter():
                        # Buscar elementos que tengan texto descriptivo largo (probables criterios)
                        if (elem.text and
                            len(elem.text.strip()) > 30 and
                            len(elem.text.strip()) < 300 and
                            any(palabra in elem.text.lower() for palabra in
                                ['económ', 'técnic', 'calidad', 'plazo', 'precio', 'punt', 'valor'])):

                            criterios.append({'descripcion': elem.text.strip()})

                # Deduplicar criterios por descripción
                criterios_unicos = []
                descripciones_vistas = set()

                for criterio in criterios:
                    if isinstance(criterio, dict):
                        desc = criterio.get('descripcion', '').lower().strip()
                    else:
                        desc = str(criterio).lower().strip()

                    # Solo agregar si no hemos visto esta descripción antes
                    if desc and desc not in descripciones_vistas and len(desc) > 5:
                        criterios_unicos.append(criterio)
                        descripciones_vistas.add(desc)

                # Limitar a máximo 8 criterios principales
                if len(criterios_unicos) > 8:
                    criterios_unicos = criterios_unicos[:8]
                    st.info(f"🔧 Limitado a 8 criterios principales de {len(criterios)} encontrados")

            datos['criterios_adjudicacion'] = criterios_unicos

            # Descripción
            for desc in root.iter():
                if any(term in desc.tag.lower() for term in ['descripcion', 'description', 'detalle']):
                    if desc.text and len(desc.text.strip()) > 20:
                        datos['descripcion'] = desc.text.strip()
                        break

            # Si no encontramos título, usar el texto más largo como descripción
            if not datos['titulo']:
                textos = []
                for elem in root.iter():
                    if elem.text and len(elem.text.strip()) > 30:
                        textos.append(elem.text.strip())
                if textos:
                    datos['titulo'] = max(textos, key=len)[:200]

            return datos

        except ET.ParseError as e:
            st.error(f"⚠️ Error al leer el XML: El archivo no es un XML válido.")
            st.error(f"Detalles técnicos: {e}")
            st.info("💡 Verifica que la URL del XML sea correcta y que el archivo esté bien formado. Si el error persiste, es posible que el XML contenga caracteres especiales no codificados correctamente.")
            return None
        except Exception as e:
            st.error(f"Error procesando XML: {e}")
            return None


    def get_contratos_data(self, limit=500):
        """Obtener datos de la tabla adjudicaciones_metabase (solo registros españoles)"""
        query = f"""
        SELECT
            id,
            titulo,
            entidad_compradora as organismo,
            fecha_publicacion,
            importe_total as presupuesto_licitacion,
            numero_licitadores as num_licitadores,
            importe_adjudicacion as precio_adjudicacion,
            adjudicatario::text as empresa_adjudicataria,
            ROUND(((importe_total - importe_adjudicacion) / NULLIF(importe_total, 0) * 100)::numeric, 2) as baja_estadistica,
            cpv::text,
            tipo_contrato,
            provincia,
            descripcion as objeto
        FROM adjudicaciones_metabase
        WHERE importe_total IS NOT NULL
        AND importe_adjudicacion IS NOT NULL
        AND importe_total > 0
        AND importe_adjudicacion > 0
        AND importe_total != importe_adjudicacion
        AND fecha_publicacion IS NOT NULL
        AND provincia NOT IN ('CZE', 'POL', 'DEU', 'FRA', 'ITA', 'PRT', 'GBR', 'NLD', 'BEL', 'AUT', 'SWE', 'DNK', 'FIN', 'NOR', 'IRL', 'GRC', 'LUX', 'HUN', 'ROU', 'BGR', 'HRV', 'SVK', 'SVN', 'EST', 'LVA', 'LTU', 'CYP', 'MLT', 'ESP')
        ORDER BY fecha_publicacion DESC
        LIMIT {limit}
        """
        return pd.read_sql(query, self.connection)

    def get_filtered_contratos_data(self, cpv_category=None, provincia=None, presupuesto=None, years=None, limit=50):
        """Obtener datos filtrados directamente desde la base de datos"""
        # Construir condiciones WHERE dinámicamente
        conditions = [
            "importe_total IS NOT NULL",
            "importe_adjudicacion IS NOT NULL",
            "importe_total > 0",
            "importe_adjudicacion > 0",
            "importe_total != importe_adjudicacion",
            "fecha_publicacion IS NOT NULL",
            "provincia NOT IN ('CZE', 'POL', 'DEU', 'FRA', 'ITA', 'PRT', 'GBR', 'NLD', 'BEL', 'AUT', 'SWE', 'DNK', 'FIN', 'NOR', 'IRL', 'GRC', 'LUX', 'HUN', 'ROU', 'BGR', 'HRV', 'SVK', 'SVN', 'EST', 'LVA', 'LTU', 'CYP', 'MLT', 'ESP')"
        ]

        # Filtro por CPV (8 dígitos exactos o 4 primeros dígitos)
        if cpv_category:
            if len(cpv_category) == 8:
                # Búsqueda exacta con CPV completo (8 dígitos)
                conditions.append(f"cpv::text ~ '{cpv_category}'")
            elif len(cpv_category) >= 4:
                # Búsqueda amplia con primeros 4 dígitos
                cpv_4_digits = cpv_category[:4]
                conditions.append(f"cpv::text ~ '{cpv_4_digits}[0-9]{{4}}'")

        # Filtro por provincia
        if provincia:
            provincia_clean = provincia.replace("'", "''")  # Escapar comillas
            conditions.append(f"LOWER(provincia) LIKE LOWER('%{provincia_clean}%')")

        # Filtro por presupuesto (0.5x a 1.5x)
        if presupuesto and presupuesto > 0:
            min_budget = presupuesto * 0.5
            max_budget = presupuesto * 1.5
            conditions.append(f"importe_total BETWEEN {min_budget} AND {max_budget}")

        # Filtro por años
        if years:
            year_conditions = [f"EXTRACT(YEAR FROM fecha_publicacion) = {year}" for year in years]
            conditions.append(f"({' OR '.join(year_conditions)})")

        where_clause = " AND ".join(conditions)

        query = f"""
        SELECT
            id,
            titulo,
            entidad_compradora as organismo,
            fecha_publicacion,
            importe_total as presupuesto_licitacion,
            numero_licitadores as num_licitadores,
            importe_adjudicacion as precio_adjudicacion,
            adjudicatario::text as empresa_adjudicataria,
            ROUND(((importe_total - importe_adjudicacion) / NULLIF(importe_total, 0) * 100)::numeric, 2) as baja_estadistica,
            cpv::text,
            tipo_contrato,
            provincia,
            descripcion as objeto
        FROM adjudicaciones_metabase
        WHERE {where_clause}
        ORDER BY fecha_publicacion DESC
        LIMIT {limit}
        """

        return pd.read_sql(query, self.connection)

    def search_previous_licitacion_same_org(self, organismo, cpv_category, presupuesto):
        """Buscar licitaciones anteriores de la misma administración con CPV similar e importe parecido"""
        if not organismo or not cpv_category:
            return None

        # Escapar comillas en el nombre del organismo
        organismo_clean = organismo.replace("'", "''")

        # Construir condiciones
        conditions = [
            "importe_total IS NOT NULL",
            "importe_adjudicacion IS NOT NULL",
            "importe_total > 0",
            "importe_adjudicacion > 0",
            "importe_total != importe_adjudicacion",
            "fecha_publicacion IS NOT NULL",
            f"LOWER(entidad_compradora) = LOWER('{organismo_clean}')"
        ]

        # Filtro por CPV (primero intentar con 8 dígitos exactos, luego 4 dígitos)
        if len(cpv_category) >= 8:
            # Intentar con CPV completo
            cpv_8_digits = cpv_category[:8]
            conditions.append(f"cpv::text ~ '{cpv_8_digits}'")
        elif len(cpv_category) >= 4:
            # Usar primeros 4 dígitos
            cpv_4_digits = cpv_category[:4]
            conditions.append(f"cpv::text ~ '{cpv_4_digits}[0-9]{{4}}'")

        # Filtro por presupuesto similar (±30%)
        if presupuesto and presupuesto > 0:
            min_budget = presupuesto * 0.7
            max_budget = presupuesto * 1.3
            conditions.append(f"importe_total BETWEEN {min_budget} AND {max_budget}")

        where_clause = " AND ".join(conditions)

        query = f"""
        SELECT
            titulo,
            entidad_compradora as organismo,
            importe_total as presupuesto_licitacion,
            importe_adjudicacion as precio_adjudicacion,
            adjudicatario as empresa_adjudicataria,
            numero_licitadores as num_licitadores,
            fecha_publicacion,
            ROUND(((importe_total - importe_adjudicacion) / NULLIF(importe_total, 0) * 100)::numeric, 2) as baja_estadistica,
            cpv::text,
            provincia
        FROM adjudicaciones_metabase
        WHERE {where_clause}
        ORDER BY fecha_publicacion DESC
        LIMIT 1
        """

        result = pd.read_sql(query, self.connection)

        if not result.empty:
            return result.iloc[0].to_dict()
        return None

    def extract_price_from_text(self, text):
        """Extraer precio de texto usando regex"""
        if pd.isna(text):
            return None

        text = str(text).replace('.', '').replace(',', '.')

        patterns = [
            r'(\d+\.?\d*)\s*€',
            r'€\s*(\d+\.?\d*)',
            r'(\d+\.?\d*)\s*euros?',
            r'euros?\s*(\d+\.?\d*)',
            r'importe[:\s]*(\d+\.?\d*)',
            r'precio[:\s]*(\d+\.?\d*)',
            r'valor[:\s]*(\d+\.?\d*)',
            r'(\d{1,3}(?:\.\d{3})*(?:,\d{2})?)'
        ]

        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                try:
                    return float(matches[0].replace(',', '.'))
                except:
                    continue
        return None

    def extract_empresa_name(self, text):
        """Extraer nombre de empresa"""
        if pd.isna(text):
            return None

        # Limpiar texto común
        text = str(text).strip()
        # Quitar patrones comunes
        text = re.sub(r'\b(S\.?L\.?|S\.?A\.?|S\.?L\.?U\.?)\b', '', text, flags=re.IGNORECASE)
        return text.strip()

    def calculate_baja_percentage(self, pbl, importe_adjudicacion):
        """Calcular porcentaje de baja: (PBL - Importe adjudicación) / PBL * 100"""
        if pd.isna(pbl) or pd.isna(importe_adjudicacion) or pbl == 0:
            return None

        baja = ((pbl - importe_adjudicacion) / pbl) * 100
        return max(0, baja)  # No permitir bajas negativas

    def find_similar_contratos_from_xml(self, xml_data, all_contratos):
        """Búsqueda avanzada de contratos similares según instrucciones de IA"""
        if not xml_data:
            return []

        # Fase 1: Búsqueda por similitud de objeto/título (principal)
        contratos_fase1 = self._search_contratos_by_object(xml_data, all_contratos)

        if len(contratos_fase1) >= 5:
            return self._filter_and_process_contratos(contratos_fase1)

        # Fase 2: Búsqueda combinada objeto + ubicación
        st.info(f"🔍 Se encontraron {len(contratos_fase1)} contratos por objeto. Expandiendo con ubicación...")
        contratos_fase2 = self._search_contratos_object_location(xml_data, all_contratos, contratos_fase1)

        if len(contratos_fase2) >= 5:
            return self._filter_and_process_contratos(contratos_fase2)

        # Fase 3: Búsqueda con CPV como apoyo (no obligatorio)
        st.info(f"🔍 Se encontraron {len(contratos_fase2)} contratos. Usando CPV como criterio adicional...")
        contratos_fase3 = self._search_contratos_cpv_support(xml_data, all_contratos, contratos_fase2)

        if len(contratos_fase3) >= 3:
            return self._filter_and_process_contratos(contratos_fase3)

        # Fase 4: Búsqueda extremadamente flexible como último recurso
        st.warning(f"🔍 Solo {len(contratos_fase3)} contratos encontrados. Aplicando búsqueda ultra-flexible...")
        contratos_fase4 = self._search_contratos_ultra_flexible(xml_data, all_contratos)

        return self._filter_and_process_contratos(contratos_fase4)

    def _search_contratos_ultra_flexible(self, xml_data, all_contratos):
        """Búsqueda ultra-flexible para encontrar cualquier contrato remotamente similar"""
        contratos_filtrados = []
        target_title = xml_data.get('titulo', '').lower()
        target_objeto = xml_data.get('objeto', '').lower()

        st.write("🚨 **BÚSQUEDA ULTRA-FLEXIBLE ACTIVADA**")
        st.write("Buscando contratos con cualquier similitud mínima...")

        # Usar datos desde 2015 para máxima flexibilidad
        all_contratos_filtered = self._filter_by_year(all_contratos, min_year=2015)
        st.write(f"- Contratos desde 2015: {len(all_contratos_filtered)}")

        # Extraer palabras clave más permisivas
        all_text = target_title + ' ' + target_objeto
        keywords = self._extract_keywords(all_text)

        # Si no tenemos keywords, usar palabras individuales
        if not keywords:
            keywords = [word for word in all_text.split() if len(word) > 2]

        st.write(f"- Buscando con keywords: {keywords[:5]}...")

        for idx, row in all_contratos_filtered.iterrows():
            score = 0

            # Obtener todo el texto del contrato
            row_title = self._extract_title_from_row(row)
            row_objeto = self._extract_objeto_from_row(row)
            combined_text = (row_title + ' ' + row_objeto).lower()

            # Cualquier keyword match da puntos
            keyword_matches = 0
            for keyword in keywords:
                if keyword in combined_text:
                    keyword_matches += 1

            # Score muy permisivo: cualquier match es válido
            if keyword_matches > 0:
                score = keyword_matches * 10  # 10 puntos por keyword

                contrato_data = self._extract_contract_data(row)
                if contrato_data is not None:
                    contrato_data['score'] = score
                    contratos_filtrados.append(contrato_data)

        st.write(f"✅ **Contratos encontrados en búsqueda ultra-flexible: {len(contratos_filtrados)}**")

        return sorted(contratos_filtrados, key=lambda x: x['score'], reverse=True)[:50]  # Máximo 50

    def _search_contratos_strict(self, xml_data, all_contratos):
        """Búsqueda estricta: CPVs exactos, ubicación, presupuesto ±30%, posteriores a 2022"""
        contratos_filtrados = []

        target_price = xml_data.get('presupuesto')
        target_location = xml_data.get('ubicacion', '')
        target_cpvs = xml_data.get('cpv', '').split(', ') if xml_data.get('cpv') else []
        target_objeto = xml_data.get('objeto', '')

        # Filtrar por año >= 2022 primero
        all_contratos_filtered = self._filter_by_year(all_contratos, min_year=2022)

        for idx, row in all_contratos_filtered.iterrows():
            score = 0

            # 1. Verificar CPVs exactos (criterio obligatorio)
            if target_cpvs:
                row_cpv = self._extract_cpv_from_row(row)
                if self._has_matching_cpv(target_cpvs, row_cpv):
                    score += 40
                else:
                    continue  # Sin CPV coincidente, saltar

            # 2. Verificar ubicación (criterio obligatorio)
            row_location = self._extract_location_from_row(row)
            if target_location and row_location:
                if self._locations_match(target_location, row_location):
                    score += 30
                else:
                    continue  # Sin ubicación coincidente, saltar

            # 3. Verificar presupuesto ±30% (criterio obligatorio)
            row_price = self._extract_price_from_row(row)
            if target_price and row_price:
                price_diff = abs(row_price - target_price) / target_price
                if price_diff <= 0.30:
                    score += 20
                else:
                    continue  # Fuera del rango de precio, saltar

            # 4. Similitud de objeto (bonus)
            if target_objeto:
                similarity = self._calculate_text_similarity(target_objeto, self._extract_objeto_from_row(row))
                if similarity > 0.3:
                    score += int(similarity * 10)

            # Extraer datos adicionales para el contrato
            contrato_data = self._extract_contract_data(row)
            if contrato_data and score >= 80:  # Umbral alto para búsqueda estricta
                contrato_data['score'] = score
                contrato_data['index'] = idx
                contratos_filtrados.append(contrato_data)

        return sorted(contratos_filtrados, key=lambda x: x['score'], reverse=True)

    def _search_contratos_expanded(self, xml_data, all_contratos):
        """Búsqueda expandida: provincias cercanas, presupuesto ±50%, todos los años"""
        contratos_filtrados = []

        target_price = xml_data.get('presupuesto')
        target_location = xml_data.get('ubicacion', '')
        target_cpvs = xml_data.get('cpv', '').split(', ') if xml_data.get('cpv') else []
        target_objeto = xml_data.get('objeto', '')

        # Obtener provincias cercanas
        nearby_provinces = self._get_nearby_provinces(target_location)

        for idx, row in all_contratos.iterrows():
            score = 0

            # 1. Verificar CPVs exactos (sigue siendo obligatorio)
            if target_cpvs:
                row_cpv = self._extract_cpv_from_row(row)
                if self._has_matching_cpv(target_cpvs, row_cpv):
                    score += 40
                else:
                    continue

            # 2. Verificar ubicación expandida
            row_location = self._extract_location_from_row(row)
            if target_location and row_location:
                if (self._locations_match(target_location, row_location) or
                    any(self._locations_match(prov, row_location) for prov in nearby_provinces)):
                    score += 25
                else:
                    continue

            # 3. Verificar presupuesto ±50%
            row_price = self._extract_price_from_row(row)
            if target_price and row_price:
                price_diff = abs(row_price - target_price) / target_price
                if price_diff <= 0.50:
                    score += 15
                else:
                    continue

            # 4. Similitud de objeto
            if target_objeto:
                similarity = self._calculate_text_similarity(target_objeto, self._extract_objeto_from_row(row))
                if similarity > 0.2:
                    score += int(similarity * 10)

            contrato_data = self._extract_contract_data(row)
            if contrato_data and score >= 60:  # Umbral reducido para búsqueda expandida
                contrato_data['score'] = score
                contrato_data['index'] = idx
                contratos_filtrados.append(contrato_data)

        return sorted(contratos_filtrados, key=lambda x: x['score'], reverse=True)

    def _search_contratos_cpv_broad(self, xml_data, all_contratos):
        """Búsqueda por CPV ampliado (4 primeros dígitos)"""
        contratos_filtrados = []

        target_cpvs = xml_data.get('cpv', '').split(', ') if xml_data.get('cpv') else []
        target_cpvs_broad = [cpv[:4] for cpv in target_cpvs if len(cpv) >= 4]

        if not target_cpvs_broad:
            return contratos_filtrados

        for idx, row in all_contratos.iterrows():
            score = 0

            # Verificar CPVs por los 3 primeros dígitos
            row_cpv = self._extract_cpv_from_row(row)
            if any(row_cpv.startswith(broad_cpv) for broad_cpv in target_cpvs_broad):
                score += 30
            else:
                continue

            contrato_data = self._extract_contract_data(row)
            if contrato_data and score >= 25:  # Umbral muy bajo para búsqueda amplia
                contrato_data['score'] = score
                contrato_data['index'] = idx
                contratos_filtrados.append(contrato_data)

        return sorted(contratos_filtrados, key=lambda x: x['score'], reverse=True)

    def _filter_by_year(self, all_contratos, min_year=2022):
        """Filtrar contratos por año mínimo"""
        filtered_contratos = []

        for idx, row in all_contratos.iterrows():
            # Buscar columnas de fecha
            fecha_encontrada = False
            for col in all_contratos.columns:
                if any(term in col.lower() for term in ['fecha', 'date', 'ano', 'year', 'anio']):
                    fecha_str = str(row.get(col, ''))
                    if fecha_str and len(fecha_str) >= 4:
                        # Extraer año de la fecha
                        year_match = re.search(r'(20\d{2})', fecha_str)
                        if year_match:
                            year = int(year_match.group(1))
                            if year >= min_year:
                                filtered_contratos.append(row)
                                fecha_encontrada = True
                                break

            # Si no encontramos fecha, incluir el contrato (será filtrado por otros criterios)
            if not fecha_encontrada:
                filtered_contratos.append(row)

        return pd.DataFrame(filtered_contratos) if filtered_contratos else all_contratos

    def _extract_cpv_from_row(self, row):
        """Extraer código CPV de una fila"""
        for col in row.index:
            if any(term in col.lower() for term in ['cpv', 'codigo', 'clasificacion']):
                value = str(row.get(col, ''))
                if value and len(value) >= 8:
                    # Extraer números de 8 dígitos o más
                    cpv_match = re.search(r'(\d{8,})', value)
                    if cpv_match:
                        return cpv_match.group(1)
        return ''

    def _extract_location_from_row(self, row):
        """Extraer ubicación de una fila"""
        location_columns = ['provincia', 'ubicacion', 'lugar', 'localidad', 'direccion', 'comunidad']
        for col in row.index:
            if any(term in col.lower() for term in location_columns):
                value = str(row.get(col, '')).strip()
                if value and len(value) > 2:
                    return value
        return ''

    def _extract_price_from_row(self, row):
        """Extraer precio de una fila"""
        price_columns = ['precio', 'importe', 'valor', 'presupuesto', 'cantidad', 'pbl']
        for col in row.index:
            if any(term in col.lower() for term in price_columns):
                value = str(row.get(col, ''))
                if value:
                    price = self.extract_price_from_text(value)
                    if price:
                        return price
        return None

    def _extract_objeto_from_row(self, row):
        """Extraer objeto/descripción de una fila"""
        objeto_columns = ['objeto', 'descripcion', 'servicio', 'titulo', 'name']
        for col in row.index:
            if any(term in col.lower() for term in objeto_columns):
                value = str(row.get(col, '')).strip()
                if value and len(value) > 10:
                    return value
        return ''

    def _has_matching_cpv(self, target_cpvs, row_cpv):
        """Verificar si hay coincidencia de CPV"""
        if not row_cpv:
            return False

        for target_cpv in target_cpvs:
            if target_cpv.strip() and row_cpv.startswith(target_cpv.strip()):
                return True
        return False

    def _locations_match(self, loc1, loc2):
        """Verificar si dos ubicaciones coinciden (incluye coincidencias parciales)"""
        if not loc1 or not loc2:
            return False

        loc1_clean = loc1.lower().strip()
        loc2_clean = loc2.lower().strip()

        # Coincidencia exacta
        if loc1_clean == loc2_clean:
            return True

        # Coincidencia parcial (una contiene a la otra)
        if loc1_clean in loc2_clean or loc2_clean in loc1_clean:
            return True

        return False

    def _calculate_text_similarity(self, text1, text2):
        """Calcular similitud entre dos textos usando TF-IDF"""
        if not text1 or not text2:
            return 0.0

        try:
            vectorizer = TfidfVectorizer(stop_words=None, ngram_range=(1, 2), max_features=1000)
            tfidf_matrix = vectorizer.fit_transform([text1, text2])
            similarity = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]
            return similarity
        except:
            return 0.0

    def _get_nearby_provinces(self, target_location):
        """Obtener provincias cercanas (simplificado - se puede expandir con datos reales)"""
        if not target_location:
            return []

        # Mapeo simplificado de provincias cercanas
        nearby_map = {
            'madrid': ['toledo', 'guadalajara', 'segovia', 'avila', 'cuenca'],
            'barcelona': ['tarragona', 'lleida', 'girona', 'valencia'],
            'valencia': ['castellon', 'alicante', 'teruel', 'cuenca', 'albacete'],
            'sevilla': ['cadiz', 'huelva', 'cordoba', 'malaga'],
            'bilbao': ['vizcaya', 'alava', 'guipuzcoa', 'cantabria'],
            # Añadir más según necesidad
        }

        target_lower = target_location.lower()
        for key, nearby in nearby_map.items():
            if key in target_lower or target_lower in key:
                return nearby

        return []

    def _extract_contract_data(self, row):
        """Extraer todos los datos relevantes del contrato"""
        # Extraer PBL e Importe de adjudicación
        pbl = self._extract_price_from_row(row)

        # Buscar importe de adjudicación específicamente
        importe_adj = None
        empresa = None

        for col in row.index:
            col_lower = col.lower()
            if 'adjudicacion' in col_lower or 'adjudicado' in col_lower:
                importe_adj = self.extract_price_from_text(str(row.get(col, '')))
            elif 'empresa' in col_lower or 'adjudicatario' in col_lower:
                empresa = self.extract_empresa_name(str(row.get(col, '')))

        # Calcular baja si tenemos los datos
        baja_percentage = None
        if pbl and importe_adj:
            baja_percentage = self.calculate_baja_percentage(pbl, importe_adj)

            # Filtrar según instrucciones: eliminar bajas > 70% o < 0.5% o con adjudicatario vacío
            if (baja_percentage > 70 or baja_percentage < 0.5 or
                not empresa or empresa.lower() in ['none', 'vacio', '', 'null']):
                return None

        if baja_percentage is not None:
            return {
                'pbl': pbl,
                'importe_adjudicacion': importe_adj,
                'baja_percentage': baja_percentage,
                'empresa': empresa,
                'precio': pbl,
                'num_licitadores': self._extract_num_licitadores(row)
            }

        return None

    def _extract_num_licitadores(self, row):
        """Extraer número de licitadores de una fila"""
        for col in row.index:
            if any(term in col.lower() for term in ['licitador', 'participante', 'oferent', 'empresa']):
                value = str(row.get(col, ''))
                # Buscar números en el texto
                numbers = re.findall(r'\d+', value)
                if numbers:
                    return int(numbers[0])
        return 1  # Default si no encontramos datos

    def _search_contratos_by_object(self, xml_data, all_contratos):
        """Búsqueda simplificada: CPV (3 dígitos) + Ubicación + Presupuesto"""
        target_price = xml_data.get('presupuesto', 0)
        target_location = xml_data.get('ubicacion', '')
        target_cpv = xml_data.get('cpv', '')
        target_objeto = xml_data.get('objeto', '')

        st.write("🔍 **BÚSQUEDA POR CPV Y UBICACIÓN:**")
        st.write(f"**CPV:** {target_cpv}")
        st.write(f"**Presupuesto:** €{target_price:,.2f}")
        st.write(f"**Ubicación:** {target_location}")

        # Extraer 3 primeros dígitos del CPV principal
        cpv_category = self._extract_cpv_category_from_multiple(target_cpv, target_objeto)
        st.write(f"**Categoría CPV (3 dígitos):** {cpv_category}")

        # Estrategia progresiva por años
        strategies = [
            {"name": "📅 2025", "year": 2025, "budget_range": (0.5, 1.5)},
            {"name": "📅 2024", "year": 2024, "budget_range": (0.5, 1.5)},
            {"name": "📅 2023", "year": 2023, "budget_range": (0.5, 1.5)},
            {"name": "📅 2022", "year": 2022, "budget_range": (0.5, 1.5)}
        ]

        for strategy in strategies:
            st.write(f"\n### {strategy['name']}")

            contratos_found = self._search_by_cpv_location(
                all_contratos,
                cpv_category,
                target_location,
                target_price,
                strategy['year'],
                strategy['budget_range']
            )

            if len(contratos_found) >= 3:
                st.success(f"✅ **Encontrados {len(contratos_found)} contratos en {strategy['name']}**")
                return contratos_found
            else:
                st.warning(f"⚠️ Solo {len(contratos_found)} contratos en {strategy['name']}. Probando año anterior...")

        return contratos_found if 'contratos_found' in locals() else []

    def _extract_cpv_category(self, cpv_string):
        """Extraer los 3 primeros dígitos del CPV principal"""
        if not cpv_string:
            return ""

        # Buscar el primer código CPV en el string
        cpv_match = re.search(r'(\d{8})', cpv_string)
        if cpv_match:
            cpv_code = cpv_match.group(1)
            return cpv_code[:4]  # Primeros 4 dígitos
        return ""

    def _extract_cpv_full(self, cpv_string):
        """Extraer el CPV completo (8 dígitos)"""
        if not cpv_string:
            return ""

        # Extraer todos los códigos CPV de 8 dígitos
        cpv_codes = re.findall(r'(\d{8})', cpv_string)

        if not cpv_codes:
            return ""

        # Devolver el primer CPV completo (8 dígitos)
        return cpv_codes[0]

    def _extract_cpv_category_from_multiple(self, cpv_string, objeto_text):
        """Extraer el CPV más relevante según el objeto cuando hay múltiples CPVs"""
        if not cpv_string:
            return ""

        # Si el CPV viene en formato JSON array: ["12345678", "87654321"]
        # Extraer todos los códigos CPV de 8 dígitos
        cpv_codes = re.findall(r'(\d{8})', cpv_string)

        if not cpv_codes:
            return ""

        if len(cpv_codes) == 1:
            # Si solo hay un CPV, devolverlo
            return cpv_codes[0][:4]

        # Si hay múltiples CPVs, intentar elegir el más relevante
        # Por simplicidad, tomamos el primero (principal)
        # En el futuro se puede mejorar para comparar con el objeto
        st.write(f"   📋 Encontrados {len(cpv_codes)} CPVs: {cpv_codes[:3]}... Usando el primero")
        return cpv_codes[0][:4]

    def _search_by_cpv_location(self, all_contratos, cpv_category, target_location, target_price, year, budget_range):
        """Buscar contratos por CPV, ubicación y presupuesto en un año específico"""
        contratos_found = []

        # Filtrar por año
        contratos_year = []
        for idx, row in all_contratos.iterrows():
            fecha = row.get('fecha_publicacion', '')
            if pd.notna(fecha):
                fecha_str = str(fecha)
                if str(year) in fecha_str:
                    contratos_year.append(row)

        if not contratos_year:
            st.write(f"   ⚠️ No hay contratos del año {year}")
            return []

        contratos_year_df = pd.DataFrame(contratos_year)
        st.write(f"   - Contratos en {year}: {len(contratos_year_df)}")

        # Filtrar por presupuesto
        if target_price > 0:
            min_budget = target_price * budget_range[0]
            max_budget = target_price * budget_range[1]
            contratos_budget = []
            for idx, row in contratos_year_df.iterrows():
                presupuesto = row.get('presupuesto_licitacion', 0)
                if pd.notna(presupuesto) and min_budget <= presupuesto <= max_budget:
                    contratos_budget.append(row)

            if not contratos_budget:
                st.write(f"   ⚠️ No hay contratos en rango €{min_budget:,.0f} - €{max_budget:,.0f}")
                return []

            contratos_budget_df = pd.DataFrame(contratos_budget)
            st.write(f"   - Presupuesto €{min_budget:,.0f} - €{max_budget:,.0f}: {len(contratos_budget_df)}")
        else:
            contratos_budget_df = contratos_year_df

        # Evaluar cada contrato por CPV y ubicación
        for idx, row in contratos_budget_df.iterrows():
            score = 0
            score_detail = {'cpv': 0, 'location': 0}

            # Score por CPV (60 puntos)
            if cpv_category and len(cpv_category) >= 4:
                row_cpv = str(row.get('cpv', ''))
                # Extraer todos los códigos CPV de 8 dígitos del contrato
                row_cpv_codes = re.findall(r'(\d{8})', row_cpv)

                # Comparar los 4 primeros dígitos de cada CPV encontrado
                cpv_match = False
                for cpv_code in row_cpv_codes:
                    if cpv_code[:4] == cpv_category:
                        cpv_match = True
                        break

                if cpv_match:
                    score += 60
                    score_detail['cpv'] = 60

            # Score por ubicación (40 puntos)
            if target_location:
                row_location = str(row.get('provincia', '')).lower()
                target_loc_lower = target_location.lower()
                if row_location and any(loc.lower() in row_location for loc in target_loc_lower.split()):
                    score += 40
                    score_detail['location'] = 40

            # Aceptar contratos con score >= 40 (al menos uno de los criterios)
            if score >= 40:
                contrato_data = self._extract_contract_data(row)
                if contrato_data is not None:
                    contrato_data['score'] = score
                    contrato_data['score_detail'] = score_detail
                    contratos_found.append(contrato_data)

        # Ordenar por score
        contratos_found.sort(key=lambda x: x['score'], reverse=True)

        # Mostrar resultados
        st.write(f"\n   **📊 Top contratos encontrados:**")
        for i, contrato in enumerate(contratos_found[:10]):
            detail = contrato['score_detail']
            st.write(f"   {i+1}. **{contrato.get('titulo', 'Sin título')[:60]}...**")
            st.write(f"      Score: {contrato['score']} (CPV: {detail['cpv']}, Ubicación: {detail['location']})")

        return contratos_found

    def _extract_main_keywords(self, text):
        """Extraer 1-2 palabras principales del objeto del contrato"""
        if not text:
            return []

        text_lower = text.lower()

        # Palabras irrelevantes que filtrar
        stop_words = {
            'de', 'del', 'la', 'el', 'en', 'y', 'para', 'con', 'por', 'su', 'al', 'los', 'las',
            'un', 'una', 'o', 'e', 'le', 'se', 'que', 'es', 'son', 'como', 'entre', 'desde',
            'hasta', 'sobre', 'bajo', 'mediante', 'durante', 'dentro', 'fuera', 'sin', 'ante',
            'obras', 'necesarias', 'tres', 'comarca', 'sierra', 'oeste', 'servicios'
        }

        # Palabras clave importantes que priorizar
        priority_keywords = [
            'aparcamiento', 'aparcamientos', 'estacionamiento', 'estacionamientos',
            'construcción', 'edificación', 'infraestructura', 'urbanización',
            'proyecto', 'redacción', 'elaboración', 'diseño',
            'ejecución', 'realización', 'desarrollo',
            'mantenimiento', 'conservación', 'reparación',
            'suministro', 'adquisición', 'compra',
            'limpieza', 'seguridad', 'informática', 'transporte'
        ]

        # Buscar palabras prioritarias primero
        found_keywords = []
        for keyword in priority_keywords:
            if keyword in text_lower and keyword not in found_keywords:
                found_keywords.append(keyword)
                if len(found_keywords) >= 2:
                    break

        # Si no encontramos suficientes, buscar otras palabras relevantes
        if len(found_keywords) < 2:
            words = re.findall(r'\b[a-záéíóúñ]{5,}\b', text_lower)
            for word in words:
                if word not in stop_words and word not in found_keywords:
                    found_keywords.append(word)
                    if len(found_keywords) >= 2:
                        break

        return found_keywords[:2]  # Máximo 2 palabras

    def _simple_search(self, xml_data, all_contratos, cpv_category, main_keywords, year_min, budget_flex, use_cpv, threshold):
        """Búsqueda simple y directa"""
        target_price = xml_data.get('presupuesto', 0)
        target_location = xml_data.get('ubicacion', '')

        # Filtros básicos
        filtered_contratos = self._filter_by_year(all_contratos, min_year=year_min)
        st.write(f"- Contratos desde {year_min}: {len(filtered_contratos)}")

        if target_price > 0:
            filtered_contratos = self._filter_by_budget_range(filtered_contratos, target_price, budget_flex)
            min_budget = target_price / budget_flex
            max_budget = target_price * budget_flex
            st.write(f"- Presupuesto {min_budget:,.0f}€ - {max_budget:,.0f}€: {len(filtered_contratos)}")

        # Evaluación de cada contrato
        contratos_scored = []

        for idx, row in filtered_contratos.iterrows():
            score = 0

            # Score por CPV (si se usa)
            cpv_score = 0
            if use_cpv and cpv_category:
                row_cpv = self._extract_cpv_from_row(row)
                if cpv_category in row_cpv:
                    cpv_score = 30
                    score += cpv_score

            # Score por palabras principales
            keyword_score = 0
            row_title = self._extract_title_from_row(row)
            row_objeto = self._extract_objeto_from_row(row)
            combined_text = (row_title + ' ' + row_objeto).lower()

            for keyword in main_keywords:
                if keyword in combined_text:
                    keyword_score += 25

            score += keyword_score

            # Score por ubicación
            location_score = 0
            if target_location:
                row_location = self._extract_location_from_row(row).lower()
                if row_location and any(loc.lower() in row_location for loc in target_location.split()):
                    location_score = 15
                    score += location_score

            if score >= threshold:
                contrato_data = self._extract_contract_data(row)
                if contrato_data is not None:
                    contrato_data['score'] = score
                    contrato_data['score_detail'] = {
                        'cpv': cpv_score,
                        'keywords': keyword_score,
                        'location': location_score
                    }
                    contratos_scored.append(contrato_data)

        # Mostrar resultados
        contratos_scored.sort(key=lambda x: x['score'], reverse=True)

        st.write(f"**📊 Top 10 contratos encontrados:**")
        for i, contrato in enumerate(contratos_scored[:10]):
            detail = contrato['score_detail']
            st.write(f"{i+1}. **{contrato.get('titulo', 'Sin título')[:60]}...**")
            st.write(f"   Score: {contrato['score']:.1f} (CPV: {detail['cpv']}, Palabras: {detail['keywords']}, Ubicación: {detail['location']})")

        return contratos_scored

    def _ai_guided_search(self, xml_data, all_contratos, year_min, budget_flex, threshold):
        """Búsqueda guiada por análisis inteligente del contenido"""
        target_title = xml_data.get('titulo', '')
        target_objeto = xml_data.get('objeto', '')
        target_price = xml_data.get('presupuesto', 0)
        target_location = xml_data.get('ubicacion', '')

        # IA: Analizar el objeto para identificar conceptos clave
        search_concepts = self._analyze_contract_nature(target_title + ' ' + target_objeto)
        st.write(f"**🧠 IA identifica conceptos:** {', '.join(search_concepts)}")

        # Filtros progresivos
        filtered_contratos = self._filter_by_year(all_contratos, min_year=year_min)
        st.write(f"- Contratos desde {year_min}: {len(filtered_contratos)}")

        if target_price > 0:
            filtered_contratos = self._filter_by_budget_range(filtered_contratos, target_price, budget_flex)
            min_budget = target_price / budget_flex
            max_budget = target_price * budget_flex
            st.write(f"- Presupuesto {min_budget:,.0f}€ - {max_budget:,.0f}€: {len(filtered_contratos)}")

        # Evaluación inteligente de cada contrato
        contratos_scored = []

        for idx, row in filtered_contratos.iterrows():
            row_title = self._extract_title_from_row(row)
            row_objeto = self._extract_objeto_from_row(row)
            row_location = self._extract_location_from_row(row)

            # IA: Calcular similitud conceptual
            similarity_score = self._calculate_conceptual_similarity(
                search_concepts, row_title + ' ' + row_objeto
            )

            # Bonus por ubicación
            location_bonus = 0
            if target_location and row_location:
                if any(loc.lower() in row_location.lower() for loc in target_location.split()):
                    location_bonus = 15

            total_score = similarity_score + location_bonus

            if total_score >= threshold:
                contrato_data = self._extract_contract_data(row)
                if contrato_data is not None:
                    contrato_data['score'] = total_score
                    contrato_data['similarity_detail'] = {
                        'conceptual': similarity_score,
                        'location': location_bonus
                    }
                    contratos_scored.append(contrato_data)

        # Mostrar resultados del análisis
        contratos_scored.sort(key=lambda x: x['score'], reverse=True)

        st.write(f"**📊 Top 10 contratos analizados:**")
        for i, contrato in enumerate(contratos_scored[:10]):
            st.write(f"{i+1}. **{contrato.get('titulo', 'Sin título')[:60]}...**")
            st.write(f"   Score: {contrato['score']:.1f} (Conceptual: {contrato['similarity_detail']['conceptual']:.1f}, Ubicación: {contrato['similarity_detail']['location']:.1f})")

        return contratos_scored

    def _analyze_contract_nature(self, text):
        """IA: Analizar inteligentemente QUÉ tipo de contrato es y QUÉ buscar"""
        text_lower = text.lower()

        # LÓGICA INTELIGENTE: ¿De qué trata REALMENTE este contrato?
        main_concepts = []

        # 1. ¿Es sobre INFRAESTRUCTURA/CONSTRUCCIÓN?
        if any(word in text_lower for word in ['construcción', 'obra', 'edificación', 'infraestructura']):
            main_concepts.append('construccion')

        # 2. ¿Es específicamente sobre APARCAMIENTOS?
        if any(word in text_lower for word in ['aparcamiento', 'estacionamiento', 'parking', 'disuasorio']):
            main_concepts.append('aparcamiento')

        # 3. ¿Incluye DISEÑO/PROYECTO?
        if any(word in text_lower for word in ['proyecto', 'redacción', 'diseño', 'plan']):
            main_concepts.append('proyecto')

        # 4. ¿Incluye EJECUCIÓN/REALIZACIÓN?
        if any(word in text_lower for word in ['ejecución', 'realización']):
            main_concepts.append('ejecucion')

        # 5. ¿Es sobre MANTENIMIENTO?
        if any(word in text_lower for word in ['mantenimiento', 'conservación']):
            main_concepts.append('mantenimiento')

        # 6. ¿Es sobre SUMINISTROS?
        if any(word in text_lower for word in ['suministro', 'adquisición', 'compra']):
            main_concepts.append('suministro')

        # 7. ¿Es sobre SERVICIOS GENERALES?
        if any(word in text_lower for word in ['servicios', 'asistencia']) and len(main_concepts) == 0:
            main_concepts.append('servicios')

        # Si no encontramos nada específico, buscar por palabras más generales
        if not main_concepts:
            # Buscar la palabra más importante (sustantivos relevantes)
            important_words = re.findall(r'\b(limpieza|seguridad|informática|transporte|energía|agua|residuos)\b', text_lower)
            if important_words:
                main_concepts.append(important_words[0])
            else:
                # Última opción: palabras más largas
                words = re.findall(r'\b[a-záéíóúñ]{6,}\b', text_lower)
                main_concepts = words[:2]  # Solo 2 palabras máximo

        # MÁXIMO 3 conceptos para búsqueda efectiva
        return main_concepts[:3]

    def _calculate_conceptual_similarity(self, search_concepts, contract_text):
        """IA: Calcular similitud conceptual SIMPLIFICADA y más permisiva"""
        if not search_concepts:
            return 0

        text_lower = contract_text.lower()

        # LÓGICA MUCHO MÁS SIMPLE Y PERMISIVA
        total_score = 0

        for concept in search_concepts:
            concept_score = 0

            # 1. Buscar el concepto directamente
            if concept in text_lower:
                concept_score = 25

            # 2. Buscar palabras relacionadas AMPLIAS
            elif concept == 'construccion':
                if any(word in text_lower for word in ['obra', 'construcción', 'edificación', 'infraestructura', 'civil', 'urbanización']):
                    concept_score = 20

            elif concept == 'aparcamiento':
                if any(word in text_lower for word in ['aparcamiento', 'estacionamiento', 'parking', 'garaje']):
                    concept_score = 20

            elif concept == 'proyecto':
                if any(word in text_lower for word in ['proyecto', 'redacción', 'elaboración', 'diseño', 'plan', 'estudio']):
                    concept_score = 20

            elif concept == 'ejecucion':
                if any(word in text_lower for word in ['ejecución', 'realización', 'desarrollo', 'construcción']):
                    concept_score = 20

            elif concept == 'mantenimiento':
                if any(word in text_lower for word in ['mantenimiento', 'conservación', 'reparación']):
                    concept_score = 20

            elif concept == 'suministro':
                if any(word in text_lower for word in ['suministro', 'adquisición', 'compra', 'provisión']):
                    concept_score = 20

            elif concept == 'servicios':
                if any(word in text_lower for word in ['servicios', 'asistencia', 'gestión']):
                    concept_score = 20

            # 3. Si es una palabra específica, buscarla directamente
            else:
                if concept in text_lower:
                    concept_score = 15
                # Búsqueda parcial muy permisiva
                elif len(concept) > 3:
                    root = concept[:4]
                    if root in text_lower:
                        concept_score = 10

            total_score += concept_score

        # CUALQUIER coincidencia da puntos decentes
        if total_score > 0:
            total_score += 15  # Bonus base por cualquier match

        return min(total_score, 100)  # Máximo 100 puntos

    def _search_contratos_object_location(self, xml_data, all_contratos, existing_contratos):
        """Búsqueda combinada por objeto y ubicación"""
        contratos_filtrados = list(existing_contratos)  # Empezar con los encontrados por objeto
        target_location = xml_data.get('ubicacion', '')

        # Si no tenemos ubicación, devolver los existentes
        if not target_location:
            return contratos_filtrados

        # Buscar más contratos que coincidan en ubicación
        for idx, row in all_contratos.iterrows():
            # Verificar que no esté ya en la lista
            row_empresa = self._extract_empresa_from_row(row)
            if any(c.get('empresa') == row_empresa for c in contratos_filtrados):
                continue

            row_location = self._extract_location_from_row(row)
            if self._locations_match(target_location, row_location):
                # Dar bonus por ubicación similar
                contrato_data = self._extract_contract_data(row)
                if contrato_data is not None:  # Verificar que no sea None
                    contrato_data['score'] = 25  # Score base por ubicación
                    contratos_filtrados.append(contrato_data)

        return sorted(contratos_filtrados, key=lambda x: x['score'], reverse=True)

    def _search_contratos_cpv_support(self, xml_data, all_contratos, existing_contratos):
        """Búsqueda con CPV como criterio de apoyo (no obligatorio)"""
        contratos_filtrados = list(existing_contratos)
        target_cpvs = xml_data.get('cpv', '').split(', ') if xml_data.get('cpv') else []

        if not target_cpvs:
            return contratos_filtrados

        # Buscar contratos adicionales que coincidan en CPV
        for idx, row in all_contratos.iterrows():
            # Verificar que no esté ya en la lista
            row_empresa = self._extract_empresa_from_row(row)
            if any(c.get('empresa') == row_empresa for c in contratos_filtrados):
                continue

            row_cpv = self._extract_cpv_from_row(row)

            # Verificar CPV exacto o los primeros 3-4 dígitos
            cpv_match = False
            for target_cpv in target_cpvs:
                if target_cpv.strip():
                    # CPV exacto
                    if row_cpv.startswith(target_cpv.strip()):
                        cpv_match = True
                        break
                    # CPV por categoría (primeros 3-4 dígitos)
                    elif len(target_cpv) >= 3 and row_cpv.startswith(target_cpv[:4]):
                        cpv_match = True
                        break

            if cpv_match:
                contrato_data = self._extract_contract_data(row)
                if contrato_data is not None:  # Verificar que no sea None
                    contrato_data['score'] = 20  # Score base por CPV
                    contratos_filtrados.append(contrato_data)

        return sorted(contratos_filtrados, key=lambda x: x['score'], reverse=True)

    def _extract_keywords(self, text):
        """Extraer palabras clave relevantes del texto"""
        if not text:
            return []

        # Palabras irrelevantes que debemos filtrar
        stop_words = {
            'de', 'del', 'la', 'el', 'en', 'y', 'para', 'con', 'por', 'su', 'al', 'los', 'las',
            'un', 'una', 'o', 'e', 'le', 'se', 'que', 'es', 'son', 'como', 'entre', 'desde',
            'hasta', 'sobre', 'bajo', 'mediante', 'durante', 'dentro', 'fuera', 'sin', 'ante'
        }

        # Extraer palabras de 3+ caracteres y filtrar stop words
        words = re.findall(r'\b[a-záéíóúñ]{3,}\b', text.lower())
        keywords = [w for w in words if w not in stop_words]

        # Retornar las más importantes (máximo 10)
        return keywords[:10]

    def _extract_semantic_keywords(self, text):
        """Extraer palabras clave semánticas importantes para la naturaleza del contrato"""
        if not text:
            return []

        # Palabras irrelevantes extendida
        stop_words = {
            'de', 'del', 'la', 'el', 'en', 'y', 'para', 'con', 'por', 'su', 'al', 'los', 'las',
            'un', 'una', 'o', 'e', 'le', 'se', 'que', 'es', 'son', 'como', 'entre', 'desde',
            'hasta', 'sobre', 'bajo', 'mediante', 'durante', 'dentro', 'fuera', 'sin', 'ante',
            'obras', 'necesarias', 'construcción', 'tres', 'comarca', 'sierra', 'oeste',
            'servicios', 'elaboración', 'proyectos', 'diseños', 'presupuestos'
        }

        # Priorizar palabras clave semánticamente importantes
        priority_words = {
            'aparcamiento', 'aparcamientos', 'estacionamiento', 'estacionamientos',
            'redacción', 'proyecto', 'ejecución', 'disuasorio', 'disuasorios',
            'arquitectura', 'ingeniería', 'construcción', 'urbanismo', 'infraestructura'
        }

        text_lower = text.lower()
        keywords = []

        # Primero buscar palabras prioritarias
        for word in priority_words:
            if word in text_lower:
                keywords.append(word)

        # Luego extraer otras palabras relevantes
        words = re.findall(r'\b[a-záéíóúñ]{4,}\b', text_lower)
        for word in words:
            if word not in stop_words and word not in keywords and len(keywords) < 8:
                keywords.append(word)

        return keywords[:8]

    def _calculate_semantic_similarity(self, semantic_keywords, text):
        """Calcular similitud semántica avanzada"""
        if not semantic_keywords or not text:
            return 0

        # Sinónimos y variaciones para mejorar la búsqueda
        synonyms = {
            'aparcamiento': ['aparcamiento', 'estacionamiento', 'parking', 'garaje'],
            'aparcamientos': ['aparcamientos', 'estacionamientos', 'parkings', 'garajes'],
            'redacción': ['redacción', 'elaboración', 'desarrollo', 'diseño'],
            'ejecución': ['ejecución', 'construcción', 'realización', 'obra'],
            'proyecto': ['proyecto', 'plan', 'diseño', 'estudio']
        }

        score = 0
        total_keywords = len(semantic_keywords)

        for keyword in semantic_keywords:
            # Buscar coincidencia exacta
            if keyword in text:
                score += 1
            else:
                # Buscar sinónimos
                if keyword in synonyms:
                    for synonym in synonyms[keyword]:
                        if synonym in text:
                            score += 0.8  # Peso menor para sinónimos
                            break
                # Buscar coincidencia parcial (raíz de palabra)
                elif len(keyword) > 5:
                    root = keyword[:5]
                    if root in text:
                        score += 0.5

        return (score / total_keywords) * 100 if total_keywords > 0 else 0

    def _filter_by_budget_range(self, contratos_df, target_budget, multiplier):
        """Filtrar contratos por rango de presupuesto"""
        if target_budget <= 0:
            return contratos_df

        min_budget = target_budget / multiplier
        max_budget = target_budget * multiplier

        filtered_contratos = []

        for idx, row in contratos_df.iterrows():
            row_price = self._extract_price_from_row(row)
            if row_price and min_budget <= row_price <= max_budget:
                filtered_contratos.append(row)

        return pd.DataFrame(filtered_contratos) if filtered_contratos else pd.DataFrame()

    def _calculate_keyword_similarity(self, keywords, text):
        """Calcular similitud basada en palabras clave - versión mejorada"""
        if not keywords or not text:
            return 0

        matches = 0
        partial_matches = 0

        for keyword in keywords:
            if keyword in text:
                matches += 1
            else:
                # Buscar coincidencias parciales (raíz de palabra)
                if len(keyword) > 4:
                    root = keyword[:4]
                    if root in text:
                        partial_matches += 0.5

        total_score = matches + partial_matches
        return (total_score / len(keywords)) * 100 if keywords else 0

    def _extract_title_from_row(self, row):
        """Extraer título/objeto del contrato de una fila"""
        title_columns = ['titulo', 'objeto', 'descripcion', 'title', 'object', 'description']
        for col in title_columns:
            if col in row.index:
                value = str(row.get(col, ''))
                if value and len(value) > 10:  # Mínimo de caracteres
                    return value
        return ''

    def _extract_objeto_from_row(self, row):
        """Extraer objeto adicional del contrato"""
        objeto_columns = ['objeto_contrato', 'objeto', 'description', 'descripcion']
        for col in objeto_columns:
            if col in row.index:
                value = str(row.get(col, ''))
                if value and len(value) > 5:
                    return value
        return ''

    def _extract_empresa_from_row(self, row):
        """Extraer nombre de empresa de una fila"""
        import json
        empresa_columns = ['empresa', 'adjudicatario', 'nombre_adjudicatario', 'empresa_adjudicataria']
        for col in empresa_columns:
            if col in row.index:
                value = str(row.get(col, ''))
                if value and len(value) > 2:
                    # Si es un JSON, extraer el campo "name"
                    if value.strip().startswith('{'):
                        try:
                            empresa_data = json.loads(value)
                            name = empresa_data.get('name', value)
                            if name and len(name) > 2:
                                return name
                        except:
                            pass
                    return value
        return ''

    def _extract_price_from_row(self, row):
        """Extraer precio del contrato de una fila"""
        price_columns = ['presupuesto', 'importe', 'precio', 'valor', 'amount']
        for col in price_columns:
            if col in row.index:
                try:
                    value = float(row.get(col, 0))
                    if value > 0:
                        return value
                except:
                    continue
        return 0

    def _extract_tipo_from_row(self, row):
        """Extraer tipo de contrato de una fila"""
        tipo_columns = ['tipo', 'tipo_contrato', 'procedimiento', 'type']
        for col in tipo_columns:
            if col in row.index:
                value = str(row.get(col, ''))
                if value:
                    return value
        return ''

    def _filter_and_process_contratos(self, contratos):
        """Filtrar y procesar la lista final de contratos"""
        # Filtrar contratos válidos
        contratos_validos = [c for c in contratos if c is not None]

        # Limitar a los 50 mejores
        return sorted(contratos_validos, key=lambda x: x.get('score', 0), reverse=True)[:50]

    def calculate_recommended_baja(self, similar_contratos, licitacion_anterior=None):
        """Calcular baja recomendada según las instrucciones de IA

        Args:
            similar_contratos: Lista de contratos similares
            licitacion_anterior: Licitación anterior de la misma administración (si existe)
        """
        # PRIORIDAD 1: Si hay licitación anterior de la misma administración, usar esa baja + 2%
        if licitacion_anterior and licitacion_anterior.get('baja_estadistica'):
            baja_anterior = licitacion_anterior['baja_estadistica']
            if baja_anterior and baja_anterior > 0:
                return min(baja_anterior + 2.0, 70.0)  # Limitado a 70%

        # PRIORIDAD 2: Cálculo normal si no hay licitación anterior
        if not similar_contratos or len(similar_contratos) == 0:
            return 15.0

        # Obtener todas las bajas válidas
        bajas = [c['baja_percentage'] for c in similar_contratos
                 if c.get('baja_percentage') is not None and c.get('baja_percentage') > 0]

        if not bajas:
            return 15.0

        # Buscar relación entre resultados según instrucciones
        if len(similar_contratos) >= 3:
            # Buscar si hay 3 o más empresas con bajas similares (±4%)
            grupos_similares = self._find_similar_baja_groups(bajas, tolerance=4.0)

            if grupos_similares:
                # Hay empresas con bajas similares: sumar 2% a la más alta
                grupo_mayor = max(grupos_similares, key=lambda g: len(g['bajas']))
                baja_mas_alta = max(grupo_mayor['bajas'])
                return min(baja_mas_alta + 2.0, 70.0)  # Limitado a 70%
            else:
                # Todas diferentes: hacer media y sumar 2%
                media = sum(bajas) / len(bajas)
                return min(media + 2.0, 70.0)
        else:
            # Menos de 3 contratos: hacer media simple
            media = sum(bajas) / len(bajas)
            return min(media + 2.0, 70.0)

    def _find_similar_baja_groups(self, bajas, tolerance=4.0):
        """Encontrar grupos de bajas similares con tolerancia dada (clustering correcto)"""
        if len(bajas) < 3:
            return []

        # Ordenar bajas para facilitar el clustering
        bajas_ordenadas = sorted(enumerate(bajas), key=lambda x: x[1])

        grupos = []
        bajas_procesadas = set()

        for i, (idx1, baja1) in enumerate(bajas_ordenadas):
            if idx1 in bajas_procesadas:
                continue

            # Iniciar un nuevo grupo
            grupo = {'bajas': [baja1], 'indices': [idx1]}
            bajas_procesadas.add(idx1)

            # Añadir bajas cercanas al rango del grupo (clustering por rango)
            for j, (idx2, baja2) in enumerate(bajas_ordenadas[i+1:], i+1):
                if idx2 in bajas_procesadas:
                    continue

                # Verificar si la baja está dentro del rango del grupo (min y max del grupo)
                min_grupo = min(grupo['bajas'])
                max_grupo = max(grupo['bajas'])

                # Si la nueva baja está dentro de tolerance de cualquier extremo del grupo
                if (baja2 - min_grupo <= tolerance) or (abs(baja2 - max_grupo) <= tolerance):
                    grupo['bajas'].append(baja2)
                    grupo['indices'].append(idx2)
                    bajas_procesadas.add(idx2)
                elif baja2 - max_grupo > tolerance:
                    # Si está muy lejos, parar de buscar (están ordenadas)
                    break

            if len(grupo['bajas']) >= 3:
                grupos.append(grupo)

        return grupos

    def get_empresa_stats(self, similar_contratos):
        """Obtener estadísticas de empresas participantes (3-7 empresas, sin None/vacíos)"""
        import json
        # Filtrar empresas válidas (sin None, vacíos, ni "NONE")
        empresas_validas = []
        for c in similar_contratos:
            if c.get('empresa'):
                empresa_raw = str(c['empresa']).strip()

                # Parsear JSON si es necesario
                if empresa_raw.startswith('['):
                    # Es un array de adjudicatarios (múltiples lotes)
                    try:
                        empresa_array = json.loads(empresa_raw)
                        for item in empresa_array:
                            if isinstance(item, dict) and 'adjudicatario' in item:
                                adj = item['adjudicatario']
                                if isinstance(adj, dict) and 'name' in adj:
                                    empresa = adj['name']
                                    if empresa and empresa.upper() not in ['NONE', 'NULL', 'N/A', ''] and len(empresa) > 3:
                                        empresas_validas.append(empresa)
                    except:
                        pass
                elif empresa_raw.startswith('{'):
                    # Es un objeto JSON simple
                    try:
                        empresa_data = json.loads(empresa_raw)
                        # Puede ser {"adjudicatario": {"name": ...}} o directamente {"name": ...}
                        if 'adjudicatario' in empresa_data and isinstance(empresa_data['adjudicatario'], dict):
                            empresa = empresa_data['adjudicatario'].get('name', empresa_raw)
                        elif 'name' in empresa_data:
                            empresa = empresa_data.get('name', empresa_raw)

                        if empresa and empresa.upper() not in ['NONE', 'NULL', 'N/A', ''] and len(empresa) > 3:
                            empresas_validas.append(empresa)
                    except:
                        pass
                else:
                    # Es texto simple
                    empresa = empresa_raw
                    if empresa and empresa.upper() not in ['NONE', 'NULL', 'N/A', ''] and len(empresa) > 3:
                        empresas_validas.append(empresa)

        if not empresas_validas:
            # No hay empresas válidas - devolver valores mínimos sin inventar datos
            bajas = [c['baja_percentage'] for c in similar_contratos if c['baja_percentage'] is not None and c['baja_percentage'] > 0]
            rango_bajas = (min(bajas), max(bajas)) if bajas else (10, 20)
            return [], 1, rango_bajas

        # Contar frecuencia de empresas reales
        empresa_counts = {}
        for empresa in empresas_validas:
            empresa_counts[empresa] = empresa_counts.get(empresa, 0) + 1

        # Obtener empresas más frecuentes (solo empresas reales del CPV)
        all_empresas = sorted(empresa_counts.items(), key=lambda x: x[1], reverse=True)

        # Seleccionar hasta 7 empresas reales (sin inventar datos)
        num_mostrar = min(7, len(all_empresas))
        top_empresas = all_empresas[:num_mostrar]

        participacion_promedio = max(3, min(7, len(set(empresas_validas))))

        # Rango de bajas
        bajas = [c['baja_percentage'] for c in similar_contratos if c['baja_percentage'] is not None and c['baja_percentage'] > 0]
        rango_bajas = (min(bajas), max(bajas)) if bajas else (10, 20)

        return top_empresas, participacion_promedio, rango_bajas

    def generate_baja_text(self, xml_data, similar_contratos, recommended_baja):
        """Generar texto de baja estadística siguiendo el formato del ejemplo"""

        # Obtener estadísticas
        top_empresas, participacion, rango_bajas = self.get_empresa_stats(similar_contratos)

        # Extraer criterios del XML
        criterios_xml = xml_data.get('criterios_adjudicacion', [])

        # Seleccionar frases aleatorias
        saludo = random.choice(self.saludos)
        despedida = random.choice(self.despedidas)

        # Construir texto siguiendo exactamente el formato del ejemplo
        texto = f"{saludo}\n"

        # Criterios de adjudicación (formato exacto del ejemplo)
        texto += "En la selección de expedientes, nos encontramos los siguientes criterios de adjudicación:\n"

        if criterios_xml and len(criterios_xml) > 0:
            for i, criterio in enumerate(criterios_xml):
                if isinstance(criterio, dict):
                    descripcion = criterio.get('descripcion', f'Criterio {i+1}')
                    peso = criterio.get('peso', '')
                    # Limpiar y formatear descripción
                    descripcion_clean = self._clean_criterio_text(descripcion)
                    peso_text = f": {peso}" if peso else ""
                    texto += f"{descripcion_clean.upper()}{peso_text}\n"
                else:
                    # Si es string simple
                    texto += f"{str(criterio).upper()}\n"
        else:
            # Criterios por defecto si no hay en el XML
            puntos_precio = random.choice([75, 80, 85])
            puntos_tecnico = 100 - puntos_precio
            texto += f"OFERTA ECONÓMICA: {puntos_precio} PUNTOS\n"
            texto += f"CRITERIOS TÉCNICOS: {puntos_tecnico} PUNTOS\n"

        # Análisis de expedientes (formato del ejemplo)
        texto += f"Al revisar expedientes previos de similar envergadura y presupuesto, hemos observado una participación promedio de {participacion} empresa"
        if participacion != 1:
            texto += "s"
        texto += ".\n"

        # Empresas más sobresalientes (formato del ejemplo)
        if top_empresas and len(top_empresas) >= 3:
            empresas_texto = ", ".join([emp[0] for emp in top_empresas[:3]])
            variaciones_intro = [
                f"Entre las empresas más sobresalientes en este campo están {empresas_texto}",
                f"Entre las compañías más destacadas en este ámbito encontramos {empresas_texto}",
                f"Las empresas con mayor presencia en este sector incluyen {empresas_texto}"
            ]
            texto += f" {random.choice(variaciones_intro)}.\n"
        elif top_empresas:
            # Menos de 3 empresas
            empresas_texto = ", ".join([emp[0] for emp in top_empresas])
            texto += f" La empresa más relevante en este campo es {empresas_texto}.\n"

        # Variaciones en ofertas (formato del ejemplo)
        if rango_bajas[0] > 0 and rango_bajas[1] > 0:
            variaciones_texto = [
                f"Notamos que las variaciones en las ofertas son notables, con un promedio de entre {rango_bajas[0]:.1f}% y  {rango_bajas[1]:.1f}%, lo que demuestra una estrategia de ofertas variada",
                f"Observamos que las diferencias en las propuestas económicas son significativas, con descuentos que oscilan entre {rango_bajas[0]:.1f}% y {rango_bajas[1]:.1f}%, evidenciando estrategias comerciales diversas",
                f"Las variaciones en las ofertas son considerables, registrándose un rango de descuentos desde {rango_bajas[0]:.1f}% hasta {rango_bajas[1]:.1f}%, reflejando un mercado competitivo"
            ]
            texto += random.choice(variaciones_texto) + ".\n"

        # Recomendación de baja (formato del ejemplo)
        variaciones_recomendacion = [
            f"Por ello, sugerimos una propuesta económica con un margen de descuento del {recommended_baja:.2f}%",
            f"En consecuencia, recomendamos una oferta económica con una baja del {recommended_baja:.2f}%",
            f"Por tanto, aconsejamos una propuesta con un descuento del {recommended_baja:.2f}%"
        ]
        texto += random.choice(variaciones_recomendacion) + ".\n"

        # Despedida
        texto += f"{despedida}"

        return texto

    def _clean_criterio_text(self, texto):
        """Limpiar y formatear texto de criterios eliminando contenido no aplicable"""
        if not texto:
            return "CRITERIO"

        texto_str = str(texto).strip()

        # Filtrar frases no aplicables o ruido común
        frases_excluir = [
            'sin descripción', 'no disponible', 'n/a', 'null', 'undefined',
            'ver documento', 'consultar', 'véase', 'según pliego',
            'criterio', 'punto', 'apartado', 'anexo'
        ]

        texto_lower = texto_str.lower()
        for frase in frases_excluir:
            if frase in texto_lower and len(texto_str) < 30:
                return "CRITERIO"

        # Eliminar números de apartados al inicio (ej: "1.", "1.1.", "a)")
        texto_limpio = re.sub(r'^[\d\.\)\s]+', '', texto_str)

        # Remover caracteres especiales excesivos pero mantener puntuación básica
        texto_limpio = re.sub(r'[^\w\sáéíóúüñÁÉÍÓÚÜÑ\-:%]', '', texto_limpio)

        # Eliminar espacios múltiples
        texto_limpio = re.sub(r'\s+', ' ', texto_limpio).strip()

        # Limitar longitud
        if len(texto_limpio) > 60:
            # Buscar punto de corte natural (coma, punto)
            corte = texto_limpio[:57].rfind(' ')
            if corte > 40:
                texto_limpio = texto_limpio[:corte] + "..."
            else:
                texto_limpio = texto_limpio[:57] + "..."

        return texto_limpio if texto_limpio and len(texto_limpio) > 3 else "CRITERIO"

    def create_excel_download(self, xml_data, similar_contratos, recommended_baja, texto_baja):
        """Crear Excel con todos los datos extraídos"""

        # Crear un nuevo workbook
        wb = Workbook()

        # Hoja 1: Resumen del análisis
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Análisis"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        sub_header_font = Font(bold=True, color="000000")
        sub_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        # Título
        ws_resumen['A1'] = "ANÁLISIS DE BAJA ESTADÍSTICA"
        ws_resumen['A1'].font = Font(bold=True, size=16)
        ws_resumen.merge_cells('A1:D1')

        # Información del contrato XML
        row = 3
        ws_resumen[f'A{row}'] = "DATOS DEL CONTRATO"
        ws_resumen[f'A{row}'].font = header_font
        ws_resumen[f'A{row}'].fill = header_fill
        ws_resumen.merge_cells(f'A{row}:D{row}')

        row += 1
        contract_data = [
            ("Título", xml_data.get('titulo', 'N/A')),
            ("Organismo", xml_data.get('organismo', 'N/A')),
            ("Presupuesto", f"€{xml_data.get('presupuesto', 0):,.2f}" if xml_data.get('presupuesto') else 'N/A'),
            ("Ubicación", xml_data.get('ubicacion', 'N/A')),
            ("CPV", xml_data.get('cpv', 'N/A')),
            ("Tipo de Procedimiento", xml_data.get('tipo_procedimiento', 'N/A'))
        ]

        for label, value in contract_data:
            ws_resumen[f'A{row}'] = label
            ws_resumen[f'A{row}'].font = sub_header_font
            ws_resumen[f'B{row}'] = str(value)
            row += 1

        # Criterios de adjudicación
        row += 1
        ws_resumen[f'A{row}'] = "CRITERIOS DE ADJUDICACIÓN"
        ws_resumen[f'A{row}'].font = header_font
        ws_resumen[f'A{row}'].fill = header_fill
        ws_resumen.merge_cells(f'A{row}:D{row}')

        row += 1
        criterios = xml_data.get('criterios_adjudicacion', [])
        if criterios:
            for i, criterio in enumerate(criterios, 1):
                if isinstance(criterio, dict):
                    desc = criterio.get('descripcion', f'Criterio {i}')
                    peso = criterio.get('peso', '')
                    criterio_text = f"{desc}" + (f" ({peso})" if peso else "")
                else:
                    criterio_text = str(criterio)

                ws_resumen[f'A{row}'] = f"{i}."
                ws_resumen[f'B{row}'] = criterio_text
                row += 1
        else:
            ws_resumen[f'A{row}'] = "No se encontraron criterios específicos"
            row += 1

        # Resultados del análisis
        row += 1
        ws_resumen[f'A{row}'] = "RESULTADOS DEL ANÁLISIS"
        ws_resumen[f'A{row}'].font = header_font
        ws_resumen[f'A{row}'].fill = header_fill
        ws_resumen.merge_cells(f'A{row}:D{row}')

        row += 1
        top_empresas, participacion, rango_bajas = self.get_empresa_stats(similar_contratos)

        analysis_data = [
            ("Baja Recomendada", f"{recommended_baja:.1f}%"),
            ("Contratos Analizados", len(similar_contratos)),
            ("Participación Media", f"{participacion} empresas"),
            ("Rango de Bajas", f"{rango_bajas[0]:.1f}% - {rango_bajas[1]:.1f}%")
        ]

        for label, value in analysis_data:
            ws_resumen[f'A{row}'] = label
            ws_resumen[f'A{row}'].font = sub_header_font
            ws_resumen[f'B{row}'] = str(value)
            row += 1

        # Empresas más activas
        if top_empresas:
            row += 1
            ws_resumen[f'A{row}'] = "EMPRESAS MÁS ACTIVAS"
            ws_resumen[f'A{row}'].font = header_font
            ws_resumen[f'A{row}'].fill = header_fill
            ws_resumen.merge_cells(f'A{row}:D{row}')

            row += 1
            for empresa, count in top_empresas[:7]:
                ws_resumen[f'A{row}'] = empresa
                ws_resumen[f'B{row}'] = f"{count} contratos"
                row += 1

        # Hoja 2: Contratos similares
        ws_contratos = wb.create_sheet("Contratos Similares")

        # Headers para contratos similares
        headers = ['Empresa', 'PBL (€)', 'Importe Adjudicación (€)', 'Baja (%)', 'Score', 'Núm. Licitadores', 'Objeto', 'Provincia']
        for col, header in enumerate(headers, 1):
            cell = ws_contratos.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Datos de contratos similares
        for row_idx, contrato in enumerate(similar_contratos, 2):
            ws_contratos.cell(row=row_idx, column=1, value=contrato.get('empresa', 'N/A'))
            ws_contratos.cell(row=row_idx, column=2, value=contrato.get('pbl', 0))
            ws_contratos.cell(row=row_idx, column=3, value=contrato.get('importe_adjudicacion', 0))
            ws_contratos.cell(row=row_idx, column=4, value=f"{contrato.get('baja_percentage', 0):.2f}%")
            ws_contratos.cell(row=row_idx, column=5, value=contrato.get('score', 0))
            ws_contratos.cell(row=row_idx, column=6, value=contrato.get('num_licitadores', 1))
            ws_contratos.cell(row=row_idx, column=7, value=contrato.get('objeto', 'N/A'))
            ws_contratos.cell(row=row_idx, column=8, value=contrato.get('provincia', 'N/A'))

        # Ajustar ancho de columnas
        from openpyxl.utils import get_column_letter
        for sheet in [ws_resumen, ws_contratos]:
            for col_idx in range(1, sheet.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)

                for row_idx in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    # Saltar celdas combinadas
                    if hasattr(cell, 'value') and cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass

                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Hoja 3: Texto generado
        ws_texto = wb.create_sheet("Informe Generado")
        ws_texto['A1'] = "INFORME DE BAJA ESTADÍSTICA"
        ws_texto['A1'].font = Font(bold=True, size=14)

        # Dividir el texto en líneas y ponerlo en celdas
        lineas = texto_baja.split('\n')
        for i, linea in enumerate(lineas, 3):
            ws_texto[f'A{i}'] = linea
            ws_texto[f'A{i}'].alignment = Alignment(wrap_text=True)

        # Ajustar ancho de la columna del texto
        ws_texto.column_dimensions['A'].width = 80

        # Guardar en memoria
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    def buscar_contratos_simples_por_cpv(self, cpv, presupuesto_min, presupuesto_max, limit=10):
        """Búsqueda simple y directa por CPV - replica el análisis manual"""
        import json

        # Extraer primeros 4 dígitos del CPV
        cpv_digits = ''.join(filter(str.isdigit, str(cpv)))[:4]

        if not cpv_digits or len(cpv_digits) < 4:
            return []

        query = f"""
        SELECT
            titulo,
            entidad_compradora as organismo,
            importe_total,
            importe_adjudicacion,
            adjudicatario,
            numero_licitadores,
            fecha_publicacion,
            ROUND(((importe_total - importe_adjudicacion) / NULLIF(importe_total, 0) * 100)::numeric, 2) as baja_estadistica,
            cpv,
            provincia
        FROM adjudicaciones_metabase
        WHERE importe_total IS NOT NULL
        AND importe_adjudicacion IS NOT NULL
        AND importe_total > 0
        AND importe_adjudicacion > 0
        AND importe_total != importe_adjudicacion
        AND cpv::text ~ '{cpv_digits}'
        AND importe_total BETWEEN {presupuesto_min} AND {presupuesto_max}
        ORDER BY fecha_publicacion DESC
        LIMIT {limit}
        """

        try:
            result = pd.read_sql(query, self.connection)

            # Convertir a lista de diccionarios con parsing de adjudicatarios
            contratos = []
            for _, row in result.iterrows():
                # Extraer empresa
                empresa = 'N/A'
                adj_raw = row['adjudicatario']
                if adj_raw:
                    try:
                        adj_str = str(adj_raw)
                        if adj_str.startswith('['):
                            adj_array = json.loads(adj_str)
                            if adj_array and len(adj_array) > 0:
                                if 'adjudicatario' in adj_array[0] and 'name' in adj_array[0]['adjudicatario']:
                                    empresa = adj_array[0]['adjudicatario']['name']
                        elif adj_str.startswith('{'):
                            adj_dict = json.loads(adj_str)
                            if 'adjudicatario' in adj_dict and isinstance(adj_dict['adjudicatario'], dict):
                                empresa = adj_dict['adjudicatario'].get('name', 'N/A')
                        else:
                            empresa = adj_str[:60] if len(adj_str) > 0 else 'N/A'
                    except:
                        empresa = str(adj_raw)[:60] if adj_raw else 'N/A'

                contrato = {
                    'titulo': row['titulo'],
                    'organismo': row['organismo'],
                    'presupuesto_licitacion': float(row['importe_total']),
                    'precio_adjudicacion': float(row['importe_adjudicacion']),
                    'empresa': empresa,
                    'empresa_adjudicataria': empresa,
                    'num_licitadores': int(row['numero_licitadores']) if row['numero_licitadores'] else 0,
                    'fecha_publicacion': str(row['fecha_publicacion']),
                    'baja_percentage': float(row['baja_estadistica']),
                    'cpv': row['cpv'],
                    'provincia': row['provincia']
                }
                contratos.append(contrato)

            return contratos
        except Exception as e:
            st.error(f"Error en búsqueda: {e}")
            return []

def main():
    st.title("📊 Analizador de Bajas Estadísticas - XML a Base de Datos")
    st.sidebar.title("Configuración")

    # Inicializar generador
    if 'generator' not in st.session_state:
        st.session_state.generator = BajaEstadisticaGenerator()

    generator = st.session_state.generator

    # Conectar a la base de datos
    if not generator.connection:
        with st.spinner("Conectando a la base de datos..."):
            if generator.connect_to_database():
                st.success("✅ Conectado a la base de datos oclemconcursos")
            else:
                st.error("❌ No se pudo conectar a la base de datos")
                return

    st.markdown("### 🔗 Análisis de Contrato desde XML o JSON")

    # Selector de tipo de fuente
    source_type = st.radio(
        "Selecciona el tipo de fuente:",
        options=["XML (URL)", "JSON (Archivo)"],
        index=0,
        help="Elige si quieres analizar desde una URL de XML o subir un archivo JSON"
    )

    if source_type == "XML (URL)":
        # Input para URL del XML
        xml_url = st.text_input(
            "Introduce la URL del XML del contrato:",
            placeholder="https://contrataciondelestado.es/FileSystem/servlet/GetDocumentByIdServlet?DocumentIdParam=...",
            help="Pega aquí el enlace completo del XML del contrato de la plataforma de contratación del estado"
        )
        json_file = None
    else:
        # File uploader para JSON
        json_file = st.file_uploader(
            "Sube el archivo JSON de la licitación:",
            type=['json'],
            help="Selecciona un archivo JSON que contenga los datos de la licitación"
        )
        xml_url = None

    # Campo opcional para número de lote
    numero_lote = st.text_input(
        "Número de lote (opcional):",
        placeholder="Ej: 1, 2, 3...",
        help="Si la licitación está dividida en lotes y solo quieres analizar uno específico, indica su número. Déjalo vacío para analizar toda la licitación."
    )

    if st.button("🚀 Analizar Contrato", type="primary"):
        datos_contrato = None
        source_name = "XML" if source_type == "XML (URL)" else "JSON"

        if source_type == "XML (URL)" and xml_url:
            with st.spinner("Descargando y analizando XML..."):
                # Extraer datos del XML
                datos_contrato = generator.extract_xml_data(xml_url, numero_lote if numero_lote else None)

        elif source_type == "JSON (Archivo)" and json_file:
            with st.spinner("Procesando archivo JSON..."):
                try:
                    # Leer el archivo JSON
                    json_content = json_file.read().decode('utf-8')
                    # Extraer datos del JSON
                    datos_contrato = generator.extract_json_data(json_content, numero_lote if numero_lote else None)
                except Exception as e:
                    st.error(f"Error leyendo archivo JSON: {e}")
                    datos_contrato = None

        if datos_contrato:
            st.success(f"✅ {source_name} procesado correctamente")

            # Mostrar datos extraídos
            st.markdown("### 📋 Datos del Contrato Extraídos")
            # Información del contrato en un expander
            with st.expander("📄 Ver información del contrato", expanded=False):
                col1, col2 = st.columns(2)

                with col1:
                    st.markdown("**📄 Información Básica**")
                    st.write(f"**Título:** {datos_contrato['titulo'][:100]}..." if len(datos_contrato['titulo']) > 100 else f"**Título:** {datos_contrato['titulo']}")
                    st.write(f"**Organismo:** {datos_contrato['organismo']}")
                    st.write(f"**Presupuesto:** €{datos_contrato['presupuesto']:,.2f}")
                    st.write(f"**CPV:** {datos_contrato['cpv']}")

                with col2:
                    st.markdown("**⚖️ Procedimiento y Criterios**")
                    st.write(f"**Tipo:** {datos_contrato['tipo_procedimiento']}")
                    if datos_contrato['criterios_adjudicacion']:
                        st.write(f"**Criterios de adjudicación ({len(datos_contrato['criterios_adjudicacion'])}):**")

                        # Mostrar todos los criterios sin límite
                        for i, criterio in enumerate(datos_contrato['criterios_adjudicacion'], 1):
                            if isinstance(criterio, dict):
                                # Si es un diccionario, mostrar la descripción
                                desc = criterio.get('descripcion', 'Sin descripción')
                                peso = criterio.get('peso', '')
                                otros = criterio.get('otros', [])

                                # Crear texto del criterio
                                criterio_texto = desc
                                if peso:
                                    criterio_texto += f" (Peso: {peso})"
                                if otros:
                                    criterio_texto += f" - {', '.join(otros[:2])}"  # Máximo 2 elementos adicionales

                                # Mostrar con longitud variable según contenido
                                max_len = 120 if peso or otros else 80
                                if len(criterio_texto) > max_len:
                                    st.write(f"{i}. {criterio_texto[:max_len]}...")
                                else:
                                    st.write(f"{i}. {criterio_texto}")
                            else:
                                # Si es una cadena simple
                                criterio_str = str(criterio)
                                if len(criterio_str) > 80:
                                    st.write(f"{i}. {criterio_str[:80]}...")
                                else:
                                    st.write(f"{i}. {criterio_str}")

                        # Si hay muchos criterios, mostrar un resumen
                        if len(datos_contrato['criterios_adjudicacion']) > 10:
                            st.info(f"💡 Se encontraron {len(datos_contrato['criterios_adjudicacion'])} criterios de adjudicación en total")

            # Buscar contratos similares en la base de datos
            with st.spinner("Buscando contratos similares en la base de datos..."):
                # Extraer criterios de búsqueda del contrato
                provincia = datos_contrato.get('ubicacion', '')
                presupuesto = datos_contrato.get('presupuesto', 0)
                years = [2025, 2024, 2023, 2022]

                # ETAPA 1: Intentar búsqueda con CPV exacto (8 dígitos)
                cpv_full = generator._extract_cpv_full(datos_contrato.get('cpv', ''))

                if cpv_full:
                    st.info(f"🔍 Etapa 1: Buscando con CPV exacto ({cpv_full})...")
                    st.session_state.contratos_data = generator.get_filtered_contratos_data(
                        cpv_category=cpv_full,  # 8 dígitos exactos
                        provincia=provincia,
                        presupuesto=presupuesto,
                        years=years,
                        limit=50
                    )

                    data = st.session_state.contratos_data
                    st.info(f"✅ Encontrados {len(data)} contratos con CPV exacto")
                else:
                    # Si no hay CPV completo, ir directamente a búsqueda amplia
                    data = pd.DataFrame()

                # Análisis de competencia
                st.markdown("### 🔍 Análisis de Competencia")

                # Función auxiliar para procesar contratos
                def process_contratos(data_df):
                    contratos_list = []
                    for idx, row in data_df.iterrows():
                        # Extraer datos directamente de las columnas SQL
                        baja_estadistica = row.get('baja_estadistica', 0)
                        empresa_raw = row.get('empresa_adjudicataria', '')

                        # Extraer solo el nombre si viene en formato JSON
                        empresa = empresa_raw
                        if empresa_raw and isinstance(empresa_raw, str):
                            try:
                                # Si es un JSON, extraer el campo "name"
                                if empresa_raw.strip().startswith('{'):
                                    import json
                                    empresa_data = json.loads(empresa_raw)
                                    # Puede ser {"adjudicatario": {"name": ...}} o directamente {"name": ...}
                                    if 'adjudicatario' in empresa_data and isinstance(empresa_data['adjudicatario'], dict):
                                        empresa = empresa_data['adjudicatario'].get('name', empresa_raw)
                                    elif 'name' in empresa_data:
                                        empresa = empresa_data.get('name', empresa_raw)
                            except:
                                # Si falla el parseo, usar el valor original
                                empresa = empresa_raw

                        # Filtrar: eliminar bajas > 70% o < 0.5% o sin empresa
                        if baja_estadistica and 0.5 <= baja_estadistica <= 70 and empresa:
                            contrato_data = {
                                'titulo': row.get('titulo', ''),
                                'organismo': row.get('organismo', ''),
                                'presupuesto_licitacion': row.get('presupuesto_licitacion', 0),
                                'precio_adjudicacion': row.get('precio_adjudicacion', 0),
                                'baja_percentage': baja_estadistica,
                                'empresa_adjudicataria': empresa,
                                'num_licitadores': row.get('num_licitadores', 1) or 1,
                                'fecha_publicacion': str(row.get('fecha_publicacion', '')),
                                'cpv': row.get('cpv', ''),
                                'provincia': row.get('provincia', ''),
                                'objeto': row.get('objeto', ''),
                                'pbl': row.get('presupuesto_licitacion', 0),
                                'importe_adjudicacion': row.get('precio_adjudicacion', 0),
                                'precio': row.get('presupuesto_licitacion', 0),
                                'empresa': empresa
                            }
                            contratos_list.append(contrato_data)
                    return contratos_list

                # Convertir DataFrame a lista de diccionarios para procesamiento
                similar_contratos = process_contratos(data)

                st.write(f"**DEBUG: Contratos procesados (Etapa 1): {len(similar_contratos)} de {len(data)} encontrados**")

                # ETAPA 2: Si hay menos de 7 contratos procesados, ampliar búsqueda a 4 dígitos
                if len(similar_contratos) < 7:
                    st.warning(f"⚠️ Solo se encontraron {len(similar_contratos)} contratos con CPV exacto (se requieren al menos 7)")
                    st.info("🔄 Etapa 2: Ampliando búsqueda a primeros 4 dígitos del CPV...")

                    # Extraer primeros 4 dígitos del CPV
                    cpv_category = generator._extract_cpv_category_from_multiple(
                        datos_contrato.get('cpv', ''),
                        datos_contrato.get('objeto', '')
                    )

                    # Buscar con CPV amplio (4 dígitos)
                    st.session_state.contratos_data = generator.get_filtered_contratos_data(
                        cpv_category=cpv_category,  # 4 dígitos
                        provincia=provincia,
                        presupuesto=presupuesto,
                        years=years,
                        limit=50
                    )

                    data = st.session_state.contratos_data
                    st.info(f"✅ Encontrados {len(data)} contratos con primeros 4 dígitos del CPV")

                    # Re-procesar con los nuevos datos
                    similar_contratos = process_contratos(data)
                    st.write(f"**DEBUG: Contratos procesados (Etapa 2): {len(similar_contratos)} de {len(data)} encontrados**")
                else:
                    st.success(f"✅ Suficientes contratos encontrados con CPV exacto ({len(similar_contratos)} contratos)")

                st.write(f"**DEBUG FINAL: Total de contratos procesados: {len(similar_contratos)}**")

                # Generar informe siempre que haya al menos 1 contrato
                if len(similar_contratos) > 0:
                    # BUSCAR LICITACIÓN ANTERIOR DE LA MISMA ADMINISTRACIÓN
                    organismo = datos_contrato.get('organismo', '')
                    cpv_full = generator._extract_cpv_full(datos_contrato.get('cpv', ''))
                    if not cpv_full:
                        cpv_full = generator._extract_cpv_category_from_multiple(
                            datos_contrato.get('cpv', ''),
                            datos_contrato.get('objeto', '')
                        )

                    licitacion_anterior = None
                    if organismo and cpv_full:
                        with st.spinner("🔍 Buscando licitaciones anteriores de la misma administración..."):
                            licitacion_anterior = generator.search_previous_licitacion_same_org(
                                organismo=organismo,
                                cpv_category=cpv_full,
                                presupuesto=presupuesto
                            )

                        if licitacion_anterior:
                            st.success("✅ ¡Encontrada licitación anterior de la misma administración!")

                            # Mostrar información destacada
                            st.markdown("### 🎯 LICITACIÓN ANTERIOR RELEVANTE")

                            # Parsear empresa_adjudicataria si es JSON
                            import json
                            empresa_anterior_raw = licitacion_anterior.get('empresa_adjudicataria', 'N/A')
                            empresa_anterior = empresa_anterior_raw
                            if empresa_anterior_raw and isinstance(empresa_anterior_raw, str):
                                try:
                                    if empresa_anterior_raw.strip().startswith('{'):
                                        empresa_data = json.loads(empresa_anterior_raw)
                                        if 'adjudicatario' in empresa_data and isinstance(empresa_data['adjudicatario'], dict):
                                            empresa_anterior = empresa_data['adjudicatario'].get('name', empresa_anterior_raw)
                                        elif 'name' in empresa_data:
                                            empresa_anterior = empresa_data.get('name', empresa_anterior_raw)
                                except:
                                    pass

                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("📅 Fecha", str(licitacion_anterior.get('fecha_publicacion', 'N/A'))[:10])
                            with col2:
                                st.metric("💰 Presupuesto", f"{licitacion_anterior.get('presupuesto_licitacion', 0):,.0f}€")
                            with col3:
                                st.metric("📉 Baja Anterior", f"{licitacion_anterior.get('baja_estadistica', 0):.2f}%")

                            st.info(f"🏆 **Adjudicatario anterior:** {empresa_anterior}")
                            st.warning(f"💡 **Recomendación basada en adjudicación anterior:** {licitacion_anterior.get('baja_estadistica', 0):.2f}% + 2% = **{licitacion_anterior.get('baja_estadistica', 0) + 2:.2f}%**")

                    # Calcular baja recomendada (priorizando licitación anterior si existe)
                    recommended_baja = generator.calculate_recommended_baja(similar_contratos, licitacion_anterior)

                    # Mostrar resultados principales
                    st.markdown("### 📊 Resultados del Análisis")

                    st.markdown("---")

                    # Información del contrato
                    st.markdown("#### 📋 Datos del Contrato")
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Organismo:** {datos_contrato.get('organismo', 'N/A')}")
                        st.write(f"**Ubicación:** {datos_contrato.get('ubicacion', 'N/A')}")
                    with col2:
                        presupuesto = datos_contrato.get('presupuesto', 0)
                        if presupuesto:
                            st.write(f"**Presupuesto:** €{presupuesto:,.2f}")
                        st.write(f"**CPV:** {datos_contrato.get('cpv', 'N/A')}")

                    st.markdown("---")

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("🎯 Baja Recomendada", f"{recommended_baja:.1f}%")
                    with col2:
                        st.metric("📈 Contratos Analizados", len(similar_contratos))
                    with col3:
                        # Verificar si similar_contratos es DataFrame o lista
                        if hasattr(similar_contratos, 'columns') and 'num_licitadores' in similar_contratos.columns:
                            avg_competitors = similar_contratos['num_licitadores'].mean()
                        elif isinstance(similar_contratos, list):
                            # Si es una lista de diccionarios, calcular promedio manualmente
                            licitadores = [c.get('num_licitadores', 0) for c in similar_contratos if isinstance(c, dict) and c.get('num_licitadores')]
                            avg_competitors = sum(licitadores) / len(licitadores) if licitadores else 0
                        else:
                            avg_competitors = 0
                        st.metric("👥 Competidores Promedio", f"{avg_competitors:.0f}")

                    st.markdown("---")

                    # Criterios de adjudicación - siempre visible
                    st.markdown("#### ⚖️ Criterios de Adjudicación")
                    if datos_contrato['criterios_adjudicacion']:
                        for i, criterio in enumerate(datos_contrato['criterios_adjudicacion'], 1):
                            if isinstance(criterio, dict):
                                desc = criterio.get('descripcion', f'Criterio {i}')
                                peso = criterio.get('peso', '')
                                st.write(f"**{i}.** {desc.upper()}: **{peso}**" if peso else f"**{i}.** {desc.upper()}")
                            else:
                                st.write(f"**{i}.** {str(criterio).upper()}")
                    else:
                        st.info(f"No se pudieron extraer criterios específicos del {source_name}")

                    st.markdown("---")

                    # Análisis de competencia histórica - siempre visible
                    st.markdown("#### 🏆 Análisis de Mercado")
                    top_empresas, participacion, rango_bajas = generator.get_empresa_stats(similar_contratos)

                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("**📈 Estadísticas de Competencia:**")
                        st.write(f"• Participación promedio: **{participacion} empresas** por licitación")
                        if rango_bajas[0] > 0:
                            st.write(f"• Rango de bajas: **{rango_bajas[0]:.1f}% - {rango_bajas[1]:.1f}%**")
                            st.write(f"• Baja media: **{(rango_bajas[0] + rango_bajas[1])/2:.1f}%**")

                    with col2:
                        if top_empresas:
                            st.markdown("**🏢 Empresas Más Activas:**")
                            for empresa, count in top_empresas[:5]:
                                st.write(f"• {empresa} ({count})")
                        else:
                            st.info("No se encontraron empresas en los contratos similares")

                    # Generar texto justificativo
                    with st.spinner("Generando informe de baja estadística..."):
                        texto_baja = generator.generate_baja_text(datos_contrato, similar_contratos, recommended_baja)

                    st.markdown("### 📝 Informe de Baja Estadística")
                    st.text_area(
                        "Texto generado para justificar la baja:",
                        texto_baja,
                        height=400,
                        help="Puedes copiar este texto para incluirlo en tu oferta"
                    )

                    # Botón para regenerar
                    if st.button("🔄 Regenerar texto con diferente redacción"):
                        nuevo_texto = generator.generate_baja_text(datos_contrato, similar_contratos, recommended_baja)
                        st.text_area("Nuevo texto generado:", nuevo_texto, height=400)

                    # Sección de descarga Excel (destacada)
                    st.markdown("---")
                    st.markdown("### 📥 Descargar Resultados")

                    col1, col2 = st.columns([2, 1])

                    with col1:
                        # Generar Excel automáticamente para descarga
                        with st.spinner("Preparando Excel..."):
                            excel_buffer = generator.create_excel_download(
                                datos_contrato, similar_contratos, recommended_baja, texto_baja
                            )

                            # Generar nombre de archivo con fecha
                            fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
                            organismo_clean = re.sub(r'[^\w\s-]', '', datos_contrato.get('organismo', 'Analisis'))[:20]
                            nombre_archivo = f"Analisis_Baja_{organismo_clean}_{fecha_actual}.xlsx"

                            st.download_button(
                                label="📊 Descargar análisis completo en Excel",
                                data=excel_buffer.getvalue(),
                                file_name=nombre_archivo,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                help="Descarga un Excel con todos los datos extraídos, contratos similares y análisis completo",
                                key=f"download_excel_{fecha_actual}",
                                use_container_width=True
                            )

                    with col2:
                        st.info(f"✅ {len(similar_contratos)} contratos incluidos")

                    st.markdown("---")

                    # Mostrar contratos similares de forma detallada
                    st.markdown(f"#### 📋 Contratos Similares Encontrados ({len(similar_contratos)})")

                    # Preparar datos para mostrar
                    import json as json_module

                    for i, contrato in enumerate(similar_contratos[:10], 1):  # Mostrar los primeros 10
                        with st.container():
                            col1, col2 = st.columns([3, 1])

                            with col1:
                                titulo = contrato.get('titulo', 'Sin título')
                                st.markdown(f"**{i}. {titulo[:80]}{'...' if len(titulo) > 80 else ''}**")
                                st.write(f"📍 **Organismo:** {contrato.get('organismo', 'N/A')}")

                                # Extraer empresa adjudicataria
                                empresa_raw = contrato.get('empresa', '')
                                empresa = 'N/A'
                                if empresa_raw:
                                    try:
                                        if str(empresa_raw).startswith('['):
                                            emp_array = json_module.loads(str(empresa_raw))
                                            if emp_array and len(emp_array) > 0 and 'adjudicatario' in emp_array[0]:
                                                empresa = emp_array[0]['adjudicatario'].get('name', 'N/A')
                                        elif str(empresa_raw).startswith('{'):
                                            emp_dict = json_module.loads(str(empresa_raw))
                                            if 'adjudicatario' in emp_dict and isinstance(emp_dict['adjudicatario'], dict):
                                                empresa = emp_dict['adjudicatario'].get('name', 'N/A')
                                        else:
                                            empresa = str(empresa_raw)[:60]
                                    except:
                                        empresa = str(empresa_raw)[:60] if empresa_raw else 'N/A'

                                if empresa and empresa != 'N/A' and len(empresa) > 3:
                                    st.write(f"🏢 **Adjudicatario:** {empresa}")

                            with col2:
                                presupuesto = contrato.get('presupuesto_licitacion', 0)
                                adjudicacion = contrato.get('precio_adjudicacion', 0)
                                baja = contrato.get('baja_percentage', 0)
                                num_lic = contrato.get('num_licitadores', 0)

                                st.write(f"💰 **Presupuesto:** €{presupuesto:,.2f}")
                                st.write(f"💵 **Adjudicación:** €{adjudicacion:,.2f}")
                                st.write(f"📉 **Baja:** {baja:.2f}%")
                                st.write(f"👥 **Licitadores:** {num_lic}")

                                fecha = contrato.get('fecha_publicacion', 'N/A')
                                if fecha != 'N/A':
                                    fecha_str = str(fecha)[:10] if len(str(fecha)) > 10 else str(fecha)
                                    st.write(f"📅 {fecha_str}")

                            st.divider()

                    if len(similar_contratos) > 10:
                        with st.expander(f"Ver los {len(similar_contratos) - 10} contratos restantes"):
                            for i, contrato in enumerate(similar_contratos[10:], 11):
                                st.write(f"**{i}. {contrato.get('titulo', 'Sin título')[:80]}**")
                                st.write(f"   • Baja: {contrato.get('baja_percentage', 0):.1f}% | Presupuesto: €{contrato.get('presupuesto_licitacion', 0):,.2f}")
                                st.write("---")

                else:
                    st.warning("⚠️ No se encontraron suficientes contratos similares en la base de datos para realizar el análisis.")
                    st.info("Esto puede deberse a que el contrato es muy específico o los datos no coinciden con los registros de la base de datos.")
        else:
            if source_type == "XML (URL)" and not xml_url:
                st.info("👆 Introduce la URL del XML del contrato para comenzar el análisis.")
            elif source_type == "JSON (Archivo)" and not json_file:
                st.info("👆 Sube un archivo JSON de la licitación para comenzar el análisis.")
            else:
                st.error(f"❌ No se pudo procesar el {source_name}. Verifica que los datos sean correctos y estén accesibles.")

if __name__ == "__main__":
    main()