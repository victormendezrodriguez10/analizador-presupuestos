import streamlit as st
import psycopg2
import pandas as pd
import json
import requests
import xml.etree.ElementTree as ET
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="Análisis de Bajas Estadísticas", page_icon="📊", layout="wide")

# Conexión a base de datos
def get_connection():
    """Crear nueva conexión a la base de datos"""
    try:
        conn = psycopg2.connect(
            host=st.secrets["postgres"]["host"],
            database=st.secrets["postgres"]["database"],
            user=st.secrets["postgres"]["user"],
            password=st.secrets["postgres"]["password"],
            port=st.secrets["postgres"]["port"]
        )
        return conn
    except Exception as e:
        st.error(f"Error al conectar con la base de datos: {e}")
        return None

def extraer_datos_xml_completo(url):
    """Extraer datos completos del XML incluyendo lotes"""
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        root = ET.fromstring(response.content)

        # Datos generales
        datos = {
            'titulo': '',
            'organismo': '',
            'ubicacion': '',
            'lotes': []
        }

        # Buscar título (intentar varios patrones)
        # 1. Primero buscar en ProcurementProject > Name (el más confiable)
        for elem in root.iter():
            if 'ProcurementProject' in elem.tag:
                for child in elem:
                    if 'Name' in child.tag and child.text:
                        texto = child.text.strip()
                        if len(texto) > 20:
                            datos['titulo'] = texto
                            break
                if datos['titulo']:
                    break

        # 2. Si no, buscar en Title tags
        if not datos['titulo']:
            for elem in root.iter():
                if 'Title' in elem.tag or 'titulo' in elem.tag:
                    if elem.text and len(elem.text.strip()) > 10:
                        datos['titulo'] = elem.text.strip()
                        break

        # Si no se encontró título, buscar en IDs descriptivos primero
        if not datos['titulo']:
            for elem in root.iter():
                if 'ID' in elem.tag and elem.text:
                    texto = elem.text.strip()
                    # Buscar IDs descriptivos (ej: "PCAP CONTRATO PERGOLA FOTOVOLTAICA")
                    if len(texto) > 20 and not texto.startswith('ES') and any(c.isalpha() for c in texto):
                        # Limpiar el texto del ID
                        titulo = texto.replace('.pdf', '').replace('.doc', '').replace('PCAP', '').replace('PPT', '').strip()
                        # Validar que sea un título válido (no cargos, personas, etc.)
                        titulo_lower = titulo.lower()
                        if (len(titulo) > 15 and
                            'http' not in titulo_lower and
                            'alcalde' not in titulo_lower and
                            'presidente' not in titulo_lower and
                            'concejal' not in titulo_lower and
                            'secretari' not in titulo_lower):
                            datos['titulo'] = titulo
                            break

        # Si aún no hay título, buscar en Name o Description largos
        if not datos['titulo']:
            for elem in root.iter():
                tag_lower = elem.tag.lower()
                # Excluir tags URI/URL
                if ('name' in tag_lower or 'description' in tag_lower) and 'uri' not in tag_lower and 'url' not in tag_lower:
                    if elem.text and len(elem.text.strip()) > 20:
                        # Evitar nombres de organismos, personas, URLs, etc.
                        texto = elem.text.strip()
                        texto_lower = texto.lower()
                        # Rechazar solo si empieza con estos términos o si es SOLO eso
                        es_valido = True
                        for termino in ['alcalde', 'presidente', 'concejal', 'secretari']:
                            if texto_lower.startswith(termino) or texto_lower == termino:
                                es_valido = False
                                break
                        # Rechazar URLs y archivos
                        if any(x in texto_lower for x in ['.pdf', '.doc', 'http://', 'https://']):
                            es_valido = False

                        if es_valido:
                            datos['titulo'] = texto
                            break

        # Buscar organismo
        for elem in root.iter():
            if 'BuyerProfileURIID' in elem.tag or 'PartyName' in elem.tag or 'Name' in elem.tag:
                if elem.text and len(elem.text.strip()) > 5 and 'http' not in elem.text:
                    datos['organismo'] = elem.text.strip()
                    break

        # Buscar ubicación
        for elem in root.iter():
            if 'CityName' in elem.tag:
                if elem.text:
                    datos['ubicacion'] = elem.text.strip()
                    break

        # Buscar lotes (solo ProcurementProjectLot, no TenderingProcess que es metadata del proceso)
        for elem in root.iter():
            if 'ProcurementProjectLot' in elem.tag:
                lote = {
                    'numero': '',
                    'titulo': '',
                    'presupuesto': 0,
                    'cpv': [],
                    'criterios': []
                }

                # Buscar ID del lote
                for child in elem.iter():
                    if 'ID' in child.tag and child.text:
                        if child.text.strip() and not child.text.startswith('ES'):
                            lote['numero'] = child.text.strip()
                            break

                # Buscar título del lote
                for child in elem.iter():
                    if 'Title' in child.tag and child.text:
                        if len(child.text.strip()) > 10:
                            lote['titulo'] = child.text.strip()
                            break

                # Si no se encontró título del lote, buscar en Name o Description
                if not lote['titulo']:
                    for child in elem.iter():
                        tag_lower = child.tag.lower()
                        # Excluir tags URI/URL
                        if ('name' in tag_lower or 'description' in tag_lower) and 'uri' not in tag_lower and 'url' not in tag_lower:
                            if child.text and len(child.text.strip()) > 20:
                                texto = child.text.strip()
                                texto_lower = texto.lower()
                                # Rechazar solo si empieza con estos términos o si es SOLO eso
                                es_valido = True
                                for termino in ['alcalde', 'presidente', 'concejal', 'secretari']:
                                    if texto_lower.startswith(termino) or texto_lower == termino:
                                        es_valido = False
                                        break
                                # Rechazar URLs y archivos
                                if any(x in texto_lower for x in ['.pdf', '.doc', 'http://', 'https://']):
                                    es_valido = False

                                if es_valido:
                                    lote['titulo'] = texto
                                    break

                # Buscar presupuesto
                for child in elem.iter():
                    if 'EstimatedOverallContractAmount' in child.tag or 'TotalAmount' in child.tag:
                        if child.text:
                            try:
                                lote['presupuesto'] = float(child.text.strip())
                                break
                            except:
                                pass

                # Buscar CPVs
                for child in elem.iter():
                    if 'ItemClassificationCode' in child.tag:
                        cpv_text = child.text
                        if cpv_text:
                            cpv_digits = ''.join(filter(str.isdigit, cpv_text))
                            if len(cpv_digits) >= 4:
                                lote['cpv'].append(cpv_digits)

                # Buscar criterios con búsqueda más flexible
                for child in elem.iter():
                    tag_lower = child.tag.lower()
                    # Buscar tanto AwardingCriteria como AwardingCriterion y sus variantes
                    if 'criteria' in tag_lower or 'criterion' in tag_lower or 'criterio' in tag_lower:
                        criterio = {}

                        # Buscar en hijos directos (no recursivo)
                        for subchild in child:
                            subtag = subchild.tag.lower()
                            if subchild.text:
                                # Buscar descripción
                                if any(x in subtag for x in ['description', 'name', 'descripcion', 'nombre']):
                                    if not criterio.get('descripcion'):
                                        criterio['descripcion'] = subchild.text.strip()
                                # Buscar peso/puntuación
                                elif any(x in subtag for x in ['weight', 'numeric', 'peso', 'punto', 'puntuacion']):
                                    if not criterio.get('peso'):
                                        criterio['peso'] = subchild.text.strip()

                        if criterio.get('descripcion'):
                            # Evitar duplicados
                            if criterio not in lote['criterios']:
                                lote['criterios'].append(criterio)

                # Solo agregar si tiene datos relevantes
                if lote['presupuesto'] > 0 or lote['cpv'] or lote['titulo']:
                    if not lote['numero']:
                        lote['numero'] = str(len(datos['lotes']) + 1)
                    datos['lotes'].append(lote)

        # Si no se encontraron lotes, buscar datos a nivel de contrato general
        if not datos['lotes']:
            # Intentar obtener un título descriptivo (verificar que no sea cargo/persona)
            titulo_valido = False
            if datos['titulo']:
                titulo_lower = datos['titulo'].lower()
                if (not any(x in titulo_lower for x in ['http', 'alcalde', 'presidente', 'concejal', 'secretari'])):
                    titulo_valido = True

            titulo_general = datos['titulo'] if titulo_valido else ''

            # Si no hay título válido, buscar en elementos ID que contengan información descriptiva
            if not titulo_general:
                for elem in root.iter():
                    if 'ID' in elem.tag and elem.text:
                        texto = elem.text.strip()
                        # Buscar IDs descriptivos (ej: "PCAP CONTRATO PERGOLA FOTOVOLTAICA")
                        if len(texto) > 20 and not texto.startswith('ES') and any(c.isalpha() for c in texto):
                            # Limpiar el texto del ID
                            titulo_general = texto.replace('.pdf', '').replace('.doc', '').replace('PCAP', '').replace('PPT', '').strip()
                            # Validar que sea un título válido (no cargos, personas, etc.)
                            titulo_lower = titulo_general.lower()
                            if (len(titulo_general) > 15 and
                                'http' not in titulo_lower and
                                'alcalde' not in titulo_lower and
                                'presidente' not in titulo_lower and
                                'concejal' not in titulo_lower and
                                'secretari' not in titulo_lower):
                                break

            if not titulo_general:
                titulo_general = 'Contrato único'

            lote_general = {
                'numero': '1',
                'titulo': titulo_general,
                'presupuesto': 0,
                'cpv': [],
                'criterios': []
            }

            # Buscar presupuesto general (priorizar EstimatedOverallContractAmount)
            presupuestos_encontrados = []
            for elem in root.iter():
                if elem.text and elem.text.strip():
                    try:
                        valor = float(elem.text.strip())
                        if 'EstimatedOverallContractAmount' in elem.tag:
                            lote_general['presupuesto'] = valor
                            break
                        elif 'TotalAmount' in elem.tag or 'BudgetAmount' in elem.tag:
                            presupuestos_encontrados.append(valor)
                    except:
                        pass

            # Si no se encontró EstimatedOverallContractAmount, usar el máximo de los otros
            if lote_general['presupuesto'] == 0 and presupuestos_encontrados:
                lote_general['presupuesto'] = max(presupuestos_encontrados)

            # Buscar CPVs generales
            for elem in root.iter():
                if 'ItemClassificationCode' in elem.tag or 'CPV' in elem.tag:
                    cpv_text = elem.get('CODE') or elem.text
                    if cpv_text:
                        cpv_digits = ''.join(filter(str.isdigit, cpv_text))
                        if len(cpv_digits) >= 4 and cpv_digits not in lote_general['cpv']:
                            lote_general['cpv'].append(cpv_digits)

            # Buscar criterios generales con búsqueda más flexible
            for elem in root.iter():
                tag_lower = elem.tag.lower()
                # Buscar tanto AwardingCriteria como AwardingCriterion y sus variantes
                if 'criteria' in tag_lower or 'criterion' in tag_lower or 'criterio' in tag_lower:
                    criterio = {}

                    # Buscar en hijos directos (no recursivo)
                    for child in elem:
                        childtag = child.tag.lower()
                        if child.text:
                            # Buscar descripción
                            if any(x in childtag for x in ['description', 'name', 'descripcion', 'nombre']):
                                if not criterio.get('descripcion'):
                                    criterio['descripcion'] = child.text.strip()
                            # Buscar peso/puntuación
                            elif any(x in childtag for x in ['weight', 'numeric', 'peso', 'punto', 'puntuacion']):
                                if not criterio.get('peso'):
                                    criterio['peso'] = child.text.strip()

                    if criterio.get('descripcion') and criterio not in lote_general['criterios']:
                        lote_general['criterios'].append(criterio)

            # Agregar lote general si tiene datos útiles
            if lote_general['presupuesto'] > 0 or lote_general['cpv']:
                datos['lotes'].append(lote_general)

        return datos
    except Exception as e:
        st.error(f"Error al procesar XML: {e}")
        return None

def extraer_palabras_clave_inteligentes(texto):
    """Extraer solo las palabras clave MÁS relevantes del título"""
    import re

    # Normalizar texto
    def normalizar(texto):
        texto = texto.lower()
        texto = re.sub(r'[áàäâ]', 'a', texto)
        texto = re.sub(r'[éèëê]', 'e', texto)
        texto = re.sub(r'[íìïî]', 'i', texto)
        texto = re.sub(r'[óòöô]', 'o', texto)
        texto = re.sub(r'[úùüû]', 'u', texto)
        return texto

    texto_norm = normalizar(texto)
    texto_norm = re.sub(r'[^a-z0-9\s]', ' ', texto_norm)
    palabras = texto_norm.split()

    # Palabras a ignorar completamente (incluyendo genéricas que no son útiles solas)
    palabras_ignorar = {
        'de', 'del', 'la', 'el', 'los', 'las', 'y', 'a', 'en', 'para', 'con', 'por', 'segun',
        'contrato', 'servicio', 'servicios', 'suministro', 'obra', 'obras', 'lote',
        'mediante', 'procedimiento', 'abierto', 'simplificado', 'menor', 'expediente',
        'una', 'uno', 'unos', 'unas', 'este', 'esta', 'estos', 'estas', 'ese', 'esa',
        'municipal', 'municipales', 'apartado', 'cuadro', 'resumen', 'pcap', 'ppt',
        'madrid', 'barcelona', 'valencia', 'sevilla', 'ayuntamiento', 'diputacion',
        'manzanares', 'fernando', 'henares', 'real', 'serie', 'numero', 'fabricadas',
        # Palabras genéricas que no son útiles solas
        'instalacion', 'instalaciones', 'construccion', 'edificio', 'edificios',
        'sistema', 'sistemas', 'equipamiento', 'equipamientos', 'mantenimiento',
        'reparacion', 'mobiliario', 'material', 'materiales', 'equipo', 'equipos'
    }

    # Términos técnicos específicos que SÍ queremos capturar (solo palabras suficientemente específicas)
    terminos_tecnicos = {
        'fotovoltaica', 'fotovoltaico', 'solar', 'climatizacion', 'calefaccion', 'refrigeracion',
        'ascensor', 'ascensores', 'elevador', 'telecomunicaciones', 'informatica', 'informatico',
        'maquinaria', 'alumbrado', 'pergola', 'pergolas'
    }

    # Bigramas técnicos específicos que queremos (tienen PRIORIDAD sobre términos individuales)
    bigramas_tecnicos = {
        'acero inoxidable', 'energia solar', 'instalacion fotovoltaica', 'pergola fotovoltaica',
        'sistema solar', 'panel solar', 'bomba calor', 'aire acondicionado', 'eficiencia energetica',
        'alumbrado publico', 'material electrico', 'centro transformacion', 'instalacion electrica',
        'sistema fotovoltaico', 'planta fotovoltaica', 'energia fotovoltaica', 'panel fotovoltaico',
        'instalacion solar', 'sistema electrico', 'equipo informatico', 'sistema climatizacion',
        'instalacion climatizacion', 'mobiliario urbano', 'mobiliario escolar', 'vehiculo electrico'
    }

    palabras_clave = set()

    # 1. Buscar bigramas técnicos específicos
    texto_busqueda = ' ' + texto_norm + ' '
    for bigrama in bigramas_tecnicos:
        if bigrama in texto_busqueda:
            palabras_clave.add(bigrama)

    # 2. Buscar términos técnicos individuales
    for palabra in palabras:
        if palabra in terminos_tecnicos:
            palabras_clave.add(palabra)

    # 3. Si no encontramos nada técnico, buscar sustantivos principales (palabras largas)
    if not palabras_clave:
        for palabra in palabras:
            if (len(palabra) > 8 and
                palabra not in palabras_ignorar and
                not palabra.isdigit()):
                palabras_clave.add(palabra)

    # 4. Limitar a máximo 3-4 palabras clave para ser más preciso
    if len(palabras_clave) > 4:
        # Priorizar las más largas y técnicas
        palabras_clave = set(sorted(palabras_clave, key=lambda x: (len(x), x in terminos_tecnicos), reverse=True)[:4])

    return palabras_clave

def calcular_similitud(titulo_base, titulo_comparar):
    """Calcular similitud entre dos títulos usando palabras clave inteligentes"""

    # Extraer palabras clave de ambos títulos
    palabras_base = extraer_palabras_clave_inteligentes(titulo_base)
    palabras_comparar = extraer_palabras_clave_inteligentes(titulo_comparar)

    if not palabras_base or not palabras_comparar:
        return 0

    # Calcular intersección considerando bigramas y palabras individuales
    # Un bigrama puede coincidir con palabras individuales
    coincidencias = 0
    for clave_base in palabras_base:
        for clave_comp in palabras_comparar:
            # Coincidencia exacta
            if clave_base == clave_comp:
                coincidencias += 2  # Peso doble para coincidencias exactas
            # Coincidencia parcial (una palabra del bigrama coincide)
            elif ' ' in clave_base or ' ' in clave_comp:
                palabras_b = set(clave_base.split())
                palabras_c = set(clave_comp.split())
                if palabras_b.intersection(palabras_c):
                    coincidencias += 1

    # Normalizar
    total = len(palabras_base) + len(palabras_comparar)
    if total == 0:
        return 0

    similitud = (coincidencias * 2) / total  # Multiplicar por 2 porque contamos de ambos lados
    return min(similitud, 1.0)  # Limitar a máximo 1.0

def buscar_contratos(cpvs, presupuesto_min, presupuesto_max, titulo_referencia="", limit=10):
    """Buscar contratos similares por CPV y filtrar por relevancia"""
    # Si es una lista de CPVs, usar todos; si es string, convertir a lista
    if isinstance(cpvs, str):
        cpvs = [cpvs]

    # Extraer primeros 3 dígitos de cada CPV (más amplio que 4)
    cpv_patterns = []
    for cpv in cpvs[:3]:  # Usar hasta 3 CPVs para no hacer la query muy compleja
        cpv_digits = ''.join(filter(str.isdigit, str(cpv)))[:3]  # Cambiado a 3 dígitos
        if cpv_digits and len(cpv_digits) >= 3:
            cpv_patterns.append(cpv_digits)

    if not cpv_patterns:
        return []

    # Crear condición OR para múltiples CPVs
    cpv_condition = " OR ".join([f"cpv::text ~ '{cpv}'" for cpv in cpv_patterns])

    # Buscar hasta 100 licitaciones
    limit_busqueda = 100

    # Ampliar rango de presupuesto para tener más opciones
    presupuesto_min_amplio = presupuesto_min * 0.5  # 25% del presupuesto objetivo
    presupuesto_max_amplio = presupuesto_max * 1.5  # 225% del presupuesto objetivo

    query = f"""
    SELECT
        titulo,
        entidad_compradora as organismo,
        importe_total,
        importe_adjudicacion,
        adjudicatario,
        numero_licitadores,
        fecha_publicacion,
        ROUND(((importe_total - importe_adjudicacion) / NULLIF(importe_total, 0) * 100)::numeric, 2) as baja,
        cpv,
        provincia
    FROM adjudicaciones_metabase
    WHERE importe_total IS NOT NULL
    AND importe_adjudicacion IS NOT NULL
    AND importe_total > 0
    AND importe_adjudicacion > 0
    AND importe_total != importe_adjudicacion
    AND ({cpv_condition})
    AND importe_total BETWEEN {presupuesto_min_amplio} AND {presupuesto_max_amplio}
    ORDER BY fecha_publicacion DESC
    LIMIT {limit_busqueda}
    """

    conn = None
    try:
        conn = get_connection()
        if not conn:
            return []

        cur = conn.cursor()
        cur.execute(query)
        columns = [desc[0] for desc in cur.description]
        results = []

        for row in cur.fetchall():
            contrato = dict(zip(columns, row))

            # Extraer nombre de empresa
            adj_raw = contrato['adjudicatario']
            empresa = 'N/A'
            if adj_raw:
                try:
                    adj_str = str(adj_raw)
                    if adj_str.startswith('['):
                        adj_array = json.loads(adj_str)
                        if adj_array and 'adjudicatario' in adj_array[0]:
                            empresa = adj_array[0]['adjudicatario'].get('name', 'N/A')
                    elif adj_str.startswith('{'):
                        adj_dict = json.loads(adj_str)
                        if 'adjudicatario' in adj_dict:
                            empresa = adj_dict['adjudicatario'].get('name', 'N/A')
                    else:
                        empresa = adj_str[:80]
                except:
                    empresa = str(adj_raw)[:80] if adj_raw else 'N/A'

            contrato['empresa'] = empresa

            # Calcular similitud si hay título de referencia
            if titulo_referencia:
                similitud = calcular_similitud(titulo_referencia, contrato['titulo'])
                contrato['similitud'] = similitud
            else:
                contrato['similitud'] = 1.0

            results.append(contrato)

        cur.close()

        # Debug: mostrar cantidad antes del filtro
        if titulo_referencia:
            st.info(f"💾 Contratos recuperados de BD: {len(results)}")

        # Filtrar y ordenar por similitud
        if titulo_referencia and results:
            # Debug: Mostrar palabras clave inteligentes extraídas
            palabras_objetivo = extraer_palabras_clave_inteligentes(titulo_referencia)
            st.info(f"🎯 Palabras clave del objetivo: {', '.join(sorted(palabras_objetivo))}")

            # Ordenar por similitud
            results.sort(key=lambda x: x['similitud'], reverse=True)

            # Debug: Mostrar top 5 similitudes
            st.write("**Top 5 contratos por similitud:**")
            for i, c in enumerate(results[:5], 1):
                st.write(f"{i}. [{c['similitud']:.1%}] {c['titulo'][:60]}")

            # Estrategia adaptativa: ajustar umbral según los resultados
            umbral = 0.08  # 8% de palabras en común (más permisivo)

            # Filtrar por umbral
            results_filtrados = [c for c in results if c['similitud'] >= umbral]
            st.info(f"✅ Con umbral 8%: {len(results_filtrados)} contratos")

            # Si quedan muy pocos resultados, relajar el umbral
            if len(results_filtrados) < 5:
                umbral = 0.05  # Bajar a 5%
                results_filtrados = [c for c in results if c['similitud'] >= umbral]
                st.info(f"✅ Con umbral 5%: {len(results_filtrados)} contratos")

            # IMPORTANTE: Nunca devolver contratos con 0% de similitud
            # Si aún no hay suficientes, tomar solo los que tienen al menos algo de similitud
            if len(results_filtrados) < 3:
                # Tomar todos los que tienen > 0% de similitud
                results_filtrados = [c for c in results if c['similitud'] > 0]
                st.info(f"✅ Con similitud > 0%: {len(results_filtrados)} contratos")
                # Si realmente no hay ninguno similar, devolver vacío
                if not results_filtrados:
                    results_filtrados = []
                    st.warning("❌ Ningún contrato tiene similitud > 0%")

            results = results_filtrados

        # Devolver solo el límite solicitado
        return results[:limit]
    except Exception as e:
        st.error(f"Error en búsqueda: {e}")
        return []
    finally:
        if conn:
            conn.close()

def generar_texto_informe(lote, contratos, baja_prom, baja_min, baja_max, empresas, num_lic_prom, datos):
    """Generar texto del informe para copiar siguiendo la estructura estándar"""
    import random

    # Variaciones para la introducción
    saludos = ["Buenos días,", "Buenas tardes,", "Estimados,"]
    intros_criterios = [
        "En la selección de expedientes, nos encontramos los siguientes criterios de adjudicación:",
        "En el análisis del expediente, identificamos los siguientes criterios de adjudicación:",
        "Para este proceso, se establecen los siguientes criterios de adjudicación:"
    ]

    # Variaciones para análisis de participación
    intros_participacion = [
        f"Al revisar expedientes previos de similar envergadura y presupuesto, hemos observado una participación promedio de {int(num_lic_prom)} empresa{'s' if int(num_lic_prom) != 1 else ''}.",
        f"Tras analizar licitaciones similares en cuanto a presupuesto y alcance, detectamos una concurrencia media de {int(num_lic_prom)} empresa{'s' if int(num_lic_prom) != 1 else ''}.",
        f"En expedientes comparables en presupuesto y características, observamos una participación promedio de {int(num_lic_prom)} empresa{'s' if int(num_lic_prom) != 1 else ''}."
    ]

    # Variaciones para empresas destacadas
    intros_empresas = [
        "Entre las empresas más sobresalientes en este campo están",
        "Las empresas con mayor actividad en este sector incluyen a",
        "Destacan en este ámbito empresas como"
    ]

    # Variaciones para análisis de ofertas
    analisis_ofertas = [
        f"Notamos que las variaciones en las ofertas son notables, con un promedio de entre {baja_min:.1f}% y {baja_max:.1f}%, lo que demuestra una estrategia de ofertas variada.",
        f"Observamos diferencias significativas en las propuestas económicas, oscilando entre {baja_min:.1f}% y {baja_max:.1f}%, evidenciando estrategias de competencia diversas.",
        f"Las ofertas presentadas muestran variabilidad considerable, situándose entre {baja_min:.1f}% y {baja_max:.1f}%, reflejando distintos enfoques competitivos."
    ]

    # Variaciones para recomendación
    recomendaciones = [
        f"Por ello, sugerimos una propuesta económica con un margen de descuento del {baja_prom:.1f}%.",
        f"En consecuencia, recomendamos plantear una oferta con un descuento aproximado del {baja_prom:.1f}%.",
        f"Considerando lo anterior, aconsejamos una baja cercana al {baja_prom:.1f}%."
    ]

    despedidas = ["Un cordial saludo", "Saludos cordiales", "Atentamente"]

    # Generar el texto siguiendo la estructura del ejemplo
    texto = f"{random.choice(saludos)}\n"
    texto += f"{random.choice(intros_criterios)}\n"

    # Criterios de adjudicación
    if lote['criterios']:
        for i, crit in enumerate(lote['criterios'], 1):
            desc = crit.get('descripcion', f'Criterio {i}')
            peso = crit.get('peso', '')
            if peso:
                # Limpiar el peso (quitar % si existe, etc.)
                peso_limpio = peso.strip().replace('%', '')
                texto += f"{desc.upper()}: {peso_limpio} puntos\n"
            else:
                texto += f"{desc.upper()}\n"
    else:
        texto += "OFERTA ECONÓMICA: 100 puntos\n"

    # Análisis de participación
    texto += f"{random.choice(intros_participacion)}\n"

    # Empresas destacadas
    if empresas:
        sorted_emp = sorted(empresas.items(), key=lambda x: x[1], reverse=True)[:5]
        empresas_texto = ", ".join([emp for emp, _ in sorted_emp[:-1]])
        if len(sorted_emp) > 1:
            empresas_texto += f" y {sorted_emp[-1][0]}"
        else:
            empresas_texto = sorted_emp[0][0]
        texto += f"{random.choice(intros_empresas)} {empresas_texto}.\n"

    # Análisis de ofertas
    texto += f"{random.choice(analisis_ofertas)}\n"

    # Recomendación
    texto += f"{random.choice(recomendaciones)}\n"

    # Despedida
    texto += f"{random.choice(despedidas)}\n"

    return texto

def crear_excel(datos_lote, contratos, baja_recomendada):
    """Crear archivo Excel con los resultados"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Título
    ws['A1'] = "ANÁLISIS DE BAJA ESTADÍSTICA"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    # Datos del contrato
    row = 3
    ws[f'A{row}'] = "Presupuesto"
    ws[f'B{row}'] = f"€{datos_lote['presupuesto']:,.2f}"
    row += 1
    ws[f'A{row}'] = "CPV"
    ws[f'B{row}'] = ', '.join(datos_lote['cpv']) if datos_lote['cpv'] else 'N/A'
    row += 2

    # Baja recomendada
    ws[f'A{row}'] = "BAJA RECOMENDADA"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'] = f"{baja_recomendada:.2f}%"
    ws[f'B{row}'].font = Font(bold=True, size=14)
    row += 2

    # Contratos similares
    ws[f'A{row}'] = "CONTRATOS SIMILARES"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    # Cabeceras
    headers = ['Título', 'Organismo', 'Presupuesto', 'Adjudicación', 'Baja %', 'Empresa']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True)
    row += 1

    # Datos
    for contrato in contratos:
        ws.cell(row, 1, contrato['titulo'][:50])
        ws.cell(row, 2, contrato['organismo'][:30])
        ws.cell(row, 3, contrato['importe_total'])
        ws.cell(row, 4, contrato['importe_adjudicacion'])
        ws.cell(row, 5, contrato['baja'])
        ws.cell(row, 6, contrato['empresa'][:40])
        row += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 40

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# Interfaz principal
st.title("📊 Análisis de Bajas Estadísticas")
st.markdown("---")

xml_url = st.text_input(
    "Introduce la URL del XML del contrato:",
    placeholder="https://contrataciondelestado.es/FileSystem/servlet/...",
    help="Pega la URL completa del XML"
)

numero_lote = st.text_input(
    "Número de lote (opcional):",
    placeholder="Déjalo vacío para analizar todos los lotes",
    help="Si solo quieres analizar un lote específico, indica su número (1, 2, 3...)"
)

if st.button("🚀 Analizar Contrato", type="primary"):
    if not xml_url:
        st.warning("Por favor, introduce una URL")
    else:
        with st.spinner("Procesando XML..."):
            datos = extraer_datos_xml_completo(xml_url)

        if not datos or not datos['lotes']:
            st.error("No se pudieron extraer lotes del XML")
        else:
            st.success(f"✅ XML procesado - {len(datos['lotes'])} lote(s) encontrado(s)")

            # Filtrar por lote si se especificó
            lotes_a_analizar = datos['lotes']
            if numero_lote:
                lotes_a_analizar = [l for l in datos['lotes'] if l['numero'] == numero_lote]
                if not lotes_a_analizar:
                    st.error(f"No se encontró el lote {numero_lote}")
                    st.stop()

            # Analizar cada lote
            for lote in lotes_a_analizar:
                st.markdown("---")
                st.markdown(f"## 📦 Lote {lote['numero']}: {lote['titulo'][:80]}")

                st.markdown(f"**Presupuesto:** €{lote['presupuesto']:,.2f}")
                st.markdown(f"**CPV:** {', '.join(lote['cpv']) if lote['cpv'] else 'No especificado'}")

                # Criterios
                st.markdown("### ⚖️ Criterios de Adjudicación")
                if lote['criterios']:
                    for i, crit in enumerate(lote['criterios'], 1):
                        desc = crit.get('descripcion', f'Criterio {i}')
                        peso = crit.get('peso', '')
                        st.write(f"**{i}.** {desc.upper()}: **{peso}**" if peso else f"**{i}.** {desc.upper()}")
                else:
                    st.info("ℹ️ No se encontraron criterios de adjudicación en el XML")

                # Buscar contratos
                if lote['cpv'] and lote['presupuesto'] > 0:
                    st.markdown("### 🔍 Análisis de Mercado")

                    pres_min = lote['presupuesto'] * 0.5
                    pres_max = lote['presupuesto'] * 1.5

                    # Mostrar el título que se va a usar
                    st.markdown(f"**🔍 Título del contrato:** `{lote['titulo']}`")

                    with st.spinner("Buscando contratos similares..."):
                        # Pasar todos los CPVs, no solo el primero
                        contratos = buscar_contratos(lote['cpv'], pres_min, pres_max, titulo_referencia=lote['titulo'], limit=10)
                        st.info(f"📊 Contratos encontrados tras filtro: {len(contratos)}")

                    if not contratos:
                        cpvs_str = ", ".join([cpv[:3] for cpv in lote['cpv'][:3]])
                        st.warning(f"⚠️ No se encontraron contratos similares para los CPVs {cpvs_str}")
                    else:
                        # Calcular estadísticas
                        bajas = [c['baja'] for c in contratos if c['baja']]
                        empresas = {}
                        for c in contratos:
                            emp = c['empresa']
                            if emp and emp != 'N/A' and len(emp) > 3:
                                empresas[emp] = empresas.get(emp, 0) + 1

                        baja_min = min(bajas)
                        baja_max = max(bajas)
                        baja_prom = sum(bajas) / len(bajas)
                        num_lic_prom = sum([c['numero_licitadores'] or 0 for c in contratos]) / len(contratos)

                        # Resultados
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("🎯 Baja Recomendada", f"{baja_prom:.2f}%")
                        with col2:
                            st.metric("📈 Contratos Analizados", len(contratos))
                        with col3:
                            st.metric("👥 Licitadores Promedio", f"{num_lic_prom:.0f}")

                        # Análisis de mercado
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("**📈 Estadísticas de Competencia:**")
                            st.write(f"• Participación promedio: **{int(num_lic_prom)} empresas** por licitación")
                            st.write(f"• Rango de bajas: **{baja_min:.1f}% - {baja_max:.1f}%**")
                            st.write(f"• Baja media: **{baja_prom:.1f}%**")

                            st.markdown("")
                            st.markdown("**💡 Recomendación:**")
                            st.write(f"Al revisar {len(contratos)} expedientes previos de similar envergadura y presupuesto, observamos variaciones en las ofertas entre **{baja_min:.1f}% y {baja_max:.1f}%**, lo que demuestra una estrategia de ofertas variada.")

                        with col2:
                            if empresas:
                                st.markdown("**🏢 Empresas Más Activas:**")
                                sorted_emp = sorted(empresas.items(), key=lambda x: x[1], reverse=True)
                                for emp, count in sorted_emp[:5]:
                                    st.write(f"• {emp} ({count})")
                            else:
                                st.info("No se identificaron empresas específicas")

                        # Generar texto del informe
                        texto_informe = generar_texto_informe(lote, contratos, baja_prom, baja_min, baja_max, empresas, num_lic_prom, datos)

                        # Botón de descarga y texto
                        st.markdown("---")
                        st.markdown("### 📝 Informe Completo")

                        col1, col2 = st.columns([1, 1])
                        with col1:
                            excel_data = crear_excel(lote, contratos, baja_prom)
                            st.download_button(
                                label="📥 Descargar análisis en Excel",
                                data=excel_data,
                                file_name=f"analisis_lote_{lote['numero']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        with col2:
                            st.info(f"✅ {len(contratos)} contratos incluidos")

                        # Texto para copiar
                        st.markdown("#### 📄 Texto del Informe (Copia y Pega)")
                        st.text_area(
                            label="Texto completo del análisis:",
                            value=texto_informe,
                            height=400,
                            help="Selecciona todo el texto (Ctrl+A / Cmd+A) y copia (Ctrl+C / Cmd+C)",
                            label_visibility="collapsed"
                        )

                        # Mostrar contratos en expander
                        with st.expander(f"📋 Ver los {len(contratos)} contratos similares encontrados"):
                            for i, c in enumerate(contratos, 1):
                                st.markdown(f"**{i}. {c['titulo'][:80]}**")
                                col1, col2, col3 = st.columns(3)
                                with col1:
                                    st.write(f"📍 {c['organismo'][:30]}")
                                    st.write(f"🏢 {c['empresa'][:40]}")
                                with col2:
                                    st.write(f"💰 €{c['importe_total']:,.2f}")
                                    st.write(f"💵 €{c['importe_adjudicacion']:,.2f}")
                                with col3:
                                    st.write(f"📉 Baja: {c['baja']:.2f}%")
                                    st.write(f"👥 Licitadores: {c['numero_licitadores'] or 0}")
                                st.divider()
                else:
                    st.warning("⚠️ No se pudo extraer CPV o presupuesto del lote")

st.markdown("---")
st.caption("📊 Análisis basado en datos del Portal de Contratación del Estado")
