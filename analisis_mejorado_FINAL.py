import streamlit as st
import psycopg2
import pandas as pd
import json
import requests
import xml.etree.ElementTree as ET
from datetime import datetime
import random
import re
import traceback
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="Análisis de Bajas Estadísticas", page_icon="📊", layout="wide")

# Conexión a base de datos
def get_connection():
    """Crear nueva conexión a la base de datos"""
    try:
        conn = psycopg2.connect(
            host=st.secrets["postgres"]["host"],
            database=st.secrets["postgres"]["database"],
            user=st.secrets["postgres"]["user"],
            password=st.secrets["postgres"]["password"],
            port=st.secrets["postgres"]["port"]
        )
        return conn
    except Exception as e:
        st.error(f"Error al conectar con la base de datos: {e}")
        return None

def get_tag_name(element):
    """Obtener nombre del tag sin namespace"""
    return element.tag.split('}')[-1] if '}' in element.tag else element.tag

def extraer_datos_xml_completo(url):
    """Extraer datos completos del XML incluyendo lotes"""
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        root = ET.fromstring(response.content)

        datos = {
            'titulo': '',
            'organismo': '',
            'ubicacion': '',
            'lotes': []
        }

        # BUSCAR TÍTULO
        for elem in root.iter():
            tag = get_tag_name(elem)
            if tag == 'ProcurementProject':
                for child in elem:
                    child_tag = get_tag_name(child)
                    if child_tag == 'Name' and child.text:
                        texto = child.text.strip()
                        if len(texto) > 15:
                            datos['titulo'] = texto
                            break
                if datos['titulo']:
                    break

        if not datos['titulo']:
            for elem in root.iter():
                tag = get_tag_name(elem)
                if tag == 'Name' and elem.text:
                    texto = elem.text.strip()
                    if len(texto) > 20 and 'http' not in texto.lower():
                        datos['titulo'] = texto
                        break

        # BUSCAR ORGANISMO
        for elem in root.iter():
            tag = get_tag_name(elem)
            if tag == 'PartyName' and elem.text:
                datos['organismo'] = elem.text.strip()
                break

        # BUSCAR UBICACIÓN (Ciudad y Provincia)
        for elem in root.iter():
            tag = get_tag_name(elem)
            if tag == 'CityName' and elem.text:
                if not datos.get('ubicacion'):  # Solo tomar la primera
                    datos['ubicacion'] = elem.text.strip()
            if tag == 'CountrySubentityCode' and elem.text:
                # Código de provincia (ej: ES-M para Madrid)
                if not datos.get('provincia_codigo'):  # Solo tomar el primero
                    datos['provincia_codigo'] = elem.text.strip()
            if tag == 'CountrySubentity' and elem.text:
                # Nombre de la provincia
                if not datos.get('provincia'):  # Solo tomar la primera
                    datos['provincia'] = elem.text.strip()

        # BUSCAR LOTES
        for elem in root.iter():
            tag = get_tag_name(elem)
            if tag == 'ProcurementProjectLot':
                lote = {
                    'numero': '',
                    'titulo': '',
                    'presupuesto': 0,
                    'cpv': [],
                    'criterios': []
                }

                for child in elem.iter():
                    child_tag = get_tag_name(child)
                    if child_tag == 'ID' and child.text:
                        lote['numero'] = child.text.strip()
                        break

                for child in elem.iter():
                    child_tag = get_tag_name(child)
                    if child_tag == 'Name' and child.text:
                        if len(child.text.strip()) > 10:
                            lote['titulo'] = child.text.strip()
                            break

                # BUSCAR PRESUPUESTO - Priorizar PBL sobre valor estimado
                importes_encontrados = {}
                for child in elem.iter():
                    child_tag = get_tag_name(child)
                    if 'Amount' in child_tag and child.text:
                        try:
                            valor = float(child.text.strip())
                            # Clasificar por tipo de importe
                            if 'TaxExclusive' in child_tag or 'LineExtension' in child_tag:
                                importes_encontrados['pbl_sin_iva'] = valor
                            elif 'Payable' in child_tag or 'TaxInclusive' in child_tag:
                                importes_encontrados['pbl_con_iva'] = valor
                            elif 'Estimated' in child_tag:
                                importes_encontrados['estimado'] = valor
                            else:
                                importes_encontrados['otro'] = valor
                        except:
                            pass

                # Priorizar: PBL sin IVA > PBL con IVA > Estimado > Otro
                if 'pbl_sin_iva' in importes_encontrados:
                    lote['presupuesto'] = importes_encontrados['pbl_sin_iva']
                elif 'pbl_con_iva' in importes_encontrados:
                    lote['presupuesto'] = importes_encontrados['pbl_con_iva']
                elif 'estimado' in importes_encontrados:
                    lote['presupuesto'] = importes_encontrados['estimado']
                elif 'otro' in importes_encontrados:
                    lote['presupuesto'] = importes_encontrados['otro']

                for child in elem.iter():
                    child_tag = get_tag_name(child)
                    if child_tag == 'ItemClassificationCode' and child.text:
                        cpv_digits = ''.join(filter(str.isdigit, child.text))
                        if len(cpv_digits) >= 4:
                            lote['cpv'].append(cpv_digits)

                for child in elem.iter():
                    child_tag = get_tag_name(child)
                    if 'Criteria' in child_tag or 'Criterion' in child_tag:
                        criterio = {}
                        for subchild in child:
                            subtag = get_tag_name(subchild)
                            if subchild.text:
                                if any(x in subtag for x in ['Description', 'Name']):
                                    criterio['descripcion'] = subchild.text.strip()
                                elif any(x in subtag for x in ['Weight', 'Numeric']):
                                    criterio['peso'] = subchild.text.strip()

                        if criterio.get('descripcion'):
                            # Filtrar solvencia y requisitos previos
                            desc_lower = criterio['descripcion'].lower()
                            palabras_excluir = ['solvencia', 'solvències', 'habilitacion', 'capacidad',
                                              'acreditacion', 'declaracion responsable', 'certificado',
                                              'clasificacion empresarial', 'experiencia acreditada']

                            # Si contiene palabras de exclusión, no es criterio de adjudicación
                            if not any(palabra in desc_lower for palabra in palabras_excluir):
                                lote['criterios'].append(criterio)

                if lote['presupuesto'] > 0 or lote['cpv']:
                    if not lote['numero']:
                        lote['numero'] = str(len(datos['lotes']) + 1)
                    datos['lotes'].append(lote)

        # Si no hay lotes, buscar datos generales
        if not datos['lotes']:
            lote_general = {
                'numero': '1',
                'titulo': datos['titulo'] or 'Contrato único',
                'presupuesto': 0,
                'cpv': [],
                'criterios': []
            }

            # BUSCAR PRESUPUESTO - Priorizar PBL sobre valor estimado
            importes_encontrados = {}
            for elem in root.iter():
                tag = get_tag_name(elem)
                if 'Amount' in tag and elem.text:
                    try:
                        valor = float(elem.text.strip())
                        # Clasificar por tipo de importe
                        if 'TaxExclusive' in tag or 'LineExtension' in tag:
                            if 'pbl_sin_iva' not in importes_encontrados or valor > importes_encontrados['pbl_sin_iva']:
                                importes_encontrados['pbl_sin_iva'] = valor
                        elif 'Payable' in tag or 'TaxInclusive' in tag:
                            if 'pbl_con_iva' not in importes_encontrados or valor > importes_encontrados['pbl_con_iva']:
                                importes_encontrados['pbl_con_iva'] = valor
                        elif 'Estimated' in tag:
                            if 'estimado' not in importes_encontrados or valor > importes_encontrados['estimado']:
                                importes_encontrados['estimado'] = valor
                        else:
                            if 'otro' not in importes_encontrados or valor > importes_encontrados['otro']:
                                importes_encontrados['otro'] = valor
                    except:
                        pass

            # Priorizar: PBL sin IVA > PBL con IVA > Estimado > Otro
            if 'pbl_sin_iva' in importes_encontrados:
                lote_general['presupuesto'] = importes_encontrados['pbl_sin_iva']
            elif 'pbl_con_iva' in importes_encontrados:
                lote_general['presupuesto'] = importes_encontrados['pbl_con_iva']
            elif 'estimado' in importes_encontrados:
                lote_general['presupuesto'] = importes_encontrados['estimado']
            elif 'otro' in importes_encontrados:
                lote_general['presupuesto'] = importes_encontrados['otro']

            for elem in root.iter():
                tag = get_tag_name(elem)
                if tag == 'ItemClassificationCode':
                    cpv_text = elem.get('listID') or elem.text
                    if cpv_text:
                        cpv_digits = ''.join(filter(str.isdigit, cpv_text))
                        if len(cpv_digits) >= 4 and cpv_digits not in lote_general['cpv']:
                            lote_general['cpv'].append(cpv_digits)

            for elem in root.iter():
                tag = get_tag_name(elem)
                if 'Criteria' in tag or 'Criterion' in tag:
                    criterio = {}
                    for child in elem:
                        child_tag = get_tag_name(child)
                        if child.text:
                            if any(x in child_tag for x in ['Description', 'Name']):
                                criterio['descripcion'] = child.text.strip()
                            elif any(x in child_tag for x in ['Weight', 'Numeric', 'Percent']):
                                criterio['peso'] = child.text.strip()

                    if criterio.get('descripcion') and criterio not in lote_general['criterios']:
                        # Filtrar solvencia y requisitos previos
                        desc_lower = criterio['descripcion'].lower()
                        palabras_excluir = ['solvencia', 'solvències', 'habilitacion', 'capacidad',
                                          'acreditacion', 'declaracion responsable', 'certificado',
                                          'clasificacion empresarial', 'experiencia acreditada']

                        # Si contiene palabras de exclusión, no es criterio de adjudicación
                        if not any(palabra in desc_lower for palabra in palabras_excluir):
                            lote_general['criterios'].append(criterio)

            if lote_general['presupuesto'] > 0 or lote_general['cpv']:
                datos['lotes'].append(lote_general)

        return datos
    except Exception as e:
        st.error(f"Error al procesar XML: {e}")
        return None

def _extract_multilang_value(value):
    """Extraer texto de objetos multiidioma (ca, es, en, oc)"""
    if isinstance(value, dict):
        # Priorizar catalán, luego español, inglés y occitano
        for lang in ['ca', 'es', 'en', 'oc']:
            if lang in value and value[lang]:
                return value[lang]
        # Si no hay idiomas, intentar con 'name' o el primer valor string
        if 'name' in value:
            return value['name']
        if 'nom' in value:
            return value['nom']
        # Devolver el primer valor string que encuentre
        for v in value.values():
            if isinstance(v, str) and v.strip():
                return v
    return value

def _find_json_value(data, key_to_find):
    """Buscar una clave en un JSON de manera recursiva (case-insensitive)"""
    if isinstance(data, dict):
        # Buscar directamente (case-insensitive)
        for key, value in data.items():
            if key.lower() == key_to_find.lower() and value:
                return value

        # Buscar recursivamente en valores anidados
        for key, value in data.items():
            if isinstance(value, (dict, list)):
                result = _find_json_value(value, key_to_find)
                if result:
                    return result
    elif isinstance(data, list):
        for item in data:
            result = _find_json_value(item, key_to_find)
            if result:
                return result

    return None

def extraer_datos_json_completo(json_data):
    """Extraer datos completos del JSON incluyendo lotes"""
    try:
        # Si es un string, parsearlo como JSON
        if isinstance(json_data, str):
            data = json.loads(json_data)
        elif isinstance(json_data, dict):
            data = json_data
        else:
            st.error("Formato JSON no válido")
            return None

        datos = {
            'titulo': '',
            'organismo': '',
            'ubicacion': '',
            'lotes': []
        }

        # BUSCAR TÍTULO - Intentar rutas específicas primero
        # Para JSONs de Diputació de Barcelona
        if 'publicacio' in data and 'dadesBasiquesPublicacio' in data['publicacio']:
            dades = data['publicacio']['dadesBasiquesPublicacio']
            if 'denominacio' in dades:
                titulo_text = _extract_multilang_value(dades['denominacio'])
                if titulo_text and len(str(titulo_text).strip()) > 15:
                    datos['titulo'] = str(titulo_text).strip()

        # Si no se encontró, buscar genéricamente
        if not datos['titulo']:
            titulo_keys = ['titulo', 'title', 'name', 'objeto', 'description', 'asunto', 'denominacion', 'denominacio']
            for key in titulo_keys:
                value = _find_json_value(data, key)
                if value:
                    titulo_text = _extract_multilang_value(value)
                    if titulo_text and len(str(titulo_text).strip()) > 15:
                        datos['titulo'] = str(titulo_text).strip()
                        break

        # BUSCAR ORGANISMO - Intentar ruta específica primero
        if 'organ' in data and isinstance(data['organ'], dict):
            if 'nom' in data['organ']:
                datos['organismo'] = str(data['organ']['nom']).strip()

        # Si no se encontró, buscar genéricamente
        if not datos['organismo']:
            organismo_keys = ['organismo', 'entidad', 'organo', 'organ', 'buyer', 'contracting_authority', 'contratante', 'administracion', 'nom']
            for key in organismo_keys:
                value = _find_json_value(data, key)
                if value:
                    # Si el valor es un objeto (como 'organ'), buscar 'nom' o 'name' dentro
                    if isinstance(value, dict) and ('nom' in value or 'name' in value):
                        datos['organismo'] = str(value.get('nom') or value.get('name')).strip()
                        break
                    else:
                        organismo_text = _extract_multilang_value(value)
                        if organismo_text:
                            datos['organismo'] = str(organismo_text).strip()
                            break

        # BUSCAR UBICACIÓN - Intentar ruta específica primero
        # Para JSONs de Diputació (llocExecucio en dadesPublicacioLot)
        if 'publicacio' in data and 'dadesPublicacioLot' in data['publicacio']:
            if len(data['publicacio']['dadesPublicacioLot']) > 0:
                lot = data['publicacio']['dadesPublicacioLot'][0]
                if 'llocExecucio' in lot:
                    ubicacion_text = _extract_multilang_value(lot['llocExecucio'])
                    if ubicacion_text:
                        datos['ubicacion'] = str(ubicacion_text).strip()

        # Si no se encontró, buscar genéricamente
        if not datos['ubicacion']:
            ubicacion_keys = ['ubicacion', 'lugar', 'provincia', 'localitat', 'location', 'place', 'region', 'city', 'address', 'llocExecucio']
            for key in ubicacion_keys:
                value = _find_json_value(data, key)
                if value:
                    ubicacion_text = _extract_multilang_value(value)
                    if ubicacion_text:
                        datos['ubicacion'] = str(ubicacion_text).strip()
                        break

        # BUSCAR LOTES (o usar el documento completo como un único lote)
        lotes_data = _find_json_value(data, 'dadesPublicacioLot') or _find_json_value(data, 'lotes') or _find_json_value(data, 'lots')

        if lotes_data and isinstance(lotes_data, list):
            # Hay lotes definidos
            for idx, lote_data in enumerate(lotes_data, 1):
                lote = extraer_lote_json(lote_data, idx, datos['titulo'])
                if lote:
                    datos['lotes'].append(lote)
        else:
            # No hay lotes, usar el documento completo como un único lote
            lote = extraer_lote_json(data, 1, datos['titulo'])
            if lote:
                datos['lotes'].append(lote)

        return datos

    except Exception as e:
        st.error(f"Error al procesar JSON: {e}")
        st.error(traceback.format_exc())
        return None

def extraer_lote_json(lote_data, numero_lote, titulo_general=''):
    """Extraer información de un lote desde JSON"""
    try:
        lote = {
            'numero': str(numero_lote),
            'titulo': '',
            'presupuesto': 0,
            'cpv': [],
            'criterios': []
        }

        # BUSCAR TÍTULO DEL LOTE
        titulo_keys = ['titulo', 'denominacion', 'denominacio', 'name', 'description']
        for key in titulo_keys:
            value = _find_json_value(lote_data, key)
            if value:
                titulo_text = _extract_multilang_value(value)
                if titulo_text and len(str(titulo_text).strip()) > 10:
                    lote['titulo'] = str(titulo_text).strip()
                    break

        # Si no se encontró título, usar el título general del contrato
        if not lote['titulo'] and titulo_general:
            lote['titulo'] = titulo_general

        # BUSCAR PRESUPUESTO
        presupuesto_keys = ['presupuesto', 'pressupost', 'pressupostLicitacio', 'pressupostBaseLicitacioAmbIva',
                          'precio', 'valor', 'importe', 'amount', 'budget', 'value', 'estimatedValue']
        for key in presupuesto_keys:
            value = _find_json_value(lote_data, key)
            if value:
                try:
                    if isinstance(value, (int, float)):
                        lote['presupuesto'] = float(value)
                        break
                    elif isinstance(value, str):
                        clean_value = re.sub(r'[^\d.,]', '', value.replace(',', '.'))
                        if clean_value:
                            lote['presupuesto'] = float(clean_value)
                            break
                except:
                    continue

        # BUSCAR CPV - Intentar ruta específica primero
        if 'cpvPrincipal' in lote_data and isinstance(lote_data['cpvPrincipal'], dict):
            if 'codi' in lote_data['cpvPrincipal']:
                cpv_code = str(lote_data['cpvPrincipal']['codi']).strip()
                if cpv_code:
                    lote['cpv'].append(cpv_code)

        # Si no se encontró, buscar CPVs secundarios
        if 'cpvsSecundaris' in lote_data and isinstance(lote_data['cpvsSecundaris'], list):
            for cpv_item in lote_data['cpvsSecundaris']:
                if isinstance(cpv_item, dict) and 'codi' in cpv_item:
                    cpv_code = str(cpv_item['codi']).strip()
                    if cpv_code:
                        lote['cpv'].append(cpv_code)

        # Si aún no hay CPV, buscar genéricamente
        if not lote['cpv']:
            cpv_keys = ['cpv', 'cpvPrincipal', 'codigo', 'codi', 'classification', 'classificationCode', 'cpv_code']
            for key in cpv_keys:
                value = _find_json_value(lote_data, key)
                if value:
                    # Si es un objeto (como cpvPrincipal), buscar 'codi' o 'codigo'
                    if isinstance(value, dict):
                        cpv_value = value.get('codi') or value.get('codigo') or value.get('code')
                        if cpv_value:
                            lote['cpv'].append(str(cpv_value).strip())
                            break
                    # Si es una lista, tomar todos los códigos
                    elif isinstance(value, list):
                        for item in value:
                            if isinstance(item, dict):
                                cpv_value = item.get('codi') or item.get('codigo') or item.get('code')
                                if cpv_value:
                                    lote['cpv'].append(str(cpv_value).strip())
                            elif isinstance(item, str) and item.strip():
                                lote['cpv'].append(item.strip())
                        if lote['cpv']:
                            break
                    else:
                        # Si es string, añadirlo directamente
                        cpv_str = str(value).strip()
                        if cpv_str:
                            lote['cpv'].append(cpv_str)
                            break

        # BUSCAR CRITERIOS DE ADJUDICACIÓN (solo verdaderos criterios de evaluación)
        # Buscar específicamente en criterisAdjudicacio primero
        criterios_data = None
        if 'criterisAdjudicacio' in lote_data:
            criterios_data = lote_data['criterisAdjudicacio']
        elif 'criteriosAdjudicacion' in lote_data:
            criterios_data = lote_data['criteriosAdjudicacion']
        else:
            # Buscar genéricamente si no hay ruta directa
            criterios_keys = ['criterios', 'criteria', 'awardingCriteria', 'evaluationCriteria']
            for key in criterios_keys:
                criterios_data = _find_json_value(lote_data, key)
                if criterios_data:
                    break

        # Palabras que indican que NO es un criterio de adjudicación (son requisitos previos)
        palabras_excluir = ['solvencia', 'solvències', 'habilitacion', 'capacidad', 'acreditacion']

        if criterios_data and isinstance(criterios_data, list):
            for criterio in criterios_data:
                if isinstance(criterio, dict):
                    # Buscar descripción
                    desc_keys = ['descripcion', 'description', 'name', 'titulo', 'criteri']
                    desc_text = None
                    for desc_key in desc_keys:
                        if desc_key in criterio and criterio[desc_key]:
                            desc_value = _extract_multilang_value(criterio[desc_key])
                            if desc_value:
                                desc_text = str(desc_value).strip()
                                break

                    # FILTRAR: Excluir si menciona solvencia o capacidad
                    if desc_text:
                        desc_lower = desc_text.lower()
                        if any(palabra in desc_lower for palabra in palabras_excluir):
                            continue  # Saltar este "criterio" (es solvencia, no criterio)

                    # Buscar peso/ponderación
                    peso_text = None
                    if 'ponderacio' in criterio:
                        peso_val = criterio['ponderacio']
                        if isinstance(peso_val, (int, float)):
                            peso_text = f"{peso_val}"
                    else:
                        peso_keys = ['peso', 'weight', 'puntos', 'points', 'percentage']
                        for peso_key in peso_keys:
                            if peso_key in criterio and criterio[peso_key]:
                                peso_val = criterio[peso_key]
                                if isinstance(peso_val, (int, float)):
                                    peso_text = f"{peso_val}"
                                else:
                                    peso_text = str(peso_val)
                                break

                    # Si tiene desglossament (formato Diputació), procesarlo
                    if 'desglossament' in criterio and isinstance(criterio['desglossament'], list):
                        for subcriterio in criterio['desglossament']:
                            if isinstance(subcriterio, dict):
                                sub_desc = None
                                if 'descripcioCriteri' in subcriterio:
                                    sub_desc = _extract_multilang_value(subcriterio['descripcioCriteri'])
                                elif 'tipusCriteri' in subcriterio:
                                    sub_desc = _extract_multilang_value(subcriterio['tipusCriteri'])

                                sub_peso = None
                                if 'puntuacio' in subcriterio:
                                    sub_peso = f"{subcriterio['puntuacio']}"

                                if sub_desc:
                                    lote['criterios'].append(f"{sub_desc}: {sub_peso} puntos" if sub_peso else sub_desc)
                    else:
                        # Añadir criterio normal
                        if desc_text:
                            criterio_str = f"{desc_text}: {peso_text} puntos" if peso_text else desc_text
                            lote['criterios'].append(criterio_str)

        return lote if lote['presupuesto'] > 0 else None

    except Exception as e:
        st.warning(f"Error al procesar lote {numero_lote}: {e}")
        return None

def extraer_palabras_clave(texto):
    """Extraer palabras clave ESPECÍFICAS más relevantes del título (VERSIÓN MEJORADA)"""
    # Normalizar texto
    texto_original = texto.lower()
    texto = re.sub(r'[áàäâ]', 'a', texto_original)
    texto = re.sub(r'[éèëê]', 'e', texto)
    texto = re.sub(r'[íìïî]', 'i', texto)
    texto = re.sub(r'[óòöô]', 'o', texto)
    texto = re.sub(r'[úùüû]', 'u', texto)
    texto = re.sub(r'[^a-z0-9\s]', ' ', texto)

    palabras = texto.split()

    # Palabras GENÉRICAS a ignorar (reducido - menos agresivo)
    ignorar = {
        # Artículos, preposiciones
        'de', 'del', 'la', 'el', 'los', 'las', 'y', 'a', 'en', 'para', 'con', 'por', 'al', 'un', 'una',
        # Palabras contractuales muy genéricas
        'contrato', 'servicio', 'servicios', 'suministro', 'lote', 'lotes',
        'mediante', 'procedimiento', 'abierto', 'simplificado', 'menor', 'contratos',
        # Entidades
        'ayuntamiento', 'diputacion', 'municipal', 'concejo', 'consell', 'junta',
        # Solo las MÁS genéricas
        'mejora', 'mejoras', 'actuacion', 'actuaciones',
        'diversos', 'diversas', 'general', 'generales', 'varios', 'varias'
    }

    # NUEVO: Bigramas técnicos prioritarios (SIEMPRE se capturan aunque tengan palabras ignoradas)
    bigramas_tecnicos = {
        # Dirección y coordinación
        'direccion obras', 'direccion ejecucion', 'direccion facultativa', 'direccion tecnica',
        'coordinacion seguridad', 'coordinacion salud', 'asistencia tecnica',
        # Gestión y sistemas
        'gestion residuos', 'gestion basuras', 'gestion recursos', 'gestion energetica',
        'sistema gestion', 'sistema informacion', 'base datos', 'bases datos',
        # Mantenimiento específico
        'mantenimiento preventivo', 'mantenimiento correctivo', 'mantenimiento integral',
        # Instalación específica
        'instalacion electrica', 'instalacion fotovoltaica', 'instalacion solar',
        'instalacion climatizacion', 'instalacion alumbrado',
        # Proyectos específicos
        'redaccion proyecto', 'redaccion proyectos', 'proyecto ejecucion',
        # Obras específicas
        'obras reforma', 'obras ampliacion', 'obras mejora', 'obras construccion',
        'ejecucion obras', 'control obras', 'supervision obras',
        # Construcción específica
        'edificio residencial', 'edificio publico', 'construccion edificio',
        # Equipos específicos
        'equipos informaticos', 'equipos electronicos', 'material oficina',
        # Control y oficina
        'oficina tecnica', 'control calidad', 'oficina obras'
    }

    # NUEVO: Palabras contextuales (palabras que se vuelven importantes si van acompañadas)
    palabras_contextuales = {
        # Palabra: [palabras que la hacen importante]
        'obras': ['direccion', 'ejecucion', 'coordinacion', 'control', 'supervision', 'reforma', 'ampliacion', 'construccion'],
        'proyecto': ['redaccion', 'desarrollo', 'ejecucion', 'basico', 'detallado'],
        'sistema': ['gestion', 'informacion', 'informatico', 'control', 'seguridad'],
        'gestion': ['residuos', 'basuras', 'recursos', 'energetica', 'administrativa'],
        'mantenimiento': ['preventivo', 'correctivo', 'integral', 'instalaciones'],
        'instalacion': ['electrica', 'fotovoltaica', 'solar', 'climatizacion', 'alumbrado'],
        'edificio': ['residencial', 'publico', 'oficinas', 'administrativo'],
        'equipos': ['informaticos', 'electronicos', 'medicos', 'deportivos'],
        'material': ['oficina', 'escolar', 'sanitario', 'deportivo'],
        'construccion': ['edificio', 'piscina', 'polideportivo', 'centro'],
        'ejecucion': ['obras', 'proyecto', 'trabajos'],
        'direccion': ['obras', 'ejecucion', 'facultativa', 'tecnica', 'proyecto'],
        'coordinacion': ['seguridad', 'salud', 'obras', 'trabajos'],
        'redaccion': ['proyecto', 'proyectos', 'memoria', 'informe'],
        'oficina': ['tecnica', 'obras', 'atencion']
    }

    # PASO 1: Detectar bigramas técnicos PRIORITARIOS (consecutivos)
    bigramas_prioritarios = []
    for i in range(len(palabras) - 1):
        bigrama = f"{palabras[i]} {palabras[i+1]}"
        # Si es bigrama técnico, SIEMPRE añadirlo
        if bigrama in bigramas_tecnicos:
            bigramas_prioritarios.append(bigrama)

    # PASO 1B: Detectar bigramas técnicos con PALABRAS DE RELLENO (hasta 3 palabras de distancia)
    # Ejemplo: "direccion de las obras" → detectar "direccion obras"
    palabras_relleno = {'de', 'del', 'la', 'el', 'los', 'las', 'y', 'a', 'en', 'para', 'con', 'por', 'al', 'un', 'una'}
    for i in range(len(palabras)):
        for j in range(i + 2, min(i + 5, len(palabras))):  # Buscar hasta 4 palabras adelante
            # Verificar que entre i y j solo haya palabras de relleno
            palabras_entre = palabras[i+1:j]
            if all(p in palabras_relleno for p in palabras_entre):
                bigrama_candidato = f"{palabras[i]} {palabras[j]}"
                if bigrama_candidato in bigramas_tecnicos and bigrama_candidato not in bigramas_prioritarios:
                    bigramas_prioritarios.append(bigrama_candidato)

    # PASO 2: Detectar bigramas contextuales (palabras que se hacen importantes juntas)
    bigramas_contextuales = []

    # PASO 2A: Bigramas contextuales consecutivos
    for i in range(len(palabras) - 1):
        palabra1 = palabras[i]
        palabra2 = palabras[i+1]

        # Verificar si palabra1 es contextual y palabra2 la activa
        if palabra1 in palabras_contextuales:
            if palabra2 in palabras_contextuales[palabra1]:
                bigrama = f"{palabra1} {palabra2}"
                if bigrama not in bigramas_prioritarios:  # Evitar duplicados
                    bigramas_contextuales.append(bigrama)

        # Verificar al revés (palabra2 contextual, palabra1 la activa)
        if palabra2 in palabras_contextuales:
            if palabra1 in palabras_contextuales[palabra2]:
                bigrama = f"{palabra1} {palabra2}"
                if bigrama not in bigramas_prioritarios and bigrama not in bigramas_contextuales:
                    bigramas_contextuales.append(bigrama)

    # PASO 2B: Bigramas contextuales con palabras de relleno en medio
    for i in range(len(palabras)):
        for j in range(i + 2, min(i + 5, len(palabras))):
            palabras_entre = palabras[i+1:j]
            if all(p in palabras_relleno for p in palabras_entre):
                palabra1 = palabras[i]
                palabra2 = palabras[j]

                # Verificar contexto
                if palabra1 in palabras_contextuales:
                    if palabra2 in palabras_contextuales[palabra1]:
                        bigrama = f"{palabra1} {palabra2}"
                        if bigrama not in bigramas_prioritarios and bigrama not in bigramas_contextuales:
                            bigramas_contextuales.append(bigrama)

                if palabra2 in palabras_contextuales:
                    if palabra1 in palabras_contextuales[palabra2]:
                        bigrama = f"{palabra1} {palabra2}"
                        if bigrama not in bigramas_prioritarios and bigrama not in bigramas_contextuales:
                            bigramas_contextuales.append(bigrama)

    # PASO 3: Detectar bigramas normales (sin palabras ignoradas)
    bigramas_normales = []
    for i in range(len(palabras) - 1):
        if len(palabras[i]) > 3 and len(palabras[i+1]) > 3:
            if palabras[i] not in ignorar and palabras[i+1] not in ignorar:
                bigrama = f"{palabras[i]} {palabras[i+1]}"
                # No duplicar si ya está en prioritarios o contextuales
                if bigrama not in bigramas_prioritarios and bigrama not in bigramas_contextuales:
                    bigramas_normales.append(bigrama)

    # Unir todos los bigramas (prioritarios primero, luego contextuales, luego normales)
    bigramas = bigramas_prioritarios + bigramas_contextuales + bigramas_normales

    # PASO 4: Palabras individuales contextuales (palabras que aparecen en bigramas contextuales)
    palabras_de_bigramas_contextuales = set()
    for bigrama in bigramas_contextuales:
        for palabra in bigrama.split():
            if len(palabra) > 4:  # Solo palabras significativas
                palabras_de_bigramas_contextuales.add(palabra)

    # PASO 5: Filtrar palabras individuales (más de 4 letras - menos restrictivo)
    palabras_individuales = []
    for p in palabras:
        if len(p) > 4:
            # Incluir si NO está ignorada O si está en bigramas contextuales
            if p not in ignorar or p in palabras_de_bigramas_contextuales:
                palabras_individuales.append(p)

    # PASO 6: Detectar palabras NÚCLEO (muy específicas de la actividad) - AMPLIADO
    palabras_nucleo = []

    # Palabras específicas de actividades (no genéricas)
    palabras_especificas = {
        # Vehículos y transporte
        'vehiculos', 'automoviles', 'camiones', 'autobuses', 'turismos', 'motos', 'furgonetas',
        # Energía específica
        'recarga', 'fotovoltaica', 'fotovoltaico', 'solar', 'eolica', 'biomasa', 'cogeneracion',
        # Servicios específicos
        'limpieza', 'jardineria', 'seguridad', 'vigilancia', 'catering', 'comedor', 'transporte',
        'mensajeria', 'lavanderia', 'desinfeccion', 'fumigacion', 'desratizacion',
        # Tecnología específica
        'software', 'hardware', 'informatica', 'telecomunicaciones', 'fibra', 'servidor',
        'base', 'datos', 'backup', 'firewall', 'router', 'switch', 'cableado',
        # Construcción específica
        'asfaltado', 'pavimentacion', 'acerado', 'alumbrado', 'alcantarillado', 'fontaneria',
        'carpinteria', 'cerrajeria', 'climatizacion', 'calefaccion', 'refrigeracion',
        # Áreas específicas
        'piscina', 'polideportivo', 'biblioteca', 'museo', 'teatro', 'auditorio',
        'residencia', 'colegio', 'escuela', 'hospital', 'centro', 'parque',
        # Servicios públicos específicos
        'residuos', 'basuras', 'reciclaje', 'contenedores', 'recogida',
        'abastecimiento', 'depuracion', 'potabilizacion', 'saneamiento',
        # NUEVAS: Técnicas y profesionales
        'facultativo', 'redaccion', 'direccion', 'coordinacion', 'supervision',
        'preventivo', 'correctivo', 'integral', 'tecnica', 'tecnicos'
    }

    for palabra in palabras_individuales:
        if palabra in palabras_especificas:
            palabras_nucleo.append(palabra)

    # PASO 7: Seleccionar las mejores palabras clave (hasta 5 en lugar de 3)
    palabras_finales = set()

    # Prioridad 1: Bigramas técnicos prioritarios (los más importantes)
    for bigrama in bigramas_prioritarios[:2]:  # Hasta 2 bigramas técnicos
        palabras_finales.add(bigrama)

    # Prioridad 2: Bigramas contextuales
    for bigrama in bigramas_contextuales[:2]:  # Hasta 2 bigramas contextuales
        if len(palabras_finales) >= 5:
            break
        palabras_finales.add(bigrama)

    # Prioridad 3: Palabras núcleo (específicas de actividad)
    if palabras_nucleo:
        for palabra in palabras_nucleo[:3]:  # Hasta 3 palabras núcleo
            if len(palabras_finales) >= 5:
                break
            palabras_finales.add(palabra)

    # Prioridad 4: Bigramas normales
    if len(palabras_finales) < 5 and bigramas_normales:
        palabras_finales.add(bigramas_normales[0])

    # Prioridad 5: Palabras individuales más largas (si faltan)
    if len(palabras_finales) < 5:
        palabras_ordenadas = sorted(palabras_individuales, key=len, reverse=True)
        for palabra in palabras_ordenadas:
            if len(palabras_finales) >= 5:
                break
            # Evitar duplicados (que la palabra no esté ya en un bigrama)
            if palabra not in ' '.join(palabras_finales):
                palabras_finales.add(palabra)

    return palabras_finales

def calcular_similitud_palabras(titulo_base, titulo_comparar):
    """Calcular similitud basada en palabras clave comunes"""
    palabras_base = extraer_palabras_clave(titulo_base)
    palabras_comp = extraer_palabras_clave(titulo_comparar)

    if not palabras_base or not palabras_comp:
        return 0

    # Palabras en común
    comunes = palabras_base.intersection(palabras_comp)

    if not comunes:
        return 0

    # Similitud = palabras comunes / promedio de palabras totales
    similitud = len(comunes) / ((len(palabras_base) + len(palabras_comp)) / 2)

    return min(similitud, 1.0)

def detectar_grupo_similar(bajas, tolerancia=4):
    """
    Detecta el grupo más grande de bajas correlativas donde cada baja
    tiene una diferencia ≤ tolerancia (4%) con la siguiente.
    Retorna el grupo más grande de bajas correlativas (mínimo 2)
    """
    if len(bajas) < 2:
        return []

    # Ordenar bajas
    bajas_ordenadas = sorted(bajas)
    grupos = []

    # Buscar todos los grupos posibles
    i = 0
    while i < len(bajas_ordenadas):
        grupo_actual = [bajas_ordenadas[i]]

        # Agregar bajas consecutivas mientras la diferencia sea ≤ tolerancia
        j = i + 1
        while j < len(bajas_ordenadas):
            diferencia = bajas_ordenadas[j] - bajas_ordenadas[j-1]
            if diferencia <= tolerancia:
                grupo_actual.append(bajas_ordenadas[j])
                j += 1
            else:
                break

        # Guardar el grupo si tiene al menos 2 elementos
        if len(grupo_actual) >= 2:
            grupos.append(grupo_actual)

        # Avanzar al siguiente grupo
        i = j if j > i + 1 else i + 1

    # Retornar el grupo más grande
    if grupos:
        return max(grupos, key=len)
    return []

def calcular_baja_recomendada(bajas):
    """
    Calcula la baja recomendada según el nuevo algoritmo:
    - Si hay 2+ bajas correlativas (diferencia consecutiva ≤4%): max del grupo + 2%
    - Si todas diferentes: media + 2%
    """
    if not bajas:
        return 0

    grupo_similar = detectar_grupo_similar(bajas, tolerancia=4)

    if grupo_similar:
        # Hay un grupo de bajas correlativas
        baja_mas_alta = max(grupo_similar)
        baja_mas_baja = min(grupo_similar)
        rango = baja_mas_alta - baja_mas_baja

        # Calcular diferencias consecutivas
        grupo_ordenado = sorted(grupo_similar)
        diferencias = [f"{grupo_ordenado[i+1] - grupo_ordenado[i]:.1f}%"
                      for i in range(len(grupo_ordenado)-1)]

        baja_recomendada = baja_mas_alta + 2
        st.info(f"✅ **Grupo de {len(grupo_similar)} bajas correlativas encontrado**: {[f'{b:.1f}%' for b in grupo_ordenado]}")
        st.info(f"📏 **Diferencias consecutivas**: {' → '.join(diferencias)} (todas ≤4%)")
        st.info(f"📊 **Cálculo**: Baja más alta ({baja_mas_alta:.2f}%) + 2% = **{baja_recomendada:.2f}%**")
    else:
        # Todas diferentes, hacer media
        media = sum(bajas) / len(bajas)
        baja_recomendada = media + 2
        st.info(f"ℹ️ **No se encontró grupo correlativo** (diferencias consecutivas >4%)")
        st.info(f"📊 **Cálculo**: Media de bajas ({media:.2f}%) + 2% = **{baja_recomendada:.2f}%**")

    return baja_recomendada

def buscar_contratos(cpvs, presupuesto_min, presupuesto_max, titulo_referencia="", limit=10, ampliada=False, provincia_origen=None, palabras_clave_manual=None):
    """Buscar contratos similares con criterios específicos"""
    if isinstance(cpvs, str):
        cpvs = [cpvs]

    # Extraer CPV según si es búsqueda ampliada o normal
    cpv_patterns = []
    if ampliada:
        # Búsqueda ampliada: usar primeros 2 dígitos (más flexible)
        for cpv in cpvs[:3]:
            cpv_digits = ''.join(filter(str.isdigit, str(cpv)))
            if len(cpv_digits) >= 2:
                cpv_patterns.append(cpv_digits[:2])  # 2 dígitos (más amplio)
        st.warning(f"🔄 **Búsqueda ampliada**: CPV primeros 2 dígitos (más flexible)")
    else:
        # Búsqueda normal: usar primeros 4 dígitos
        for cpv in cpvs[:3]:
            cpv_digits = ''.join(filter(str.isdigit, str(cpv)))
            if len(cpv_digits) >= 4:
                cpv_patterns.append(cpv_digits[:4])  # 4 dígitos

    if not cpv_patterns:
        st.warning("❌ No se pudieron extraer CPVs válidos")
        return []

    # Eliminar duplicados
    cpv_patterns = list(set(cpv_patterns))

    if ampliada:
        st.info(f"🔍 **Buscando con CPV**: {', '.join(cpv_patterns)} (primeros 2 dígitos)")
    else:
        st.info(f"🔍 **Buscando con CPV**: {', '.join(cpv_patterns)} (primeros 4 dígitos)")

    cpv_condition = " OR ".join([f"cpv::text ~ '^{cpv}'" for cpv in cpv_patterns])

    # Presupuesto objetivo
    presupuesto_objetivo = (presupuesto_min + presupuesto_max) / 2

    # Rango de presupuesto según si es búsqueda ampliada o normal
    if ampliada:
        # Búsqueda ampliada: ±100% del objetivo (más flexible)
        presupuesto_min_rango = presupuesto_objetivo * 0.3
        presupuesto_max_rango = presupuesto_objetivo * 2.0
        if provincia_origen:
            st.warning(f"🔄 **Búsqueda ampliada** - Rango presupuesto (±100%): €{presupuesto_min_rango:,.0f} - €{presupuesto_max_rango:,.0f}")
            st.info(f"💡 **Manteniendo**: Palabra clave + Provincia ({provincia_origen})")
        else:
            st.warning(f"🔄 **Búsqueda ampliada** - Rango presupuesto (±100%): €{presupuesto_min_rango:,.0f} - €{presupuesto_max_rango:,.0f}")
    else:
        # Búsqueda normal: ±30% del objetivo
        presupuesto_min_rango = presupuesto_objetivo * 0.7
        presupuesto_max_rango = presupuesto_objetivo * 1.3
        st.info(f"💰 **Rango presupuesto (±30%)**: €{presupuesto_min_rango:,.0f} - €{presupuesto_max_rango:,.0f}")

    query = f"""
    SELECT
        titulo,
        entidad_compradora as organismo,
        importe_total,
        importe_adjudicacion,
        adjudicatario,
        numero_licitadores,
        fecha_publicacion,
        ROUND(((importe_total - importe_adjudicacion) / NULLIF(importe_total, 0) * 100)::numeric, 2) as baja,
        cpv,
        INITCAP(LOWER(TRIM(provincia))) as provincia
    FROM adjudicaciones_metabase
    WHERE importe_total IS NOT NULL
    AND importe_adjudicacion IS NOT NULL
    AND importe_total > 0
    AND importe_adjudicacion > 0
    AND importe_total != importe_adjudicacion
    AND adjudicatario IS NOT NULL
    AND adjudicatario != 'null'
    AND adjudicatario != ''
    AND ({cpv_condition})
    AND importe_total BETWEEN {presupuesto_min_rango} AND {presupuesto_max_rango}
    AND ROUND(((importe_total - importe_adjudicacion) / NULLIF(importe_total, 0) * 100)::numeric, 2) > 0.5
    AND ROUND(((importe_total - importe_adjudicacion) / NULLIF(importe_total, 0) * 100)::numeric, 2) < 70
    ORDER BY fecha_publicacion DESC
    LIMIT 300
    """

    conn = None
    try:
        conn = get_connection()
        if not conn:
            return []

        cur = conn.cursor()
        cur.execute(query)
        columns = [desc[0] for desc in cur.description]
        results = []

        for row in cur.fetchall():
            contrato = dict(zip(columns, row))

            # Extraer nombre de empresa
            adj_raw = contrato['adjudicatario']
            empresa = 'N/A'
            if adj_raw:
                try:
                    adj_str = str(adj_raw)
                    if adj_str.startswith('['):
                        adj_array = json.loads(adj_str)
                        if adj_array and 'adjudicatario' in adj_array[0]:
                            empresa = adj_array[0]['adjudicatario'].get('name', 'N/A')
                    elif adj_str.startswith('{'):
                        adj_dict = json.loads(adj_str)
                        if 'adjudicatario' in adj_dict:
                            empresa = adj_dict['adjudicatario'].get('name', 'N/A')
                    else:
                        empresa = adj_str[:80]
                except:
                    empresa = str(adj_raw)[:80] if adj_raw else 'N/A'

            contrato['empresa'] = empresa
            results.append(contrato)

        cur.close()

        st.info(f"💾 **Contratos recuperados de BD**: {len(results)}")

        if not results:
            st.error("❌ No se encontraron contratos con ese CPV y presupuesto.")
            return []

        # FILTRAR POR SIMILITUD DE PALABRAS CLAVE
        if titulo_referencia or palabras_clave_manual:
            # Guardar results originales ANTES de filtrar por palabras clave (para fallback geográfico)
            results_sin_filtro_palabras = results.copy()

            # Usar palabras clave manuales si están disponibles, si no extraerlas automáticamente
            if palabras_clave_manual:
                # Procesar palabras clave manuales
                palabras_objetivo = set([p.strip().lower() for p in palabras_clave_manual.split(',') if p.strip()])
                st.info(f"🎯 **Palabras clave manuales**: {', '.join(sorted(palabras_objetivo))}")
            else:
                palabras_objetivo = extraer_palabras_clave(titulo_referencia)
                st.info(f"🎯 **Palabras clave extraídas automáticamente**: {', '.join(sorted(palabras_objetivo))}")

            # Función para normalizar texto (para búsqueda)
            def normalizar_para_busqueda(texto):
                if not texto:
                    return ''
                texto = texto.lower().strip()
                # Quitar acentos
                texto = re.sub(r'[áàäâ]', 'a', texto)
                texto = re.sub(r'[éèëê]', 'e', texto)
                texto = re.sub(r'[íìïî]', 'i', texto)
                texto = re.sub(r'[óòöô]', 'o', texto)
                texto = re.sub(r'[úùüû]', 'u', texto)
                return texto

            # Calcular palabras coincidentes y similitud para cada contrato
            for c in results:
                if palabras_clave_manual:
                    # BÚSQUEDA DIRECTA EN TÍTULO (sin extraer palabras clave)
                    # Normalizar título del contrato
                    titulo_normalizado = normalizar_para_busqueda(c['titulo'])

                    comunes = set()
                    # Buscar cada palabra manual en el título
                    for palabra_objetivo in palabras_objetivo:
                        # Normalizar palabra objetivo
                        palabra_normalizada = normalizar_para_busqueda(palabra_objetivo)

                        # Buscar si la palabra está contenida en el título
                        if palabra_normalizada in titulo_normalizado:
                            comunes.add(palabra_objetivo)

                    c['num_palabras_comunes'] = len(comunes)
                    c['palabras_comunes'] = comunes
                    if palabras_objetivo:
                        c['similitud'] = len(comunes) / len(palabras_objetivo)  # Porcentaje de palabras encontradas
                    else:
                        c['similitud'] = 0
                else:
                    # Usar sistema automático de extracción de palabras clave
                    palabras_contrato = extraer_palabras_clave(c['titulo'])
                    comunes = palabras_objetivo.intersection(palabras_contrato)
                    c['num_palabras_comunes'] = len(comunes)
                    c['palabras_comunes'] = comunes
                    c['similitud'] = calcular_similitud_palabras(titulo_referencia, c['titulo'])

            # FILTRAR: solo contratos con al menos 1 palabra en común
            results_filtrados = [c for c in results if c['num_palabras_comunes'] > 0]

            st.info(f"🔍 **Contratos con palabras clave en común**: {len(results_filtrados)}")

            if not results_filtrados:
                st.warning("⚠️ No se encontraron contratos con palabras clave similares")
                st.write("**Mostrando los 5 más recientes sin filtro:**")
                results = results[:5]
                for i, c in enumerate(results, 1):
                    fecha_str = str(c['fecha_publicacion'])[:10] if c['fecha_publicacion'] else 'N/A'
                    st.write(f"{i}. [{fecha_str}] {c['titulo'][:70]}")
                return results

            results = results_filtrados

            # Función para normalizar texto (quitar acentos, minúsculas, espacios)
            def normalizar_texto(texto):
                if not texto:
                    return ''
                texto = texto.lower().strip()
                # Quitar acentos
                texto = re.sub(r'[áàäâ]', 'a', texto)
                texto = re.sub(r'[éèëê]', 'e', texto)
                texto = re.sub(r'[íìïî]', 'i', texto)
                texto = re.sub(r'[óòöô]', 'o', texto)
                texto = re.sub(r'[úùüû]', 'u', texto)
                # Quitar caracteres especiales excepto espacios
                texto = re.sub(r'[^a-z0-9\s]', '', texto)
                return texto

            # Función para comparar provincias de manera flexible
            def provincias_coinciden(prov1, prov2):
                """Compara dos provincias de manera flexible"""
                if not prov1 or not prov2:
                    return False

                prov1_norm = normalizar_texto(prov1)
                prov2_norm = normalizar_texto(prov2)

                # 1. Coincidencia exacta
                if prov1_norm == prov2_norm:
                    return True

                # 2. Uno contiene al otro
                if prov1_norm in prov2_norm or prov2_norm in prov1_norm:
                    return True

                # 3. Extraer palabras clave de cada provincia
                # Eliminar palabras comunes geográficas
                palabras_ignorar = {'provincia', 'comunidad', 'autonoma', 'ciudad', 'de', 'del', 'la', 'las', 'el', 'los'}

                palabras1 = set([p for p in prov1_norm.split() if p not in palabras_ignorar and len(p) > 2])
                palabras2 = set([p for p in prov2_norm.split() if p not in palabras_ignorar and len(p) > 2])

                # Si alguna palabra clave coincide
                if palabras1 and palabras2:
                    if palabras1.intersection(palabras2):
                        return True

                # 4. Diccionario de provincias bilingües
                equivalencias = {
                    'valencia': ['valencia', 'valencia', 'valenciana'],
                    'alicante': ['alicante', 'alacant'],
                    'castellon': ['castellon', 'castello'],
                    'barcelona': ['barcelona'],
                    'girona': ['girona', 'gerona'],
                    'lleida': ['lleida', 'lerida'],
                    'tarragona': ['tarragona'],
                    'vizcaya': ['vizcaya', 'bizkaia'],
                    'guipuzcoa': ['guipuzcoa', 'gipuzkoa'],
                    'alava': ['alava', 'araba'],
                    'navarra': ['navarra', 'nafarroa'],
                    'coruna': ['coruna', 'corunha'],
                    'orense': ['orense', 'ourense'],
                    'pontevedra': ['pontevedra'],
                    'lugo': ['lugo'],
                    'baleares': ['baleares', 'balears', 'illes', 'islas']
                }

                # Buscar en equivalencias
                for key, variantes in equivalencias.items():
                    if any(v in prov1_norm for v in variantes) and any(v in prov2_norm for v in variantes):
                        return True

                return False

            # Calcular proximidad geográfica mejorada
            if provincia_origen:
                st.info(f"📍 **Provincia de origen**: {provincia_origen}")
                st.info(f"🔍 **Provincia normalizada para búsqueda**: '{normalizar_texto(provincia_origen)}'")

                # Contador para debug
                provincias_encontradas = {}  # provincia_original: provincia_normalizada
                contratos_misma_provincia = 0
                ejemplos_match = []  # Para mostrar ejemplos de matches exitosos

                for c in results:
                    provincia_contrato = c.get('provincia', '')

                    if provincia_contrato:
                        provincias_encontradas[provincia_contrato] = normalizar_texto(provincia_contrato)

                    # Usar función de comparación flexible
                    if provincias_coinciden(provincia_origen, provincia_contrato):
                        c['proximidad'] = 1
                        contratos_misma_provincia += 1
                        if len(ejemplos_match) < 3:
                            ejemplos_match.append(f"'{provincia_contrato}' ✅ match con '{provincia_origen}'")
                    else:
                        c['proximidad'] = 0

                # Mostrar info de debug detallada
                if contratos_misma_provincia > 0:
                    st.success(f"✅ **{contratos_misma_provincia} contratos encontrados en {provincia_origen}**")
                    if ejemplos_match:
                        st.info(f"🔍 **Ejemplos de matches**: {' | '.join(ejemplos_match)}")
                else:
                    st.warning(f"⚠️ **No se encontraron contratos en {provincia_origen}**")
                    if provincias_encontradas:
                        # Mostrar las primeras 10 provincias con su normalización
                        provincias_debug = []
                        for prov_orig, prov_norm in sorted(list(provincias_encontradas.items()))[:10]:
                            provincias_debug.append(f"{prov_orig} ('{prov_norm}')")
                        st.info(f"🗺️ **Provincias en resultados**:\n" + "\n".join([f"- {p}" for p in provincias_debug]))
                        st.error(f"❌ **Buscando**: '{provincia_norm}' - **No coincide con ninguna**")
            else:
                # Sin provincia origen, todos tienen misma proximidad
                for c in results:
                    c['proximidad'] = 0

            # SISTEMA DE FILTRADO POR NIVELES CON PRIORIZACIÓN INTELIGENTE
            # Nivel 1: Palabras clave + Zona + Recientes (últimos 2 años)
            fecha_limite = datetime.now() - pd.DateOffset(years=2)
            nivel_1 = [c for c in results if c['num_palabras_comunes'] > 0 and c['proximidad'] == 1
                      and c['fecha_publicacion'] and c['fecha_publicacion'] >= fecha_limite]

            # Nivel 2: Palabras clave + Zona (sin filtro de fecha)
            nivel_2 = [c for c in results if c['num_palabras_comunes'] > 0 and c['proximidad'] == 1]

            # Nivel 3: Solo palabras clave (otras provincias)
            nivel_3 = [c for c in results if c['num_palabras_comunes'] > 0 and c['proximidad'] == 0]

            # Debug de niveles
            if provincia_origen:
                st.info(f"📊 **Niveles disponibles**: Nivel 1: {len(nivel_1)}, Nivel 2: {len(nivel_2)}, Nivel 3: {len(nivel_3)}")

            # ESTRATEGIA INTELIGENTE: Priorizar SIEMPRE misma provincia si existe
            if len(nivel_1) >= limit:
                # Ideal: Hay suficientes contratos recientes de la misma zona
                results_finales = nivel_1
                st.success(f"✅ **Nivel 1**: {len(nivel_1)} contratos (Palabras clave + Misma zona + Recientes)")
            elif len(nivel_2) >= limit:
                # Bueno: Hay suficientes contratos de la misma zona (aunque no sean recientes)
                results_finales = nivel_2
                st.info(f"ℹ️ **Nivel 2**: {len(nivel_2)} contratos (Palabras clave + Misma zona)")
            elif len(nivel_2) > 0:
                # Hay algunos contratos de la misma zona pero no suficientes
                # PRIORIZAR: Mostrar primero los de la misma zona, luego completar con otros
                results_finales = nivel_2 + nivel_3
                st.warning(f"⚠️ **Nivel mixto**: {len(nivel_2)} contratos de misma zona + {len(nivel_3)} de otras zonas")
                st.info(f"💡 **Se priorizan los {len(nivel_2)} contratos de la misma provincia**")
            else:
                # No hay ningún contrato de la misma zona
                results_finales = nivel_3
                if provincia_origen:
                    st.warning(f"⚠️ **Nivel 3**: {len(nivel_3)} contratos (No se encontraron en {provincia_origen})")
                else:
                    st.warning(f"⚠️ **Nivel 3**: {len(nivel_3)} contratos (Solo palabras clave)")

            # ORDENAR: PRIMERO por proximidad (misma provincia primero), LUEGO por palabras comunes, LUEGO por fecha
            results_finales.sort(key=lambda x: (
                x.get('proximidad', 0),  # 1 = misma provincia, 0 = otra provincia
                x['num_palabras_comunes'],
                x['fecha_publicacion'] if x['fecha_publicacion'] else datetime(1900, 1, 1)
            ), reverse=True)

            # Debug: Mostrar los primeros 3 contratos antes de enviar
            if provincia_origen and len(results_finales) >= 3:
                st.info("🔍 **Debug - Primeros 3 contratos después de ordenar:**")
                for idx, c in enumerate(results_finales[:3], 1):
                    prov = c.get('provincia', 'N/A')
                    prox = c.get('proximidad', 0)
                    palabras = c.get('num_palabras_comunes', 0)
                    st.text(f"  {idx}. Provincia: {prov} | Proximidad: {prox} | Palabras: {palabras}")

            st.success(f"✅ **Mostrando los {min(limit, len(results_finales))} contratos más relevantes**")

            results = results_finales

            # Mostrar los primeros 10
            # Contar cuántos son de la misma provincia en los 10 primeros
            contratos_mostrar = results[:10]
            num_misma_provincia = sum(1 for c in contratos_mostrar if c.get('proximidad', 0) == 1)
            num_otras_provincias = len(contratos_mostrar) - num_misma_provincia

            if provincia_origen and num_misma_provincia > 0:
                st.write(f"**Contratos encontrados:** {num_misma_provincia} de {provincia_origen} (📍), {num_otras_provincias} de otras provincias (📌)")
            else:
                st.write("**Contratos encontrados (ordenados por relevancia):**")

            for i, c in enumerate(contratos_mostrar, 1):
                fecha_str = str(c['fecha_publicacion'])[:10] if c['fecha_publicacion'] else 'N/A'
                # Usar las palabras comunes ya calculadas
                palabras_comunes = c.get('palabras_comunes', set())
                num_coincidencias = c.get('num_palabras_comunes', 0)
                provincia_str = c.get('provincia', 'N/A')
                proximidad_icon = "📍" if c.get('proximidad', 0) == 1 else "📌"

                st.write(f"{i}. [{num_coincidencias} palabra{'s' if num_coincidencias != 1 else ''} coincidente{'s' if num_coincidencias != 1 else ''}] {proximidad_icon} [{provincia_str}] [{fecha_str}] {c['titulo'][:60]}")
                st.write(f"   💡 Palabras clave coincidentes: {', '.join(sorted(palabras_comunes))}")

        else:
            # Sin título, solo ordenar por fecha
            results.sort(key=lambda x: x['fecha_publicacion'] if x['fecha_publicacion'] else datetime(1900, 1, 1), reverse=True)

        return results[:limit]

    except Exception as e:
        st.error(f"❌ Error en búsqueda: {e}")
        st.code(traceback.format_exc())
        return []
    finally:
        if conn:
            conn.close()

def generar_texto_informe(lote, contratos, baja_prom, baja_min, baja_max, empresas, num_lic_prom, datos):
    """Generar texto del informe para copiar siguiendo la estructura estándar"""

    # Variaciones para la introducción
    saludos = ["Buenos días,", "Buenas tardes,", "Estimados,"]
    intros_criterios = [
        "En la selección de expedientes, nos encontramos los siguientes criterios de adjudicación:",
        "En el análisis del expediente, identificamos los siguientes criterios de adjudicación:",
        "Para este proceso, se establecen los siguientes criterios de adjudicación:"
    ]

    # Variaciones para análisis de participación
    intros_participacion = [
        f"Al revisar expedientes previos de similar envergadura y presupuesto, hemos observado una participación promedio de {int(num_lic_prom)} empresa{'s' if int(num_lic_prom) != 1 else ''}.",
        f"Tras analizar licitaciones similares en cuanto a presupuesto y alcance, detectamos una concurrencia media de {int(num_lic_prom)} empresa{'s' if int(num_lic_prom) != 1 else ''}.",
        f"En expedientes comparables en presupuesto y características, observamos una participación promedio de {int(num_lic_prom)} empresa{'s' if int(num_lic_prom) != 1 else ''}."
    ]

    # Variaciones para empresas destacadas
    intros_empresas = [
        "Entre las empresas más sobresalientes en este campo están",
        "Las empresas con mayor actividad en este sector incluyen a",
        "Destacan en este ámbito empresas como"
    ]

    # Variaciones para análisis de ofertas
    analisis_ofertas = [
        f"Notamos que las variaciones en las ofertas son notables, con un promedio de entre {baja_min:.1f}% y {baja_max:.1f}%, lo que demuestra una estrategia de ofertas variada.",
        f"Observamos diferencias significativas en las propuestas económicas, oscilando entre {baja_min:.1f}% y {baja_max:.1f}%, evidenciando estrategias de competencia diversas.",
        f"Las ofertas presentadas muestran variabilidad considerable, situándose entre {baja_min:.1f}% y {baja_max:.1f}%, reflejando distintos enfoques competitivos."
    ]

    # Variaciones para recomendación
    recomendaciones = [
        f"Por ello, sugerimos una propuesta económica con un margen de descuento del {baja_prom:.1f}%.",
        f"En consecuencia, recomendamos plantear una oferta con un descuento aproximado del {baja_prom:.1f}%.",
        f"Considerando lo anterior, aconsejamos una baja cercana al {baja_prom:.1f}%."
    ]

    despedidas = ["Un cordial saludo", "Saludos cordiales", "Atentamente"]

    # Generar el texto siguiendo la estructura del ejemplo
    texto = f"{random.choice(saludos)}\n"
    texto += f"{random.choice(intros_criterios)}\n"

    # Criterios de adjudicación
    if lote['criterios']:
        for i, crit in enumerate(lote['criterios'], 1):
            # Manejar tanto strings como diccionarios
            if isinstance(crit, dict):
                desc = crit.get('descripcion', f'Criterio {i}')
                peso = crit.get('peso', '')
                if peso:
                    # Limpiar el peso (quitar % si existe, etc.)
                    peso_limpio = peso.strip().replace('%', '')
                    texto += f"{desc.upper()}: {peso_limpio} puntos\n"
                else:
                    texto += f"{desc.upper()}\n"
            else:
                # Es un string (formato JSON)
                texto += f"{str(crit).upper()}\n"
    else:
        texto += "OFERTA ECONÓMICA: 100 puntos\n"

    # Análisis de participación
    texto += f"{random.choice(intros_participacion)}\n"

    # Empresas destacadas (priorizando las de la misma provincia)
    if empresas:
        # Obtener provincia de origen (normalizada)
        provincia_origen = ''
        if datos:
            provincia_origen = (datos.get('provincia') or datos.get('ubicacion') or '').strip().lower()

        # Ordenar empresas: primero por provincia, luego por frecuencia
        def ordenar_empresas(item):
            nombre, info = item
            # Si info es un dict (nuevo formato con provincia)
            if isinstance(info, dict):
                es_misma_provincia = 1 if info['provincia'] and info['provincia'] == provincia_origen else 0
                frecuencia = info['frecuencia']
            else:
                # Formato antiguo (solo frecuencia)
                es_misma_provincia = 0
                frecuencia = info
            # Devolver tupla: primero prioridad provincia (1=misma, 0=otra), luego frecuencia
            return (es_misma_provincia, frecuencia)

        sorted_emp = sorted(empresas.items(), key=ordenar_empresas, reverse=True)[:5]
        empresas_texto = ", ".join([emp for emp, _ in sorted_emp[:-1]])
        if len(sorted_emp) > 1:
            empresas_texto += f" y {sorted_emp[-1][0]}"
        else:
            empresas_texto = sorted_emp[0][0]
        texto += f" {random.choice(intros_empresas)} {empresas_texto}.\n"

    # Análisis de ofertas
    texto += f"{random.choice(analisis_ofertas)}\n"

    # Recomendación
    texto += f"{random.choice(recomendaciones)}\n"

    # Despedida
    texto += f"{random.choice(despedidas)}\n"

    return texto

def crear_excel(datos_lote, contratos, baja_recomendada):
    """Crear archivo Excel con los resultados"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Título
    ws['A1'] = "ANÁLISIS DE BAJA ESTADÍSTICA"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    # Datos del contrato
    row = 3
    ws[f'A{row}'] = "Presupuesto"
    ws[f'B{row}'] = f"€{datos_lote['presupuesto']:,.2f}"
    row += 1
    ws[f'A{row}'] = "CPV"
    ws[f'B{row}'] = ', '.join(datos_lote['cpv']) if datos_lote['cpv'] else 'N/A'
    row += 2

    # Baja recomendada
    ws[f'A{row}'] = "BAJA RECOMENDADA"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'] = f"{baja_recomendada:.2f}%"
    ws[f'B{row}'].font = Font(bold=True, size=14)
    row += 2

    # Contratos similares
    ws[f'A{row}'] = "CONTRATOS SIMILARES"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:J{row}')
    row += 1

    # Cabeceras
    headers = ['Título', 'Organismo', 'Provincia', 'Presupuesto', 'Adjudicación', 'Baja %', 'Empresa', 'Licitadores', 'Fecha', 'CPV']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True)
    row += 1

    # Datos
    for contrato in contratos:
        ws.cell(row, 1, contrato['titulo'])  # Título completo
        ws.cell(row, 2, contrato['organismo'])  # Organismo completo
        ws.cell(row, 3, contrato.get('provincia', 'N/A'))  # Provincia
        ws.cell(row, 4, contrato['importe_total'])
        ws.cell(row, 5, contrato['importe_adjudicacion'])
        ws.cell(row, 6, contrato['baja'])
        ws.cell(row, 7, contrato['empresa'])  # Empresa completa
        ws.cell(row, 8, contrato.get('numero_licitadores', 'N/A'))  # Licitadores
        fecha_pub = str(contrato['fecha_publicacion'])[:10] if contrato.get('fecha_publicacion') else 'N/A'
        ws.cell(row, 9, fecha_pub)  # Fecha
        ws.cell(row, 10, str(contrato.get('cpv', 'N/A')))  # CPV
        row += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 60  # Título más ancho
    ws.column_dimensions['B'].width = 40  # Organismo
    ws.column_dimensions['C'].width = 15  # Provincia
    ws.column_dimensions['D'].width = 15  # Presupuesto
    ws.column_dimensions['E'].width = 15  # Adjudicación
    ws.column_dimensions['F'].width = 10  # Baja %
    ws.column_dimensions['G'].width = 40  # Empresa
    ws.column_dimensions['H'].width = 12  # Licitadores
    ws.column_dimensions['I'].width = 12  # Fecha
    ws.column_dimensions['J'].width = 20  # CPV

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# Sistema de autenticación
def check_login():
    """Verificar si el usuario está autenticado"""
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        st.title("🔐 Acceso a Análisis de Bajas Estadísticas")
        st.markdown("---")
        st.markdown("### Introduce tus credenciales")

        with st.form("login_form"):
            email = st.text_input("Email", placeholder="usuario@empresa.com")
            password = st.text_input("Contraseña", type="password")
            submit = st.form_submit_button("Iniciar Sesión")

            if submit:
                # Obtener credenciales de secrets
                valid_email = st.secrets.get("auth", {}).get("email", "")
                valid_password = st.secrets.get("auth", {}).get("password", "")

                if email == valid_email and password == valid_password:
                    st.session_state.authenticated = True
                    st.success("✅ Acceso concedido")
                    st.rerun()
                else:
                    st.error("❌ Email o contraseña incorrectos")

        st.stop()

# Verificar autenticación antes de mostrar la app
check_login()

# Interfaz principal
st.title("📊 Análisis de Bajas Estadísticas")
st.markdown("---")

# Selector de tipo de fuente
source_type = st.radio(
    "Selecciona el tipo de fuente:",
    options=["XML (URL)", "JSON (Archivo)", "Manual"],
    index=0,
    help="Elige si quieres analizar desde una URL de XML, subir un archivo JSON o introducir los datos manualmente"
)

xml_url = None
json_file = None
datos_manuales = None

if source_type == "XML (URL)":
    # Input para URL del XML
    xml_url = st.text_input(
        "Introduce la URL del XML del contrato:",
        placeholder="https://contrataciondelestado.es/FileSystem/servlet/...",
        help="Pega la URL completa del XML"
    )
elif source_type == "JSON (Archivo)":
    # File uploader para JSON
    json_file = st.file_uploader(
        "Sube el archivo JSON de la licitación:",
        type=['json'],
        help="Selecciona un archivo JSON que contenga los datos de la licitación"
    )
else:  # Manual
    st.markdown("### 📝 Introduce los datos del contrato")

    with st.form("formulario_manual"):
        st.markdown("#### Datos Generales del Contrato")

        col1, col2 = st.columns(2)
        with col1:
            titulo_contrato = st.text_input(
                "Título del contrato *",
                placeholder="Ej: Suministro de material de oficina",
                help="Título completo del contrato"
            )
            organismo = st.text_input(
                "Organismo contratante *",
                placeholder="Ej: Ayuntamiento de Madrid",
                help="Nombre del organismo que realiza la contratación"
            )

        with col2:
            ubicacion = st.text_input(
                "Ubicación",
                placeholder="Ej: Madrid",
                help="Ciudad o provincia de ejecución"
            )
            provincia = st.text_input(
                "Provincia",
                placeholder="Ej: Madrid",
                help="Provincia para búsqueda de contratos cercanos"
            )

        st.markdown("---")
        st.markdown("#### Datos del Lote")

        titulo_lote = st.text_input(
            "Título del lote",
            placeholder="Ej: Material de oficina (dejar vacío si coincide con el título del contrato)",
            help="Si el contrato no tiene lotes, dejar vacío"
        )

        col1, col2 = st.columns(2)
        with col1:
            presupuesto = st.number_input(
                "Presupuesto (€) *",
                min_value=0.0,
                step=1000.0,
                format="%.2f",
                help="Presupuesto base de licitación"
            )

        with col2:
            cpv_input = st.text_input(
                "Código CPV *",
                placeholder="Ej: 30190000 o 30190000, 30191000",
                help="Código CPV (puedes introducir varios separados por comas)"
            )

        st.markdown("#### Criterios de Adjudicación")
        st.markdown("Introduce los criterios de adjudicación (uno por línea, formato: *Descripción: Peso puntos*)")

        criterios_input = st.text_area(
            "Criterios",
            placeholder="Ej:\nOferta económica: 60 puntos\nMejoras técnicas: 30 puntos\nPlazo de entrega: 10 puntos",
            height=150,
            help="Un criterio por línea. Formato: Descripción: Peso puntos"
        )

        submit_manual = st.form_submit_button("✅ Validar Datos", type="primary")

        if submit_manual:
            # Validar campos obligatorios
            errores = []
            if not titulo_contrato:
                errores.append("El título del contrato es obligatorio")
            if not organismo:
                errores.append("El organismo contratante es obligatorio")
            if presupuesto <= 0:
                errores.append("El presupuesto debe ser mayor que 0")
            if not cpv_input:
                errores.append("El código CPV es obligatorio")

            if errores:
                for error in errores:
                    st.error(f"❌ {error}")
            else:
                # Procesar CPVs
                cpvs = [cpv.strip() for cpv in cpv_input.split(',') if cpv.strip()]

                # Procesar criterios
                criterios = []
                if criterios_input:
                    for linea in criterios_input.strip().split('\n'):
                        if linea.strip():
                            criterios.append(linea.strip())

                # Crear estructura de datos compatible
                datos_manuales = {
                    'titulo': titulo_contrato,
                    'organismo': organismo,
                    'ubicacion': ubicacion or 'No especificado',
                    'provincia': provincia or '',
                    'lotes': [{
                        'numero': '1',
                        'titulo': titulo_lote if titulo_lote else titulo_contrato,
                        'presupuesto': presupuesto,
                        'cpv': cpvs,
                        'criterios': criterios
                    }]
                }

                st.success("✅ Datos validados correctamente")
                st.session_state.datos_manuales = datos_manuales

# Campo de palabras clave manual (disponible para todas las opciones)
st.markdown("---")
st.markdown("### 🔑 Palabras Clave para Búsqueda (Opcional)")
palabras_clave_manual = st.text_input(
    "Introduce palabras clave específicas para buscar contratos similares",
    placeholder="Ej: limpieza, jardineria, edificios",
    help="Separadas por comas. Si lo dejas vacío, se extraerán automáticamente del título del contrato"
)

if st.button("🚀 Analizar Contrato", type="primary"):
    if source_type == "XML (URL)" and not xml_url:
        st.warning("Por favor, introduce una URL")
    elif source_type == "JSON (Archivo)" and not json_file:
        st.warning("Por favor, sube un archivo JSON")
    elif source_type == "Manual" and not st.session_state.get('datos_manuales'):
        st.warning("Por favor, completa y valida el formulario primero")
    else:
        datos = None

        if source_type == "XML (URL)":
            with st.spinner("Procesando XML..."):
                datos = extraer_datos_xml_completo(xml_url)
        elif source_type == "JSON (Archivo)":
            with st.spinner("Procesando JSON..."):
                try:
                    # Leer el archivo JSON
                    json_content = json_file.read().decode('utf-8')
                    datos = extraer_datos_json_completo(json_content)
                except Exception as e:
                    st.error(f"Error leyendo archivo JSON: {e}")
                    datos = None
        else:  # Manual
            datos = st.session_state.get('datos_manuales')

        if not datos or not datos['lotes']:
            if source_type == "Manual":
                st.error(f"Error al procesar los datos manuales")
            else:
                source_name = "XML" if source_type == "XML (URL)" else "JSON"
                st.error(f"No se pudieron extraer lotes del {source_name}")
        else:
            if source_type == "Manual":
                st.success(f"✅ Datos manuales procesados - {len(datos['lotes'])} lote(s) encontrado(s)")
            else:
                source_name = "XML" if source_type == "XML (URL)" else "JSON"
                st.success(f"✅ {source_name} procesado - {len(datos['lotes'])} lote(s) encontrado(s)")

            # Mostrar datos extraídos
            with st.expander("📋 Datos extraídos del documento"):
                st.write(f"**Título:** {datos.get('titulo', 'No detectado')}")
                st.write(f"**Organismo:** {datos.get('organismo', 'No detectado')}")
                st.write(f"**Ubicación:** {datos.get('ubicacion', 'No detectado')}")
                if datos.get('provincia'):
                    st.write(f"**📍 Provincia:** {datos.get('provincia')}")
                else:
                    st.write(f"**📍 Provincia:** No detectada (no se aplicará filtro geográfico)")

            # Analizar cada lote
            for lote in datos['lotes']:
                st.markdown("---")
                st.markdown(f"## 📦 Lote {lote['numero']}: {lote['titulo'][:80] if lote['titulo'] else 'Sin título'}")

                st.markdown(f"**Presupuesto:** €{lote['presupuesto']:,.2f}")
                st.markdown(f"**CPV:** {', '.join(lote['cpv']) if lote['cpv'] else 'No especificado'}")

                # Mostrar palabras clave (manuales o automáticas)
                if palabras_clave_manual:
                    palabras_mostrar = set([p.strip().lower() for p in palabras_clave_manual.split(',') if p.strip()])
                    st.markdown(f"**🔑 Palabras clave (manuales):** {', '.join(sorted(palabras_mostrar))}")
                elif lote['titulo']:
                    palabras_clave = extraer_palabras_clave(lote['titulo'])
                    if palabras_clave:
                        st.markdown(f"**🔑 Palabras clave (automáticas):** {', '.join(sorted(palabras_clave))}")

                # Criterios
                st.markdown("### ⚖️ Criterios de Adjudicación")
                if lote['criterios']:
                    for i, crit in enumerate(lote['criterios'], 1):
                        # Manejar tanto strings como diccionarios
                        if isinstance(crit, dict):
                            desc = crit.get('descripcion', f'Criterio {i}')
                            peso = crit.get('peso', '')
                            st.write(f"**{i}.** {desc}: **{peso}**" if peso else f"**{i}.** {desc}")
                        else:
                            # Es un string
                            st.write(f"**{i}.** {crit}")
                else:
                    st.info("ℹ️ No se encontraron criterios de adjudicación")

                # Buscar contratos
                if lote['cpv'] and lote['presupuesto'] > 0:
                    st.markdown("### 🔍 Búsqueda de Contratos Similares")

                    pres_min = lote['presupuesto'] * 0.5
                    pres_max = lote['presupuesto'] * 1.5

                    # Mostrar provincia que se usará para la búsqueda
                    provincia_busqueda = datos.get('provincia')
                    if provincia_busqueda:
                        st.info(f"🌍 **Buscando contratos con filtro geográfico**: {provincia_busqueda}")
                    else:
                        st.info(f"🌍 **Buscando contratos sin filtro geográfico** (provincia no detectada en el documento)")

                    # Búsqueda normal
                    with st.spinner("Buscando contratos..."):
                        contratos = buscar_contratos(
                            lote['cpv'],
                            pres_min,
                            pres_max,
                            titulo_referencia=lote['titulo'],
                            limit=10,
                            ampliada=False,
                            provincia_origen=provincia_busqueda,
                            palabras_clave_manual=palabras_clave_manual if palabras_clave_manual else None
                        )

                    # Si hay menos de 3 contratos, hacer búsqueda ampliada
                    if len(contratos) < 3:
                        st.warning(f"⚠️ Solo se encontraron {len(contratos)} contrato(s). Ampliando búsqueda...")
                        if provincia_busqueda:
                            st.info(f"🔄 **Ampliando CPV (2 dígitos) y presupuesto (±100%), manteniendo palabra clave + provincia**")
                        else:
                            st.info(f"🔄 **Ampliando CPV (2 dígitos) y presupuesto (±100%), manteniendo palabra clave**")
                        with st.spinner("Buscando con criterios ampliados..."):
                            contratos = buscar_contratos(
                                lote['cpv'],
                                pres_min,
                                pres_max,
                                titulo_referencia=lote['titulo'],
                                limit=10,
                                ampliada=True,
                                provincia_origen=provincia_busqueda,
                                palabras_clave_manual=palabras_clave_manual if palabras_clave_manual else None
                            )

                        if len(contratos) < 3:
                            st.error(f"❌ Solo se encontraron {len(contratos)} contrato(s) incluso con búsqueda ampliada")

                    if contratos:
                        # Calcular estadísticas
                        bajas = [c['baja'] for c in contratos if c['baja']]

                        if bajas:
                            baja_min = min(bajas)
                            baja_max = max(bajas)

                            # Usar nuevo algoritmo de cálculo
                            baja_prom = calcular_baja_recomendada(bajas)

                            num_lic_prom = sum([c['numero_licitadores'] or 0 for c in contratos]) / len(contratos)

                            st.markdown("### 📊 Resultados")

                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("🎯 Baja Recomendada", f"{baja_prom:.2f}%")
                            with col2:
                                st.metric("📈 Contratos Analizados", len(contratos))
                            with col3:
                                st.metric("👥 Licitadores Promedio", f"{num_lic_prom:.0f}")

                            st.markdown(f"**Rango de bajas:** {baja_min:.1f}% - {baja_max:.1f}%")

                            # Generar diccionario de empresas con información de provincia
                            empresas_data = {}
                            for c in contratos:
                                emp = c['empresa']
                                if emp and emp != 'N/A' and len(emp) > 3:
                                    if emp not in empresas_data:
                                        empresas_data[emp] = {
                                            'frecuencia': 0,
                                            'provincia': c.get('provincia', '').strip().lower() if c.get('provincia') else ''
                                        }
                                    empresas_data[emp]['frecuencia'] += 1

                            # Generar texto del informe
                            texto_informe = generar_texto_informe(lote, contratos, baja_prom, baja_min, baja_max, empresas_data, num_lic_prom, datos)

                            # Sección de descarga y texto
                            st.markdown("---")
                            st.markdown("### 📝 Informe Generado")

                            col1, col2 = st.columns([1, 1])
                            with col1:
                                excel_data = crear_excel(lote, contratos, baja_prom)
                                st.download_button(
                                    label="📥 Descargar análisis en Excel",
                                    data=excel_data,
                                    file_name=f"analisis_lote_{lote['numero']}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                            with col2:
                                st.info(f"✅ {len(contratos)} contratos incluidos en el análisis")

                            # Texto para copiar
                            st.markdown("#### 📄 Texto del Informe (Copia y Pega)")
                            st.text_area(
                                label="Texto completo del análisis:",
                                value=texto_informe,
                                height=300,
                                help="Copia este texto para usar en tu informe"
                            )

                            # Mostrar contratos
                            with st.expander(f"📋 Ver los {len(contratos)} contratos encontrados"):
                                for i, c in enumerate(contratos, 1):
                                    st.markdown(f"### {i}. {c['titulo']}")

                                    col1, col2 = st.columns(2)
                                    with col1:
                                        st.write(f"**📍 Organismo:** {c['organismo']}")
                                        st.write(f"**🏢 Adjudicatario:** {c['empresa']}")
                                        st.write(f"**📍 Provincia:** {c.get('provincia', 'N/A')}")
                                        st.write(f"**🔢 CPV:** {c.get('cpv', 'N/A')}")
                                    with col2:
                                        st.write(f"**💰 Presupuesto:** €{c['importe_total']:,.2f}")
                                        st.write(f"**💵 Adjudicación:** €{c['importe_adjudicacion']:,.2f}")
                                        st.write(f"**📉 Baja:** {c['baja']:.2f}%")
                                        fecha = str(c['fecha_publicacion'])[:10] if c['fecha_publicacion'] else 'N/A'
                                        st.write(f"**📅 Fecha:** {fecha}")
                                        num_lic = c.get('numero_licitadores', 'N/A')
                                        st.write(f"**👥 Licitadores:** {num_lic if num_lic else 'N/A'}")

                                    st.divider()
                else:
                    st.warning("⚠️ No se pudo extraer CPV o presupuesto del lote")

st.markdown("---")
st.caption("📊 Análisis basado en datos del Portal de Contratación del Estado")
